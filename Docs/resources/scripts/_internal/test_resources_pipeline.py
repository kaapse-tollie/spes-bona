from __future__ import annotations

import csv
import importlib.util
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path("/Users/depro/Documents/Paradox Interactive/Victoria 3/mod")
REPO = ROOT / "Spes Bona - A Southern Africa Flavour Pack"
PUBLIC_ROOT = REPO / "Docs/resources"
AUDIT_DIR = PUBLIC_ROOT / "audit"
DERIVED_DIR = PUBLIC_ROOT / "data/derived"
BUILDER = PUBLIC_ROOT / "scripts/_internal/build_resources_workbook.py"
CLI = PUBLIC_ROOT / "scripts/resources.py"
WORKBOOK = PUBLIC_ROOT / "RESOURCES.xlsx"
README = PUBLIC_ROOT / "README.md"
REPORT = AUDIT_DIR / "test_report.md"

ADJUSTMENTS = AUDIT_DIR / "adjustments.csv"
PRIORITY_ROWS = AUDIT_DIR / "priority_rows.csv"
ROW_AUDIT = AUDIT_DIR / "row_audit.csv"
FINAL_RESOURCE_CAPS = DERIVED_DIR / "final_resource_caps.csv"
REGIONAL_TOTALS = DERIVED_DIR / "regional_resource_totals.csv"
ARABLE_STATE_MEANS = DERIVED_DIR / "arable_state_means.csv"
ARABLE_SHARED_DENOMINATOR = DERIVED_DIR / "arable_shared_denominator.csv"
ARABLE_COMPARATOR_DIAGNOSTICS = DERIVED_DIR / "arable_comparator_diagnostics.csv"
ARABLE_TARGET_CAPACITY = DERIVED_DIR / "arable_target_capacity_rows.csv"
ARABLE_COMPARATOR_CAPACITY = DERIVED_DIR / "arable_comparator_capacity_rows.csv"
ARABLE_RESOURCE_EXPECTATIONS = DERIVED_DIR / "arable_resource_expectations.csv"
WOOD_TARGET_CAPACITY = DERIVED_DIR / "wood_target_capacity_rows.csv"
WOOD_COMPARATOR_CAPACITY = DERIVED_DIR / "wood_comparator_capacity_rows.csv"
RUBBER_TARGET_CAPACITY = DERIVED_DIR / "rubber_target_capacity_rows.csv"
RUBBER_COMPARATOR_CAPACITY = DERIVED_DIR / "rubber_comparator_capacity_rows.csv"
RESOURCE_ADJUSTMENTS = DERIVED_DIR / "resource_adjustments.csv"
RESOURCE_DENOMINATORS = DERIVED_DIR / "resource_denominators.csv"
TARGET_OBSERVATIONS = DERIVED_DIR / "target_observations.csv"
GDP_SELECTION = DERIVED_DIR / "gdp_selection_rows.csv"
STATE_DELTA_SUMMARY = DERIVED_DIR / "state_delta_summary.csv"
STATE_RESOURCE_DELTAS = DERIVED_DIR / "state_resource_deltas.csv"
STATE_DELTA_REPORT = AUDIT_DIR / "sb_state_delta_report.md"
TARGET_DATA_VALIDATION = AUDIT_DIR / "target_data_validation.csv"
STATE_REGIONAL_ADVANTAGES = AUDIT_DIR / "state_regional_advantages.csv"
STATE_COUNTERFACTUAL_AUDIT = AUDIT_DIR / "state_resource_counterfactual_audit.csv"
STATE_PASS_TRACKER = AUDIT_DIR / "state_pass_tracker.csv"
FAMILY_REWRITE_LOG = AUDIT_DIR / "family_rewrite_log.csv"
STATE_REVIEW_STATUS = AUDIT_DIR / "state_review_status.csv"
DATA_README = PUBLIC_ROOT / "data/README.md"
RAW_DATA_README = PUBLIC_ROOT / "data/raw/README.md"
DERIVED_DATA_README = PUBLIC_ROOT / "data/derived/README.md"
NON_ARABLE_BENCHMARKS = PUBLIC_ROOT / "data/raw/non_arable_benchmark_cases.csv"
WOOD_WEIGHTS = PUBLIC_ROOT / "data/raw/wood_land_class_weights.csv"
RAW_WOOD_TARGET_CAPACITY = PUBLIC_ROOT / "data/raw/wood_target_capacity_rows.csv"
RAW_WOOD_COMPARATOR_CAPACITY = PUBLIC_ROOT / "data/raw/wood_comparator_capacity_rows.csv"
RUBBER_WEIGHTS = PUBLIC_ROOT / "data/raw/rubber_land_class_weights.csv"
RAW_RUBBER_TARGET_CAPACITY = PUBLIC_ROOT / "data/raw/rubber_target_capacity_rows.csv"
RAW_RUBBER_COMPARATOR_CAPACITY = PUBLIC_ROOT / "data/raw/rubber_comparator_capacity_rows.csv"
ARABLE_WEIGHTS = PUBLIC_ROOT / "data/raw/arable_land_class_weights.csv"
ARABLE_BASKETS = PUBLIC_ROOT / "data/raw/arable_baskets.csv"
RAW_ARABLE_TARGET_CAPACITY = PUBLIC_ROOT / "data/raw/arable_target_capacity_rows.csv"
RAW_ARABLE_COMPARATOR_CAPACITY = PUBLIC_ROOT / "data/raw/arable_comparator_capacity_rows.csv"
RAW_HISTORICAL = PUBLIC_ROOT / "data/raw/historical_anchors.csv"
RAW_MODERN = PUBLIC_ROOT / "data/raw/modern_maxima.csv"
RAW_ADJUSTMENT_INPUTS = PUBLIC_ROOT / "data/raw/adjustment_inputs.csv"
RAW_COUNTEREVIDENCE = PUBLIC_ROOT / "data/raw/counterevidence_cases.csv"


@dataclass
class CheckResult:
    name: str
    status: str
    detail: str


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def load_builder():
    spec = importlib.util.spec_from_file_location("resources_builder", BUILDER)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def status_line(result: CheckResult) -> str:
    return f"- **{result.status}** `{result.name}`: {result.detail}"


def find_row_by_first_cell(ws, value: str) -> int | None:
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == value:
            return row_idx
    return None


def extract_table(ws, section_title: str) -> tuple[list[str], list[dict[str, Any]]]:
    title_row = find_row_by_first_cell(ws, section_title)
    if title_row is None:
        return [], []
    header_row = title_row + 1
    headers: list[str] = []
    col = 1
    while True:
        value = ws.cell(row=header_row, column=col).value
        if value in (None, ""):
            break
        headers.append(str(value))
        col += 1
    rows: list[dict[str, Any]] = []
    row_idx = header_row + 1
    while True:
        first = ws.cell(row=row_idx, column=1).value
        second = ws.cell(row=row_idx, column=2).value
        if first in (None, "") and second in (None, ""):
            break
        rows.append({headers[col_idx - 1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)})
        row_idx += 1
    return headers, rows


def duplicate_active_keys(rows: list[dict[str, str]], logical_key_fields: list[str]) -> list[str]:
    duplicates: list[str] = []
    seen: set[tuple[str, ...]] = set()
    for row in rows:
        if row.get("row_status") not in ("", "active"):
            continue
        key = tuple(str(row.get(field, "")) for field in logical_key_fields)
        if key in seen:
            duplicates.append(" / ".join(key))
        seen.add(key)
    return duplicates


def run_tests() -> str:
    builder = load_builder()
    readme_text = README.read_text(encoding="utf-8")
    data_readme_text = DATA_README.read_text(encoding="utf-8") if DATA_README.exists() else ""
    raw_readme_text = RAW_DATA_README.read_text(encoding="utf-8") if RAW_DATA_README.exists() else ""
    derived_readme_text = DERIVED_DATA_README.read_text(encoding="utf-8") if DERIVED_DATA_README.exists() else ""
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=False)
    overview_ws = workbook["Overview"]
    state_sheet_names = [info["official_name"] for info in builder.STATE_INFO]
    state_sheet_map = {name: workbook[name] for name in state_sheet_names}

    final_caps = read_csv_rows(FINAL_RESOURCE_CAPS)
    regional_totals = read_csv_rows(REGIONAL_TOTALS)
    state_delta_summary = read_csv_rows(STATE_DELTA_SUMMARY)
    state_resource_deltas = read_csv_rows(STATE_RESOURCE_DELTAS)
    adjustments = read_csv_rows(ADJUSTMENTS)
    priority_rows = read_csv_rows(PRIORITY_ROWS)
    row_audit = read_csv_rows(ROW_AUDIT)
    arable_state_means = read_csv_rows(ARABLE_STATE_MEANS)
    arable_shared = read_csv_rows(ARABLE_SHARED_DENOMINATOR)
    arable_comparator_diagnostics = read_csv_rows(ARABLE_COMPARATOR_DIAGNOSTICS)
    arable_target_capacity_rows = read_csv_rows(ARABLE_TARGET_CAPACITY)
    arable_comparator_capacity_rows = read_csv_rows(ARABLE_COMPARATOR_CAPACITY)
    arable_resource_expectations = read_csv_rows(ARABLE_RESOURCE_EXPECTATIONS)
    wood_target_capacity_rows = read_csv_rows(WOOD_TARGET_CAPACITY)
    wood_comparator_capacity_rows = read_csv_rows(WOOD_COMPARATOR_CAPACITY)
    rubber_target_capacity_rows = read_csv_rows(RUBBER_TARGET_CAPACITY)
    rubber_comparator_capacity_rows = read_csv_rows(RUBBER_COMPARATOR_CAPACITY)
    resource_adjustments = read_csv_rows(RESOURCE_ADJUSTMENTS)
    resource_denominators = read_csv_rows(RESOURCE_DENOMINATORS)
    target_observations = read_csv_rows(TARGET_OBSERVATIONS)
    gdp_selection_rows = read_csv_rows(GDP_SELECTION)
    target_data_validation = read_csv_rows(TARGET_DATA_VALIDATION)
    state_regional_advantages = read_csv_rows(STATE_REGIONAL_ADVANTAGES)
    state_counterfactual_audit = read_csv_rows(STATE_COUNTERFACTUAL_AUDIT)
    state_pass_tracker = read_csv_rows(STATE_PASS_TRACKER)
    family_rewrite_log = read_csv_rows(FAMILY_REWRITE_LOG)
    state_review_status = read_csv_rows(STATE_REVIEW_STATUS)
    non_arable_benchmarks = read_csv_rows(NON_ARABLE_BENCHMARKS) if NON_ARABLE_BENCHMARKS.exists() else []
    raw_historical_rows = read_csv_rows(RAW_HISTORICAL)
    raw_modern_rows = read_csv_rows(RAW_MODERN)
    raw_arable_target_rows = read_csv_rows(RAW_ARABLE_TARGET_CAPACITY)
    raw_wood_target_rows = read_csv_rows(RAW_WOOD_TARGET_CAPACITY)
    raw_rubber_target_rows = read_csv_rows(RAW_RUBBER_TARGET_CAPACITY)
    raw_adjustment_rows = read_csv_rows(RAW_ADJUSTMENT_INPUTS)
    raw_counterevidence_rows = read_csv_rows(RAW_COUNTEREVIDENCE)

    live_values = builder.parse_live_state_resources()

    results: list[CheckResult] = []

    results.append(CheckResult("public cli entrypoint exists", "PASS" if CLI.exists() else "FAIL", f"Expected {CLI}"))
    results.append(CheckResult("new arable raw files exist", "PASS" if ARABLE_WEIGHTS.exists() and RAW_ARABLE_TARGET_CAPACITY.exists() and RAW_ARABLE_COMPARATOR_CAPACITY.exists() else "FAIL", "Expected land-class, target-capacity, and comparator-capacity raw files."))
    results.append(CheckResult("new wood raw files exist", "PASS" if WOOD_WEIGHTS.exists() and RAW_WOOD_TARGET_CAPACITY.exists() and RAW_WOOD_COMPARATOR_CAPACITY.exists() else "FAIL", "Expected wood land-class, target-capacity, and comparator-capacity raw files."))
    results.append(CheckResult("new rubber raw files exist", "PASS" if RUBBER_WEIGHTS.exists() and RAW_RUBBER_TARGET_CAPACITY.exists() and RAW_RUBBER_COMPARATOR_CAPACITY.exists() else "FAIL", "Expected rubber land-class, target-capacity, and comparator-capacity raw files."))
    results.append(CheckResult("non-arable benchmark registry exists", "PASS" if NON_ARABLE_BENCHMARKS.exists() and len(non_arable_benchmarks) > 0 else "FAIL", "Expected non_arable_benchmark_cases.csv to be the authoritative non-land comparator registry."))

    readme_checks = [
        "## introduction",
        "## why",
        "## methodology",
        "## analysis",
        "## problems and what we did",
        "## how to modify / run",
        "## references",
        "append-only",
        "state-pass",
        "effective commercial agricultural hectares",
        "effective commercial forestry hectares",
        "latent rubber",
        "gbr 1940",
        "1940-equivalent",
        "peak override",
        "resources.xlsx",
        "non_arable_benchmark_cases.csv",
        "earliest commercial activity year",
        "representative gdp-equivalent year",
        "representative-year lag",
    ]
    missing_readme = [token for token in readme_checks if token not in readme_text.lower()]
    results.append(CheckResult("readme follows the explanatory paper-style structure", "PASS" if not missing_readme else "FAIL", "Missing README terms: " + ", ".join(missing_readme)))
    readme_drift_hits = [
        token
        for token in ["todo:", "current readme had become too thin", "±2 year band", "+/- 2"]
        if token in readme_text.lower()
    ]
    results.append(CheckResult("readme no longer carries v2 drift language or the old GDP anchor", "PASS" if not readme_drift_hits else "FAIL", "Unexpected README text: " + ", ".join(readme_drift_hits)))

    gdp_anchor_rows = read_csv_rows(PUBLIC_ROOT / "data/raw/gdp_reference_anchor.csv")
    gdp_anchor_failures = []
    if len(gdp_anchor_rows) != 1:
        gdp_anchor_failures.append(f"expected 1 anchor row, got {len(gdp_anchor_rows)}")
    else:
        anchor_row = gdp_anchor_rows[0]
        if anchor_row.get("reference_geography") != "GBR":
            gdp_anchor_failures.append("reference geography should remain GBR")
        if anchor_row.get("reference_year") != "1940":
            gdp_anchor_failures.append(f"reference year should be 1940, got {anchor_row.get('reference_year')}")
    results.append(CheckResult("frozen GDP anchor is GBR 1940", "PASS" if not gdp_anchor_failures else "FAIL", "; ".join(gdp_anchor_failures)))

    z_rule_failures = []
    if abs(builder.UNIVERSAL_Z_E_COEFFICIENT - 0.00275) > 1e-15:
        z_rule_failures.append("unexpected e coefficient")
    if abs(builder.UNIVERSAL_Z_PROXY_LAG_COEFFICIENT - 0.01) > 1e-15:
        z_rule_failures.append("unexpected proxy-lag coefficient")
    expected_keeps = [
        ((1930, 1960), 0.8000),
        ((1940, 1953), 0.8700),
        ((1949, 1960), 0.8330),
        ((1960, 1960), 0.8326),
        ((1970, 1990), 0.5167),
        ((2020, 2020), 0.0332),
    ]
    for (earliest_year, representative_year), expected_keep in expected_keeps:
        actual_penalty = builder.universal_z_penalty(earliest_year, representative_year)
        actual_keep = None if actual_penalty is None else (1.0 - actual_penalty)
        if actual_keep is None or abs(actual_keep - expected_keep) > 5e-4:
            z_rule_failures.append(
                f"{earliest_year}/{representative_year}: expected keep about {expected_keep}, got {actual_keep}"
            )
    if builder.universal_z_penalty(1930, 1935) != 0.0:
        z_rule_failures.append("1930/1935 should produce zero chronology penalty")
    results.append(CheckResult("universal quantity-resource Z formula matches the locked chronology calibration", "PASS" if not z_rule_failures else "FAIL", "; ".join(z_rule_failures[:10])))

    visible_sheets = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    expected_visible_sheets = ["Overview", *state_sheet_names]
    results.append(CheckResult("overview plus one sheet per SB state", "PASS" if visible_sheets == expected_visible_sheets else "FAIL", f"Visible sheets: {visible_sheets}"))

    required_overview_sections = ["SB totals before and after", "Major tag changes", "State pass progress"]
    missing_overview_sections = [section for section in required_overview_sections if find_row_by_first_cell(overview_ws, section) is None]
    results.append(CheckResult("overview sheet sections", "PASS" if not missing_overview_sections else "FAIL", "Missing sections: " + ", ".join(missing_overview_sections)))

    missing_state_sections = [
        state
        for state, ws in state_sheet_map.items()
        if find_row_by_first_cell(ws, "Totals") is None or find_row_by_first_cell(ws, "Resource rows") is None
    ]
    results.append(CheckResult("each state sheet has totals and resource table sections", "PASS" if not missing_state_sections else "FAIL", "Missing sections on: " + ", ".join(missing_state_sections)))

    formula_cells = []
    for ws in [overview_ws, *state_sheet_map.values()]:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"{ws.title}!{cell.coordinate}")
    results.append(CheckResult("visible sheets are generated output", "PASS" if not formula_cells else "FAIL", f"Formula cells on visible sheets: {len(formula_cells)}"))

    missing_docs = [str(path) for path in [DATA_README, RAW_DATA_README, DERIVED_DATA_README] if not path.exists()]
    results.append(CheckResult("data readmes exist", "PASS" if not missing_docs else "FAIL", "Missing docs: " + ", ".join(missing_docs)))
    data_readme_checks = [
        ("data/README classifies the package", "formula-driving" in data_readme_text.lower() and "generated output" in data_readme_text.lower()),
        ("raw data README classifies inputs", "formula-driving" in raw_readme_text.lower() and "supporting source material" in raw_readme_text.lower()),
        ("derived data README classifies outputs", "generated output" in derived_readme_text.lower() and "review surface" in derived_readme_text.lower()),
    ]
    for check_name, passed in data_readme_checks:
        results.append(CheckResult(check_name, "PASS" if passed else "FAIL", "README classification text missing."))

    ds_store_hits = [str(path) for path in PUBLIC_ROOT.rglob(".DS_Store")]
    results.append(CheckResult("no .DS_Store files remain under Docs/resources", "PASS" if not ds_store_hits else "FAIL", "; ".join(ds_store_hits[:10])))

    raw_file_headers = {
        "historical_anchors.csv": set(raw_historical_rows[0].keys()) if raw_historical_rows else set(),
        "modern_maxima.csv": set(raw_modern_rows[0].keys()) if raw_modern_rows else set(),
        "arable_target_capacity_rows.csv": set(raw_arable_target_rows[0].keys()) if raw_arable_target_rows else set(),
        "wood_target_capacity_rows.csv": set(raw_wood_target_rows[0].keys()) if raw_wood_target_rows else set(),
    }
    raw_validation_failures = [
        name
        for name, headers in raw_file_headers.items()
        if not set(builder.TARGET_VALIDATION_FIELDS) <= headers
    ]
    results.append(CheckResult("formula-driving raw files expose validation columns", "PASS" if not raw_validation_failures else "FAIL", "Missing validation columns in: " + ", ".join(raw_validation_failures)))

    results.append(CheckResult("target data validation output exists and is populated", "PASS" if TARGET_DATA_VALIDATION.exists() and len(target_data_validation) > 0 else "FAIL", f"Rows={len(target_data_validation)}"))

    validation_rule_failures = []
    for row in target_data_validation:
        scope = row["evidence_scope"]
        discount = float(row["localization_discount"] or 0)
        if scope == "state_localized" and abs(discount - 1.0) > 1e-9:
            validation_rule_failures.append(f"{row['state']} / {row['resource']} / {row['land_class']} expected 1.0")
        if scope in {"regional_proxy", "national_fallback"} and not (0 < discount < 1.0):
            validation_rule_failures.append(f"{row['state']} / {row['resource']} / {row['land_class']} expected <1.0")
        if row["slot_support_status"] == "broad_potential_only" and row["drives_x"] != "no":
            validation_rule_failures.append(f"{row['state']} / {row['resource']} / {row['land_class']} broad potential still drives x")
    results.append(CheckResult("validation discount and drive-x rules hold", "PASS" if not validation_rule_failures else "FAIL", "; ".join(validation_rule_failures[:12])))

    placeholder_hits = []
    for row in target_data_validation:
        text = " ".join([row.get("validation_note", ""), row.get("review_action", "")]).lower()
        if any(token in text for token in ["not yet", "unresolved", "placeholder"]):
            placeholder_hits.append(f"{row['state']} / {row['resource']} / {row['land_class']}")
    results.append(CheckResult("validation notes ship without placeholder language", "PASS" if not placeholder_hits else "FAIL", "; ".join(placeholder_hits[:10])))

    new_audit_tables_present = all(path.exists() for path in [STATE_REGIONAL_ADVANTAGES, STATE_COUNTERFACTUAL_AUDIT, STATE_PASS_TRACKER, FAMILY_REWRITE_LOG, STATE_REVIEW_STATUS])
    results.append(CheckResult("new audit tables exist", "PASS" if new_audit_tables_present else "FAIL", "Missing one of the tracker/counterfactual/rewrite/state review audit tables."))
    results.append(CheckResult("regional advantages and state review status are populated", "PASS" if len(state_regional_advantages) == len(builder.STATE_INFO) and len(state_review_status) == len(builder.STATE_INFO) else "FAIL", f"advantages={len(state_regional_advantages)}, review_status={len(state_review_status)}"))

    lifecycle_required = set(builder.LIFECYCLE_FIELDNAMES)
    lifecycle_failures = []
    raw_tables_to_check = [
        ("historical_anchors.csv", raw_historical_rows, builder.TARGET_OBSERVATION_LOGICAL_KEY_FIELDS),
        ("modern_maxima.csv", raw_modern_rows, builder.TARGET_OBSERVATION_LOGICAL_KEY_FIELDS),
        ("arable_target_capacity_rows.csv", raw_arable_target_rows, builder.ARABLE_TARGET_CAPACITY_LOGICAL_KEY_FIELDS),
        ("wood_target_capacity_rows.csv", raw_wood_target_rows, builder.WOOD_TARGET_CAPACITY_LOGICAL_KEY_FIELDS),
        ("rubber_target_capacity_rows.csv", raw_rubber_target_rows, builder.RUBBER_TARGET_CAPACITY_LOGICAL_KEY_FIELDS),
        ("adjustment_inputs.csv", raw_adjustment_rows, builder.ADJUSTMENT_INPUT_LOGICAL_KEY_FIELDS),
        ("counterevidence_cases.csv", raw_counterevidence_rows, builder.COUNTEREVIDENCE_LOGICAL_KEY_FIELDS),
    ]
    duplicate_key_failures = []
    for name, rows, logical_key_fields in raw_tables_to_check:
        headers = set(rows[0].keys()) if rows else set()
        if not lifecycle_required <= headers:
            lifecycle_failures.append(name)
        duplicate_key_failures.extend([f"{name}: {value}" for value in duplicate_active_keys(rows, logical_key_fields)])
    results.append(CheckResult("lifecycle columns exist on maintained evidence tables", "PASS" if not lifecycle_failures else "FAIL", "Missing lifecycle fields in: " + ", ".join(lifecycle_failures)))
    results.append(CheckResult("no maintained logical key has more than one active evidence row", "PASS" if not duplicate_key_failures else "FAIL", "; ".join(duplicate_key_failures[:10])))
    results.append(
        CheckResult(
            "adjustment schema carries the documented-working floor flag",
            "PASS" if "documented_working_floor_eligible" in builder.ADJUSTMENT_INPUT_FIELDNAMES else "FAIL",
            "Expected documented_working_floor_eligible in adjustment input schema.",
        )
    )

    chronology_pair_failures = []
    for row in raw_adjustment_rows:
        if row.get("row_status") not in ("", "active"):
            continue
        if not builder.quantity_resource_uses_universal_z(row["resource"]):
            continue
        earliest_year = row.get("earliest_commercial_activity_year", "")
        flagship_year = row.get("flagship_scale_year", "")
        if bool(earliest_year) != bool(flagship_year):
            chronology_pair_failures.append(f"{row['state']} / {row['resource']}")
    results.append(CheckResult("quantity-resource chronology moderation rows set both commercial years together", "PASS" if not chronology_pair_failures else "FAIL", "; ".join(chronology_pair_failures[:10])))

    active_floor_by_key = {}
    for row in raw_adjustment_rows:
        if row.get("row_status") not in ("", "active"):
            continue
        minimum_floor = int(float(row.get("minimum_operating_floor_cap") or 0))
        active_floor_by_key[(row["state"], row["resource"])] = minimum_floor
    ceil_failures = []
    for row in final_caps:
        if row["status"] in {"explicit exception", "denominator unavailable", "review required", "constrained zero"}:
            continue
        adjusted_cap = row.get("adjusted_cap", "")
        if adjusted_cap in ("", None):
            continue
        floor = active_floor_by_key.get((row["state"], row["resource"]), 0)
        expected_cap = max(math.ceil(float(adjusted_cap)), 0)
        if floor:
            expected_cap = max(expected_cap, floor)
        actual_cap = int(float(row["final_audited_cap"]))
        if actual_cap != expected_cap:
            ceil_failures.append(f"{row['state']} / {row['resource']}: expected {expected_cap}, got {actual_cap}")
    results.append(CheckResult("integer cap conversion now uses ceil with active floors preserved", "PASS" if not ceil_failures else "FAIL", "; ".join(ceil_failures[:10])))

    arable_resources = {resource for category, resource in builder.BINARY_RESOURCES if category == "Arable Resource"}
    arable_selection_rows = [row for row in gdp_selection_rows if row["resource"] in arable_resources or row["resource"] == "Arable Land"]
    results.append(CheckResult("arable rows are removed from gdp selection output", "PASS" if not arable_selection_rows else "FAIL", f"Unexpected arable GDP rows: {len(arable_selection_rows)}"))
    wood_selection_rows = [row for row in gdp_selection_rows if row["resource"] == "Wood"]
    results.append(CheckResult("wood rows are removed from gdp selection output", "PASS" if not wood_selection_rows else "FAIL", f"Unexpected wood GDP rows: {len(wood_selection_rows)}"))
    rubber_selection_rows = [row for row in gdp_selection_rows if row["resource"] == "Rubber (undiscovered)"]
    results.append(CheckResult("latent-rubber rows are removed from gdp selection output", "PASS" if not rubber_selection_rows else "FAIL", f"Unexpected rubber GDP rows: {len(rubber_selection_rows)}"))

    expected_target_capacity_rows = len(builder.STATE_INFO) * len(builder.ARABLE_LAND_CLASSES)
    results.append(CheckResult("arable target capacity rows cover all states and land classes", "PASS" if len(arable_target_capacity_rows) == expected_target_capacity_rows else "FAIL", f"Rows={len(arable_target_capacity_rows)}, expected={expected_target_capacity_rows}"))
    results.append(CheckResult("arable comparator capacity rows are populated", "PASS" if len(arable_comparator_capacity_rows) > 0 else "FAIL", f"Rows={len(arable_comparator_capacity_rows)}"))
    expected_wood_target_rows = len(builder.STATE_INFO) * len(builder.WOOD_LAND_CLASSES)
    results.append(CheckResult("wood target capacity rows cover all states and land classes", "PASS" if len(wood_target_capacity_rows) == expected_wood_target_rows else "FAIL", f"Rows={len(wood_target_capacity_rows)}, expected={expected_wood_target_rows}"))
    results.append(CheckResult("wood comparator capacity rows are populated", "PASS" if len(wood_comparator_capacity_rows) > 0 else "FAIL", f"Rows={len(wood_comparator_capacity_rows)}"))
    expected_rubber_target_rows = len(builder.STATE_INFO) * len(builder.RUBBER_LAND_CLASSES)
    results.append(CheckResult("rubber target capacity rows cover all states and land classes", "PASS" if len(rubber_target_capacity_rows) == expected_rubber_target_rows else "FAIL", f"Rows={len(rubber_target_capacity_rows)}, expected={expected_rubber_target_rows}"))
    results.append(CheckResult("rubber comparator capacity rows are populated", "PASS" if len(rubber_comparator_capacity_rows) >= 20 * len(builder.RUBBER_LAND_CLASSES) else "FAIL", f"Rows={len(rubber_comparator_capacity_rows)}"))

    shared_row = arable_shared[0]
    shared_value = float(shared_row["shared_sb_arable_effective_hectares_per_cap"])
    state_mean_values = [float(row["weighted_state_effective_hectares_per_cap"]) for row in arable_state_means if row["weighted_state_effective_hectares_per_cap"] not in ("", None)]
    expected_shared = sum(state_mean_values) / len(state_mean_values)
    results.append(CheckResult("shared arable denominator matches simple mean of state means", "PASS" if abs(shared_value - expected_shared) < 1e-6 else "FAIL", f"shared={shared_value}, expected={expected_shared}"))

    arable_denominator_failures = []
    for row in final_caps:
        if row["resource"] != "Arable Land":
            continue
        if abs(float(row["denominator_units_per_cap"]) - shared_value) > 1e-6:
            arable_denominator_failures.append(row["state"])
    results.append(CheckResult("all arable rows use the shared land-capacity denominator", "PASS" if not arable_denominator_failures else "FAIL", "States with non-shared arable denominator: " + ", ".join(arable_denominator_failures)))

    arable_contract_failures = []
    for row in final_caps:
        if row["resource"] != "Arable Land":
            continue
        if row["selection_mode"] or row["gdp_geography_used"] or row["gdp_anchor_year"]:
            arable_contract_failures.append(f"{row['state']} / GDP metadata should be blank")
        if abs(float(row["output_addition_y"] or 0)) > 0 or abs(float(row["plausibility_haircut_z"] or 0)) > 0:
            arable_contract_failures.append(f"{row['state']} / YZ should be zero")
    results.append(CheckResult("arable rows use direct land-capacity X with no GDP fields or legacy Y/Z", "PASS" if not arable_contract_failures else "FAIL", "; ".join(arable_contract_failures[:10])))

    land_z_failures = []
    for row in final_caps:
        if row["resource"] not in {"Arable Land", "Wood", "Rubber (undiscovered)"}:
            continue
        if abs(float(row["plausibility_haircut_z"] or 0)) > 1e-9:
            land_z_failures.append(f"{row['state']} / {row['resource']}")
    results.append(CheckResult("hectare families bypass the universal quantity-resource Z rule", "PASS" if not land_z_failures else "FAIL", "; ".join(land_z_failures[:10])))

    arable_caps = {row["state"]: int(float(row["final_audited_cap"])) for row in final_caps if row["resource"] == "Arable Land"}
    spot_failures = []
    if arable_caps.get("Eastern Cape", 0) < 30:
        spot_failures.append("Eastern Cape should be materially above the old single-digit result")
    if arable_caps.get("Cape Colony", 0) < 30:
        spot_failures.append("Cape Colony should stay a major commercial farming state")
    if not (5 <= arable_caps.get("Northern Cape", 0) <= 20):
        spot_failures.append("Northern Cape should stay constrained")
    if not (1 <= arable_caps.get("Botswana", 0) <= 10):
        spot_failures.append("Botswana should stay low but nonzero")
    if arable_caps.get("Namaqualand", 0) > 2:
        spot_failures.append("Namaqualand should stay very low")
    if arable_caps.get("Eastern Transvaal", 0) <= arable_caps.get("West Transvaal", 0):
        spot_failures.append("Eastern Transvaal should outrank West Transvaal")
    results.append(CheckResult("arable spot-check outcomes are playable and directionally plausible", "PASS" if not spot_failures else "FAIL", "; ".join(spot_failures)))

    arable_expectation_headers = set(arable_resource_expectations[0].keys()) if arable_resource_expectations else set()
    required_expectation_headers = {
        "researched_plausible",
        "live_enabled_in_state",
        "basket_membership_status",
        "audit_relevance",
        "gameplay_note",
    }
    results.append(CheckResult("arable resource expectations are now a gameplay audit surface", "PASS" if required_expectation_headers <= arable_expectation_headers else "FAIL", "Missing fields: " + ", ".join(sorted(required_expectation_headers - arable_expectation_headers))))

    expectation_map = {(row["state"], row["resource"]): row for row in arable_resource_expectations}
    basket_logic_failures = []
    eastern_cape_wheat = expectation_map.get(("Eastern Cape", "Wheat Farm"))
    if eastern_cape_wheat is None or not (
        eastern_cape_wheat["researched_plausible"] == "yes"
        and eastern_cape_wheat["live_enabled_in_state"] == "yes"
        and eastern_cape_wheat["basket_membership_status"] == "researched_and_live"
    ):
        basket_logic_failures.append("Eastern Cape / Wheat Farm")
    cape_tea = expectation_map.get(("Cape Colony", "Tea Plantation"))
    if cape_tea is None or not (
        cape_tea["researched_plausible"] == "no"
        and cape_tea["live_enabled_in_state"] == "no"
        and cape_tea["basket_membership_status"] == "excluded"
    ):
        basket_logic_failures.append("Cape Colony / Tea Plantation")
    results.append(CheckResult("arable basket/live expectations reflect the synced gameplay state", "PASS" if not basket_logic_failures else "FAIL", "Basket logic failures: " + ", ".join(basket_logic_failures)))

    overview_headers, overview_rows = extract_table(overview_ws, "SB totals before and after")
    overview_map = {str(row["Resource"]): row for row in overview_rows}
    overview_mismatches = []
    expected_overview = {row["resource"]: row for row in regional_totals if row["region"] == "SB Scope"}
    for resource, expected in expected_overview.items():
        actual = overview_map.get(resource)
        if actual is None:
            overview_mismatches.append(resource)
            continue
        if (
            str(actual["Vanilla baseline"]) != expected["vanilla_total"]
            or str(actual["SB update"]) != expected["final_sb_total"]
            or str(actual["Delta vs vanilla"]) != expected["absolute_delta"]
        ):
            overview_mismatches.append(resource)
    results.append(CheckResult("overview totals mirror regional_resource_totals.csv", "PASS" if not overview_mismatches else "FAIL", "Mismatches: " + ", ".join(overview_mismatches[:10])))

    tag_headers, tag_rows = extract_table(overview_ws, "Major tag changes")
    tag_map = {(str(row["Tag"]), str(row["Resource"])): row for row in tag_rows}
    tag_mismatches = []
    expected_tag_rows = [row for row in regional_totals if row["region"] in builder.WORKBOOK_TAG_REGIONS]
    for expected in expected_tag_rows:
        key = (expected["region"], expected["resource"])
        actual = tag_map.get(key)
        if actual is None:
            tag_mismatches.append(f"{expected['region']} / {expected['resource']}")
            continue
        if (
            str(actual["Vanilla baseline"]) != expected["vanilla_total"]
            or str(actual["SB update"]) != expected["final_sb_total"]
            or str(actual["Delta vs vanilla"]) != expected["absolute_delta"]
        ):
            tag_mismatches.append(f"{expected['region']} / {expected['resource']}")
    results.append(CheckResult("major tag changes mirror regional_resource_totals.csv", "PASS" if not tag_mismatches else "FAIL", "Mismatches: " + ", ".join(tag_mismatches[:10])))

    progress_headers, progress_rows = extract_table(overview_ws, "State pass progress")
    progress_map = {str(row["State"]): row for row in progress_rows}
    progress_failures = []
    for tracker_row in state_pass_tracker:
        actual = progress_map.get(tracker_row["state"])
        if actual is None:
            progress_failures.append(f"missing {tracker_row['state']}")
            continue
        expected_status = builder.tracker_status_for_workbook(tracker_row["pass_status"])
        if (
            str(actual["Pass order"]) != str(tracker_row["pass_order"])
            or str(actual["Status"]) != expected_status
            or str(actual["Completed rows"]) != str(tracker_row["completed_rows"])
            or str(actual["Changed rows"]) != str(tracker_row["changed_rows"])
        ):
            progress_failures.append(tracker_row["state"])
    results.append(CheckResult("overview progress block mirrors state_pass_tracker.csv", "PASS" if not progress_failures else "FAIL", "Mismatches: " + ", ".join(progress_failures[:10])))

    csv_map = {(row["state"], row["resource"]): row for row in final_caps}
    expectation_sheet_map = {(row["state"], row["resource"]): row for row in arable_resource_expectations}
    counterfactual_audit_map = {(row["state"], row["resource"]): row for row in state_counterfactual_audit}
    vanilla_priors = builder.load_vanilla_priors()
    state_sheet_mismatches = []
    basis_failures = []
    for state, ws in state_sheet_map.items():
        headers, sheet_rows = extract_table(ws, "Resource rows")
        if "Basis" not in headers:
            basis_failures.append(f"{state} / Basis column missing")
        _, sheet_rows = extract_table(ws, "Resource rows")
        sheet_map = {str(row["Resource"]): row for row in sheet_rows}
        for category, resource in builder.WORKBOOK_STATE_RESOURCE_ORDER:
            actual = sheet_map.get(resource)
            if actual is None:
                state_sheet_mismatches.append(f"{state} / missing {resource}")
                continue
            if category == "Arable Resource":
                expected_vanilla = str(vanilla_priors[state].get(resource, "no")).lower()
                expected_sb = expectation_sheet_map[(state, resource)]["researched_plausible"]
                if str(actual["Vanilla baseline"]).lower() != expected_vanilla or str(actual["SB update"]).lower() != expected_sb:
                    state_sheet_mismatches.append(f"{state} / {resource}")
                if str(actual.get("Basis", "")) != builder.visible_basis_label(counterfactual_audit_map[(state, resource)]["driving_basis"]):
                    basis_failures.append(f"{state} / {resource}")
                continue
            expected = csv_map[(state, resource)]
            if str(actual["Vanilla baseline"]) != expected["vanilla_cap"] or str(actual["SB update"]) != expected["final_audited_cap"]:
                state_sheet_mismatches.append(f"{state} / {resource}")
            if str(actual.get("Basis", "")) != builder.visible_basis_label(counterfactual_audit_map[(state, resource)]["driving_basis"]):
                basis_failures.append(f"{state} / {resource}")
    results.append(CheckResult("state sheets mirror vanilla baselines and audited SB updates", "PASS" if not state_sheet_mismatches else "FAIL", "Mismatches: " + ", ".join(state_sheet_mismatches[:12])))
    results.append(CheckResult("state sheets expose the public Basis column", "PASS" if not basis_failures else "FAIL", "Basis issues: " + ", ".join(basis_failures[:12])))

    sync_expected_states = {
        row["state"]
        for row in state_pass_tracker
        if row["live_synced"] == "yes" and row["pass_status"] == builder.TRACKER_ACCEPTED
    }
    live_mismatches = []
    ignored_live_mismatches = []
    for row in final_caps:
        expected = int(float(row["final_audited_cap"]))
        actual = live_values[row["state"]].get(row["resource"], 0)
        if expected == actual:
            continue
        detail = f"{row['state']} / {row['resource']}: csv={expected}, live={actual}"
        if builder.SYNC_LIVE_ON_BUILD or row["state"] in sync_expected_states:
            live_mismatches.append(detail)
        else:
            ignored_live_mismatches.append(detail)
    if builder.SYNC_LIVE_ON_BUILD:
        results.append(CheckResult("final caps match live state file", "PASS" if not live_mismatches else "FAIL", f"{len(live_mismatches)} mismatches against the live resource file."))
    elif sync_expected_states:
        detail = f"{len(live_mismatches)} mismatches on accepted synced states."
        if ignored_live_mismatches:
            detail += f" Ignored {len(ignored_live_mismatches)} mismatches on in-review or unsynced states."
        results.append(CheckResult("final caps match live state file", "PASS" if not live_mismatches else "FAIL", detail))
    else:
        results.append(CheckResult("live state file sync remains frozen", "PASS", f"Auto-sync disabled; {len(ignored_live_mismatches)} live mismatches are expected during the audit pass."))

    arable_live_mismatches = []
    expectation_map = {(row["state"], row["resource"]): row for row in arable_resource_expectations}
    for state in sync_expected_states:
        for _category, resource in builder.BINARY_RESOURCES:
            expected = expectation_map[(state, resource)]["researched_plausible"]
            actual = str(live_values[state].get(resource, "no")).lower()
            if actual != expected:
                arable_live_mismatches.append(f"{state} / {resource}: expected {expected}, live {actual}")
    if sync_expected_states:
        results.append(CheckResult("accepted synced states match live arable resources", "PASS" if not arable_live_mismatches else "FAIL", "; ".join(arable_live_mismatches[:12])))
    else:
        results.append(CheckResult("accepted synced states match live arable resources", "PASS", "No accepted synced states yet."))

    builder_text = BUILDER.read_text(encoding="utf-8")
    wood_builder_failures = []
    if "RAW_WOOD_TARGET_CAPACITY_CSV" not in builder_text or "RAW_WOOD_COMPARATOR_CAPACITY_CSV" not in builder_text:
        wood_builder_failures.append("wood capacity raw paths missing from builder")
    wood_denominator = next((row for row in resource_denominators if row["resource_family"] == "Wood"), None)
    if wood_denominator is None or wood_denominator["method"] != "mean_of_effective_commercial_forestry_comparator_pool":
        wood_builder_failures.append("wood denominator method")
    if wood_denominator is None or not (20000 <= float(wood_denominator["denominator_units_per_cap"] or 0) <= 40000):
        wood_builder_failures.append("wood denominator outside the expected playable range")
    results.append(CheckResult("wood uses dedicated effective-forestry denominator path", "PASS" if not wood_builder_failures else "FAIL", "Wood path failures: " + ", ".join(wood_builder_failures)))

    non_arable_builder_failures = []
    if "non_arable_benchmark_cases.csv" not in builder_text:
        non_arable_builder_failures.append("builder still points at legacy mining benchmark file")
    if "Gold Mine" not in {row["resource_family"] for row in resource_denominators}:
        non_arable_builder_failures.append("Gold Mine denominator row missing")
    gold_mine_denominator = next((row for row in resource_denominators if row["resource_family"] == "Gold Mine"), None)
    if gold_mine_denominator is None or gold_mine_denominator["status"] != "formula-driven":
        non_arable_builder_failures.append("Gold Mine should inherit a formula-driven gold comparator family")
    results.append(CheckResult("non-arable benchmark registry and gold-mine denominator path are wired", "PASS" if not non_arable_builder_failures else "FAIL", "; ".join(non_arable_builder_failures)))

    if any(row["resource"] == "Wood" for row in target_observations):
        results.append(CheckResult("public target observations no longer expose formula-driving wood estate rows", "FAIL", "Wood rows still present in target_observations.csv"))
    else:
        results.append(CheckResult("public target observations no longer expose formula-driving wood estate rows", "PASS", "Wood rows are absent from target_observations.csv"))

    target_headers = set(target_observations[0].keys()) if target_observations else set()
    required_target_headers = {
        "discounted_normalized_1940_output",
        "discounted_wheat_equivalent_1940_output",
        "drives_x",
        "review_action",
        *builder.TARGET_VALIDATION_FIELDS,
    }
    results.append(CheckResult("target observation output exposes validation and discounted evidence fields", "PASS" if required_target_headers <= target_headers else "FAIL", "Missing fields: " + ", ".join(sorted(required_target_headers - target_headers))))

    wood_row_failures = []
    for row in final_caps:
        if row["resource"] != "Wood":
            continue
        if row["selection_mode"] or row["gdp_geography_used"] or row["gdp_anchor_year"]:
            wood_row_failures.append(f"{row['state']} / GDP metadata should be blank")
    results.append(CheckResult("wood rows use the land-capacity contract with no GDP metadata", "PASS" if not wood_row_failures else "FAIL", "; ".join(wood_row_failures[:10])))

    restoration_failures = []
    for collection, group_field in [(wood_target_capacity_rows, "state"), (wood_comparator_capacity_rows, "comparator_geography")]:
        grouped: dict[str, dict[str, float]] = {}
        for row in collection:
            group = row[group_field]
            payload = grouped.setdefault(group, {"plantation": 0.0, "restoration": 0.0})
            land_class = row["land_class"]
            effective = float(row["effective_area_ha"] or 0)
            if land_class in {"high_suitability_plantation", "moderate_suitability_plantation"}:
                payload["plantation"] += effective
            elif land_class == "restorable_commercial_forest":
                payload["restoration"] += effective
        for group, payload in grouped.items():
            if payload["restoration"] > (payload["plantation"] * 0.5) + 1e-6:
                restoration_failures.append(group)
    results.append(CheckResult("wood restoration allowance never exceeds the 50 percent cap", "PASS" if not restoration_failures else "FAIL", "Cap failures: " + ", ".join(restoration_failures[:10])))

    wood_caps = {row["state"]: int(float(row["final_audited_cap"])) for row in final_caps if row["resource"] == "Wood"}
    wood_spot_failures = []
    if wood_caps.get("Cape Colony", 0) < 3:
        wood_spot_failures.append("Cape Colony should rise above the old estate-footprint result")
    if wood_caps.get("Eastern Cape", 0) <= wood_caps.get("Cape Colony", 0):
        wood_spot_failures.append("Eastern Cape should outrank Cape Colony")
    if wood_caps.get("Eastern Transvaal", 0) < wood_caps.get("Eastern Cape", 0):
        wood_spot_failures.append("Eastern Transvaal should remain the strongest wood state")
    if wood_caps.get("Botswana", 0) != 0:
        wood_spot_failures.append("Botswana should stay at zero")
    if wood_caps.get("Namaqualand", 0) != 0:
        wood_spot_failures.append("Namaqualand should stay at zero")
    results.append(CheckResult("wood spot-check outcomes are directionally plausible", "PASS" if not wood_spot_failures else "FAIL", "; ".join(wood_spot_failures)))

    rubber_policy_failures = []
    latent_rubber_caps = {row["state"]: int(float(row["final_audited_cap"])) for row in final_caps if row["resource"] == "Rubber (undiscovered)"}
    for state in ["Lourenço Marques", "Zambezi"]:
        row = next((item for item in final_caps if item["state"] == state and item["resource"] == "Rubber (undiscovered)"), None)
        if row is None:
            rubber_policy_failures.append(f"missing {state} / Rubber (undiscovered)")
            continue
        if int(float(row["final_audited_cap"])) <= 0:
            rubber_policy_failures.append(f"{state} / Rubber (undiscovered) should be formula-driven and nonzero under v3")
        if row["status"] != "formula-driven":
            rubber_policy_failures.append(f"{state} / Rubber (undiscovered) should now be formula-driven")
    dry_rubber_row = next((item for item in final_caps if item["state"] == "Namaqualand" and item["resource"] == "Rubber (undiscovered)"), None)
    if dry_rubber_row is None or int(float(dry_rubber_row["final_audited_cap"])) != 0 or dry_rubber_row["status"] != "explicit exception":
        rubber_policy_failures.append("Namaqualand / Rubber (undiscovered) should remain a defended zero")
    discovered_rubber_failures = []
    for row in final_caps:
        if row["resource"] != "Rubber (discovered)":
            continue
        if int(float(row["final_audited_cap"])) != 0:
            discovered_rubber_failures.append(f"{row['state']} discovered rubber should remain zero")
        if row["status"] != "explicit exception":
            discovered_rubber_failures.append(f"{row['state']} discovered rubber should remain explicit exception")
    results.append(CheckResult("latent rubber is hectare-based while discovered rubber stays exceptional", "PASS" if not (rubber_policy_failures or discovered_rubber_failures) else "FAIL", "; ".join((rubber_policy_failures + discovered_rubber_failures)[:12])))

    rubber_denominator = next((row for row in resource_denominators if row["resource_family"] == "Rubber"), None)
    rubber_denominator_failures = []
    if rubber_denominator is None:
        rubber_denominator_failures.append("Rubber denominator row missing")
    else:
        if rubber_denominator["status"] != "formula-driven":
            rubber_denominator_failures.append("Rubber denominator should be formula-driven")
        if rubber_denominator["method"] != "mean_of_effective_latent_rubber_comparator_pool":
            rubber_denominator_failures.append("Rubber denominator method mismatch")
    results.append(CheckResult("latent-rubber denominator is built from the comparator hectare pool", "PASS" if not rubber_denominator_failures else "FAIL", "; ".join(rubber_denominator_failures)))

    validation_map = {(row["state"], row["resource"], row["land_class"]): row for row in target_data_validation}
    provenance_failures = []
    mozambique_row = validation_map.get(("Lourenço Marques", "Wood", "high_suitability_plantation"))
    if mozambique_row is None:
        provenance_failures.append("missing Lourenço Marques / Wood validation row")
    else:
        if mozambique_row["drives_x"] != "no":
            provenance_failures.append("Lourenço Marques / Wood still drives x from broad Mozambique forestry potential")
        if mozambique_row["slot_support_status"] != "broad_potential_only":
            provenance_failures.append("Lourenço Marques / Wood should remain broad_potential_only")
    northern_transvaal_validation = validation_map.get(("Northern Transvaal", "Wood", "high_suitability_plantation"))
    if wood_caps.get("Northern Transvaal", 0) > 0 and (northern_transvaal_validation is None or northern_transvaal_validation["drives_x"] != "yes"):
        provenance_failures.append("Northern Transvaal nonzero wood cap lacks accepted bounded forestry evidence")
    zambezi_validation = validation_map.get(("Zambezi", "Wood", "high_suitability_plantation"))
    if wood_caps.get("Zambezi", 0) > 0 and (zambezi_validation is None or zambezi_validation["drives_x"] != "yes"):
        provenance_failures.append("Zambezi nonzero wood cap lacks accepted bounded fallback evidence")
    drakensberg_row = next((row for row in final_caps if row["state"] == "Drakensberg" and row["resource"] == "Wood"), None)
    if drakensberg_row is None or drakensberg_row["status"] not in {"constrained zero", "explicit exception"}:
        provenance_failures.append("Drakensberg / Wood should remain a defended zero")
    rubber_lm_row = validation_map.get(("Lourenço Marques", "Rubber (undiscovered)", "high_suitability_plantation"))
    if rubber_lm_row is None or rubber_lm_row["drives_x"] != "yes":
        provenance_failures.append("Lourenço Marques / Rubber (undiscovered) lacks accepted bounded plantation evidence")
    rubber_zam_row = validation_map.get(("Zambezi", "Rubber (undiscovered)", "high_suitability_plantation"))
    if rubber_zam_row is None or rubber_zam_row["drives_x"] != "yes":
        provenance_failures.append("Zambezi / Rubber (undiscovered) lacks accepted bounded plantation evidence")
    rubber_dry_row = validation_map.get(("Namaqualand", "Rubber (undiscovered)", "high_suitability_plantation"))
    if rubber_dry_row is None or rubber_dry_row["drives_x"] != "no":
        provenance_failures.append("Namaqualand / Rubber (undiscovered) should remain non-driving")
    results.append(CheckResult("sentinel provenance scenarios behave as intended", "PASS" if not provenance_failures else "FAIL", "; ".join(provenance_failures)))

    row_audit_map = {(row["state"], row["resource"]): row for row in row_audit}
    results.append(CheckResult("row audit file exists", "PASS" if ROW_AUDIT.exists() and len(row_audit) == len(final_caps) else "FAIL", f"row_audit rows={len(row_audit)}, final_caps rows={len(final_caps)}"))

    missing_audit_fields = []
    for row in row_audit:
        required = ["representative_max_year", "year_selection_reason", "counterevidence_status", "plausibility_1936_status", "observed_output_x_reason", "y_reason", "y_counterevidence_trigger", "y_quantification_method", "z_reason", "audit_class"]
        if any(row[field] in ("", None) for field in required):
            missing_audit_fields.append(f"{row['state']} / {row['resource']}")
    results.append(CheckResult("row audit fields are populated", "PASS" if not missing_audit_fields else "FAIL", f"Rows missing audit fields: {len(missing_audit_fields)}"))

    tracker_status_map = {row["state"]: row["pass_status"] for row in state_pass_tracker}
    unsupported_zero_rows = []
    for row in final_caps:
        if int(float(row["final_audited_cap"])) != 0 or float(row["observed_output_x"] or 0) != 0:
            continue
        if tracker_status_map.get(row["state"]) == builder.TRACKER_NOT_STARTED:
            continue
        audit = row_audit_map[(row["state"], row["resource"])]
        if audit["counterevidence_status"] == "not_reviewed" and row["exception_status"] == "":
            unsupported_zero_rows.append(f"{row['state']} / {row['resource']}")
    results.append(CheckResult("hard zeros carry explicit audit support or exception status", "PASS" if not unsupported_zero_rows else "FAIL", f"Unsupported hard zeros: {len(unsupported_zero_rows)}"))

    expected_public_rows = len(builder.STATE_INFO) * len(builder.WORKBOOK_STATE_RESOURCE_ORDER)
    results.append(CheckResult("state counterfactual audit covers the full public workbook surface", "PASS" if STATE_COUNTERFACTUAL_AUDIT.exists() and len(state_counterfactual_audit) == expected_public_rows else "FAIL", f"audit_rows={len(state_counterfactual_audit)}, expected={expected_public_rows}"))

    counterfactual_field_failures = []
    changed_row_without_citations = []
    for row in state_counterfactual_audit:
        required = ["historical_label", "counterfactual_label", "driving_basis", "decision", "row_category"]
        if any(row[field] in ("", None) for field in required):
            counterfactual_field_failures.append(f"{row['state']} / {row['resource']}")
        if (
            row["decision"] != "keep"
            and int(row.get("state_pass_index") or 0) > 0
            and not (row["citation_1_title"] or row["citation_2_title"])
        ):
            changed_row_without_citations.append(f"{row['state']} / {row['resource']}")
    results.append(CheckResult("every audited public row has both labels and one driving basis", "PASS" if not counterfactual_field_failures else "FAIL", "; ".join(counterfactual_field_failures[:12])))
    results.append(CheckResult("no changed row lacks citations", "PASS" if not changed_row_without_citations else "FAIL", "; ".join(changed_row_without_citations[:12])))
    arable_counterevidence_failures = []
    arable_resources = {resource for category, resource in builder.BINARY_RESOURCES if category == "Arable Resource"}
    counterfactual_audit_map = {(row["state"], row["resource"]): row for row in state_counterfactual_audit}
    for row in raw_counterevidence_rows:
        if row.get("row_status") not in ("", "active") or row["resource"] not in arable_resources:
            continue
        audit_row = counterfactual_audit_map.get((row["state"], row["resource"]))
        if (
            audit_row is None
            or audit_row["proposed_value"] != "no"
            or int(audit_row.get("state_pass_index") or 0) <= 0
        ):
            continue
        surfaced_titles = {audit_row["citation_1_title"], audit_row["citation_2_title"]}
        if row["source_a_title"] not in surfaced_titles and row["source_b_title"] not in surfaced_titles:
            arable_counterevidence_failures.append(f"{row['state']} / {row['resource']}")
    results.append(
        CheckResult(
            "arable counterevidence rows surface in the public audit when they defend a no row",
            "PASS" if not arable_counterevidence_failures else "FAIL",
            "; ".join(arable_counterevidence_failures[:12]),
        )
    )
    arable_positive_counterevidence_failures = []
    for row in raw_counterevidence_rows:
        if row.get("row_status") not in ("", "active") or row["resource"] not in arable_resources:
            continue
        if "enabled" not in str(row.get("decision", "")).lower():
            continue
        audit_row = counterfactual_audit_map.get((row["state"], row["resource"]))
        if (
            audit_row is None
            or audit_row["proposed_value"] != "yes"
            or int(audit_row.get("state_pass_index") or 0) <= 0
        ):
            continue
        surfaced_titles = {audit_row["citation_1_title"], audit_row["citation_2_title"]}
        if row["source_a_title"] not in surfaced_titles and row["source_b_title"] not in surfaced_titles:
            arable_positive_counterevidence_failures.append(f"{row['state']} / {row['resource']}")
    results.append(
        CheckResult(
            "arable counterevidence rows surface in the public audit when they defend a yes row",
            "PASS" if not arable_positive_counterevidence_failures else "FAIL",
            "; ".join(arable_positive_counterevidence_failures[:12]),
        )
    )

    tracker_failures = []
    if len(state_pass_tracker) != len(builder.STATE_INFO):
        tracker_failures.append(f"expected {len(builder.STATE_INFO)} tracker rows, got {len(state_pass_tracker)}")
    tracker_map = {row["state"]: row for row in state_pass_tracker}
    for info in builder.STATE_INFO:
        state = info["official_name"]
        row = tracker_map.get(state)
        if row is None:
            tracker_failures.append(f"missing {state}")
            continue
        if int(row["pass_order"]) != builder.STATE_PASS_ORDER[state]:
            tracker_failures.append(f"{state} order mismatch")
    results.append(CheckResult("state pass tracker rows are present and in fixed order", "PASS" if not tracker_failures else "FAIL", "; ".join(tracker_failures[:12])))
    tracker_state_failures = []
    valid_tracker_statuses = {
        builder.TRACKER_NOT_STARTED,
        builder.TRACKER_IN_REVIEW,
        builder.TRACKER_ACCEPTED,
        builder.TRACKER_RERUN_REQUIRED,
    }
    for row in state_pass_tracker:
        status = row["pass_status"]
        state = row["state"]
        if status not in valid_tracker_statuses:
            tracker_state_failures.append(f"{state} invalid status {status}")
        if status == builder.TRACKER_ACCEPTED and row.get("last_completed_pass_index") in ("", None):
            tracker_state_failures.append(f"{state} accepted without a completed pass index")
        if status != builder.TRACKER_ACCEPTED and row.get("live_synced") == "yes":
            tracker_state_failures.append(f"{state} live_synced=yes before acceptance")
    if any(row["live_synced"] == "yes" for row in state_pass_tracker) and any(
        row["pass_status"] != builder.TRACKER_ACCEPTED for row in state_pass_tracker
    ):
        tracker_state_failures.append("live sync should remain frozen until every state is accepted")
    results.append(
        CheckResult(
            "loop tracker remains coherent across reset and in-progress passes",
            "PASS" if not tracker_state_failures else "FAIL",
            "; ".join(tracker_state_failures[:12]),
        )
    )
    reset_baseline = all(row["pass_status"] == builder.TRACKER_NOT_STARTED for row in state_pass_tracker) and not family_rewrite_log
    tracker_surface_detail = (
        "Tracker is still at the reset baseline."
        if reset_baseline
        else "Tracker/family rewrite surfaces are carrying stateful loop progress."
    )
    results.append(
        CheckResult(
            "loop surfaces persist stateful progress after the v3 reset",
            "PASS",
            tracker_surface_detail,
        )
    )

    supersede_failures = []
    for _name, rows, _logical_key_fields in raw_tables_to_check:
        row_ids = {row["row_id"] for row in rows}
        for row in rows:
            if row.get("supersedes_row_id") and row["supersedes_row_id"] not in row_ids:
                supersede_failures.append(row["supersedes_row_id"])
    results.append(CheckResult("superseded rows remain present after a change", "PASS" if not supersede_failures else "FAIL", "; ".join(supersede_failures[:10])))

    family_rewrite_failures = []
    for row in family_rewrite_log:
        if row.get("rerun_completed") != "yes":
            continue
        affected_states = [state.strip() for state in str(row.get("affected_states", "")).split(";") if state.strip()]
        for state in affected_states:
            tracker_row = tracker_map.get(state)
            if tracker_row is None or tracker_row.get("last_completed_pass_index") in ("", None):
                family_rewrite_failures.append(f"{row['resource_family']} / {state}")
    results.append(CheckResult("no family rewrite is marked complete unless affected completed states were rerun", "PASS" if not family_rewrite_failures else "FAIL", "; ".join(family_rewrite_failures[:10])))

    regional_expected: dict[tuple[str, str], tuple[int, int]] = {}
    vanilla_priors = builder.load_vanilla_priors()
    vanilla_totals = builder.load_adjusted_vanilla_totals(vanilla_priors)
    final_caps_map = {(row["state"], row["resource"]): int(float(row["final_audited_cap"])) for row in final_caps}
    for region, states in builder.REGIONS.items():
        for resource in builder.SUMMARY_RESOURCES:
            vanilla_total = vanilla_totals[region][resource]
            final_total = sum(final_caps_map[(state, resource)] for state in states)
            regional_expected[(region, resource)] = (vanilla_total, final_total)
    regional_mismatches = []
    for row in regional_totals:
        expected = regional_expected[(row["region"], row["resource"])]
        if int(float(row["vanilla_total"])) != expected[0] or int(float(row["final_sb_total"])) != expected[1]:
            regional_mismatches.append(f"{row['region']} / {row['resource']}")
    results.append(CheckResult("regional totals match final caps aggregation", "PASS" if not regional_mismatches else "FAIL", f"{len(regional_mismatches)} regional mismatches."))

    state_summary_failures = []
    if len(state_delta_summary) != len(builder.STATE_INFO):
        state_summary_failures.append(f"expected {len(builder.STATE_INFO)} state rows, got {len(state_delta_summary)}")
    summary_map = {row["state"]: row for row in state_delta_summary}
    for row in builder.STATE_INFO:
        state = row["official_name"]
        summary = summary_map.get(state)
        if summary is None:
            state_summary_failures.append(f"missing {state}")
            continue
        expected_total = sum(int(float(cap_row["final_audited_cap"])) for cap_row in final_caps if cap_row["state"] == state)
        if int(float(summary["final_audited_total"])) != expected_total:
            state_summary_failures.append(f"{state} total mismatch")
    results.append(CheckResult("state delta summary matches final caps aggregation", "PASS" if not state_summary_failures else "FAIL", "; ".join(state_summary_failures[:10])))

    state_delta_failures = []
    if len(state_resource_deltas) != len(final_caps):
        state_delta_failures.append(f"expected {len(final_caps)} rows, got {len(state_resource_deltas)}")
    delta_map = {(row["state"], row["resource"]): row for row in state_resource_deltas}
    for row in final_caps:
        delta_row = delta_map.get((row["state"], row["resource"]))
        if delta_row is None:
            state_delta_failures.append(f"missing {row['state']} / {row['resource']}")
            continue
        if int(float(delta_row["final_audited_cap"])) != int(float(row["final_audited_cap"])):
            state_delta_failures.append(f"{row['state']} / {row['resource']} audited cap mismatch")
    report_text = STATE_DELTA_REPORT.read_text(encoding="utf-8") if STATE_DELTA_REPORT.exists() else ""
    if "## Cape Colony" not in report_text or "## Eastern Transvaal" not in report_text:
        state_delta_failures.append("markdown state delta report missing expected headings")
    results.append(CheckResult("state delta exports exist and mirror final caps", "PASS" if not state_delta_failures else "FAIL", "; ".join(state_delta_failures[:10])))

    results.append(CheckResult("priority rows file exists", "PASS" if PRIORITY_ROWS.exists() and len(priority_rows) > 0 else "FAIL", f"Priority rows: {len(priority_rows)}"))

    passes = sum(1 for result in results if result.status == "PASS")
    fails = len(results) - passes
    report = "\n".join(
        [
            "# Public Resource Audit Test Report",
            "",
            "- Date: 2026-04-22",
            f"- Passes: {passes}",
            f"- Fails: {fails}",
            "",
            "## Checks",
            "",
            *[status_line(result) for result in results],
            "",
            "## Current Priority Rows",
            "",
            *[f"- `{row['priority']}` `{row['state']} / {row['resource']}`: {row['issue_type']}" for row in priority_rows[:15]],
            "",
        ]
    )
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(report, encoding="utf-8")
    return report
