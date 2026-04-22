from __future__ import annotations

import csv
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
SCRIPTS_DIR = SCRIPT_DIR.parent
PUBLIC_ROOT = SCRIPTS_DIR.parent
REPO = PUBLIC_ROOT.parents[1]
RAW_DIR = PUBLIC_ROOT / "data/raw"
DERIVED_DIR = PUBLIC_ROOT / "data/derived"
AUDIT_DIR = PUBLIC_ROOT / "audit"

RAW_HISTORICAL_CSV = RAW_DIR / "historical_anchors.csv"
RAW_MODERN_CSV = RAW_DIR / "modern_maxima.csv"
RAW_GROWTH_CSV = RAW_DIR / "growth_anchor_series.csv"
RAW_BENCHMARK_CSV = RAW_DIR / "non_arable_benchmark_cases.csv"
RAW_WOOD_COMPARATOR_CSV = RAW_DIR / "wood_comparator_cases.csv"
RAW_WOOD_LAND_WEIGHTS_CSV = RAW_DIR / "wood_land_class_weights.csv"
RAW_WOOD_TARGET_CAPACITY_CSV = RAW_DIR / "wood_target_capacity_rows.csv"
RAW_WOOD_COMPARATOR_CAPACITY_CSV = RAW_DIR / "wood_comparator_capacity_rows.csv"
RAW_RANKINGS_CSV = RAW_DIR / "agri_rankings.csv"
RAW_VANILLA_PRIORS_CSV = RAW_DIR / "vanilla_priors.csv"
RAW_ARABLE_BASKETS_CSV = RAW_DIR / "arable_baskets.csv"
RAW_ARABLE_LAND_WEIGHTS_CSV = RAW_DIR / "arable_land_class_weights.csv"
RAW_ARABLE_TARGET_CAPACITY_CSV = RAW_DIR / "arable_target_capacity_rows.csv"
RAW_ARABLE_COMPARATOR_CAPACITY_CSV = RAW_DIR / "arable_comparator_capacity_rows.csv"
RAW_STATE_METADATA_CSV = RAW_DIR / "state_metadata.csv"
RAW_ADJUSTMENT_INPUTS_CSV = RAW_DIR / "adjustment_inputs.csv"
RAW_COUNTEREVIDENCE_CSV = RAW_DIR / "counterevidence_cases.csv"
RAW_FAMILY_CONSTANTS_CSV = RAW_DIR / "family_normalising_constants.csv"
RAW_GDP_REFERENCE_CSV = RAW_DIR / "gdp_reference_anchor.csv"
RAW_GDP_SERIES_CSV = RAW_DIR / "gdp_per_capita_series.csv"
RAW_GDP_MAP_CSV = RAW_DIR / "gdp_geography_map.csv"
LEGACY_OVERRIDES_CSV = AUDIT_DIR / "overrides.csv"

STATE_FILE = REPO / "map_data/state_regions/04_subsaharan_africa.txt"
OUTPUT = PUBLIC_ROOT / "RESOURCES.xlsx"
PUBLIC_TEST_REPORT = AUDIT_DIR / "test_report.md"
SYNC_LIVE_ON_BUILD = False

OLD_PUBLIC_OUTPUTS = [
    REPO / "Docs/RESOURCES.xlsx",
    REPO / "Docs/RESOURCES.md",
    DERIVED_DIR / "agri_comparator_averages.csv",
    DERIVED_DIR / "agri_comparator_maxima.csv",
    DERIVED_DIR / "comparator_raw_data.csv",
    DERIVED_DIR / "final_caps.csv",
    DERIVED_DIR / "growth_rates.csv",
    DERIVED_DIR / "mining_comparator_averages.csv",
    DERIVED_DIR / "mining_comparator_maxima.csv",
    DERIVED_DIR / "regional_totals.csv",
    DERIVED_DIR / "target_raw_data.csv",
    RAW_DIR / "mining_benchmark_cases.csv",
    RAW_DIR / "qualitative_scores.csv",
    RAW_DIR / "quantitative_read_overrides.csv",
    AUDIT_DIR / "overrides.csv",
    AUDIT_DIR / "audit_priorities.csv",
]

STATE_INFO = [
    {"sheet_key": "W_Cape", "state_id": "STATE_CAPE_COLONY", "official_name": "Cape Colony", "vanilla_proxy_id": "STATE_CAPE_COLONY", "vanilla_proxy_name": "Cape Colony"},
    {"sheet_key": "NCape_NW", "state_id": "STATE_NORTHERN_CAPE", "official_name": "Northern Cape", "vanilla_proxy_id": "STATE_NORTHERN_CAPE", "vanilla_proxy_name": "Northern Cape"},
    {"sheet_key": "E_Cape", "state_id": "STATE_EASTERN_CAPE", "official_name": "Eastern Cape", "vanilla_proxy_id": "STATE_EASTERN_CAPE", "vanilla_proxy_name": "Eastern Cape"},
    {"sheet_key": "Gauteng", "state_id": "STATE_TRANSVAAL", "official_name": "West Transvaal", "vanilla_proxy_id": "STATE_TRANSVAAL", "vanilla_proxy_name": "Transvaal"},
    {"sheet_key": "Mpm_ESW", "state_id": "STATE_EAST_TRANSVAAL", "official_name": "Eastern Transvaal", "vanilla_proxy_id": "STATE_TRANSVAAL", "vanilla_proxy_name": "Transvaal"},
    {"sheet_key": "Limpopo", "state_id": "STATE_NORTHERN_TRANSVAAL", "official_name": "Northern Transvaal", "vanilla_proxy_id": "STATE_TRANSVAAL", "vanilla_proxy_name": "Transvaal"},
    {"sheet_key": "Free_State", "state_id": "STATE_VRYSTAAT", "official_name": "Transorangia", "vanilla_proxy_id": "STATE_VRYSTAAT", "vanilla_proxy_name": "Transorangia"},
    {"sheet_key": "Lesotho", "state_id": "STATE_DRAKENSBERG", "official_name": "Drakensberg", "vanilla_proxy_id": "STATE_ZULULAND", "vanilla_proxy_name": "Zululand"},
    {"sheet_key": "Botswana", "state_id": "STATE_BOTSWANA", "official_name": "Botswana", "vanilla_proxy_id": "STATE_BOTSWANA", "vanilla_proxy_name": "Botswana"},
    {"sheet_key": "S_Moz", "state_id": "STATE_LOURENCO_MARQUES", "official_name": "Lourenço Marques", "vanilla_proxy_id": "STATE_LOURENCO_MARQUES", "vanilla_proxy_name": "Lourenço Marques"},
    {"sheet_key": "Zimbabwe", "state_id": "STATE_ZAMBEZI", "official_name": "Zambezi", "vanilla_proxy_id": "STATE_ZAMBEZI", "vanilla_proxy_name": "Zambezi"},
    {"sheet_key": "N_Namibia", "state_id": "STATE_HEREROLAND", "official_name": "Hereroland", "vanilla_proxy_id": "STATE_HEREROLAND", "vanilla_proxy_name": "Hereroland"},
    {"sheet_key": "S_Namibia", "state_id": "STATE_NAMAQUALAND", "official_name": "Namaqualand", "vanilla_proxy_id": "STATE_NAMAQUALAND", "vanilla_proxy_name": "Namaqualand"},
]
STATE_INFO_BY_NAME = {row["official_name"]: row for row in STATE_INFO}
STATE_INFO_BY_SHEET = {row["sheet_key"]: row for row in STATE_INFO}

NUMERIC_RESOURCES = [
    ("Arable", "Arable Land"),
    ("Capped Resource", "Coal Mine"),
    ("Capped Resource", "Fishing"),
    ("Capped Resource", "Gold Mine"),
    ("Capped Resource", "Iron Mine"),
    ("Capped Resource", "Lead Mine"),
    ("Capped Resource", "Sulfur Mine"),
    ("Capped Resource", "Whaling"),
    ("Capped Resource", "Wood"),
    ("Special Resource", "Gold Fields (discovered)"),
    ("Special Resource", "Gold Fields (undiscovered)"),
    ("Special Resource", "Oil (discovered)"),
    ("Special Resource", "Oil (undiscovered)"),
    ("Special Resource", "Rubber (discovered)"),
    ("Special Resource", "Rubber (undiscovered)"),
]

BINARY_RESOURCES = [
    ("Arable Resource", "Wheat Farm"),
    ("Arable Resource", "Livestock Ranch"),
    ("Arable Resource", "Vineyard"),
    ("Arable Resource", "Tea Plantation"),
    ("Arable Resource", "Maize Farm"),
    ("Arable Resource", "Millet Farm"),
    ("Arable Resource", "Sugar Plantation"),
    ("Arable Resource", "Tobacco Plantation"),
    ("Arable Resource", "Banana Plantation"),
    ("Arable Resource", "Coffee Plantation"),
    ("Arable Resource", "Cotton Plantation"),
    ("Arable Resource", "Rice Farm"),
    ("Arable Resource", "Dye Plantation"),
    ("Arable Resource", "Opium"),
    ("Arable Resource", "Silk Plantation"),
]

LAND_ECONOMY_RESOURCES = {resource for _, resource in BINARY_RESOURCES}

RESOURCE_TO_BUILDING = {
    "Cotton Plantation": "building_cotton_plantation",
    "Dye Plantation": "building_dye_plantation",
    "Livestock Ranch": "building_livestock_ranch",
    "Maize Farm": "building_maize_farm",
    "Millet Farm": "building_millet_farm",
    "Opium": "building_opium_plantation",
    "Rice Farm": "building_rice_farm",
    "Silk Plantation": "building_silk_plantation",
    "Sugar Plantation": "building_sugar_plantation",
    "Tea Plantation": "building_tea_plantation",
    "Tobacco Plantation": "building_tobacco_plantation",
    "Vineyard": "building_vineyard",
    "Wheat Farm": "building_wheat_farm",
    "Banana Plantation": "building_banana_plantation",
    "Coffee Plantation": "building_coffee_plantation",
}

CAP_BUILDINGS = {
    "Coal Mine": "building_coal_mine",
    "Fishing": "building_fishing_wharf",
    "Gold Mine": "building_gold_mine",
    "Iron Mine": "building_iron_mine",
    "Lead Mine": "building_lead_mine",
    "Sulfur Mine": "building_sulfur_mine",
    "Whaling": "building_whaling_station",
    "Wood": "building_logging_camp",
}

SPECIAL_RESOURCE_CONFIG = {
    "Gold Fields": {"type": "building_gold_field", "depleted_type": "building_gold_mine"},
    "Oil": {"type": "building_oil_rig"},
    "Rubber": {"type": "building_rubber_plantation"},
}

SUMMARY_RESOURCES = [resource for _, resource in NUMERIC_RESOURCES]
REGIONS = {
    "CAP": ["Cape Colony", "Northern Cape", "Eastern Cape"],
    "TRN": ["West Transvaal", "Eastern Transvaal", "Northern Transvaal"],
    "ORA": ["Transorangia"],
    "SAF": ["Cape Colony", "Northern Cape", "Eastern Cape", "West Transvaal", "Eastern Transvaal", "Northern Transvaal", "Transorangia"],
    "SWA": ["Hereroland", "Namaqualand"],
    "SB Scope": [row["official_name"] for row in STATE_INFO],
}
WORKBOOK_TAG_REGIONS = ["CAP", "TRN", "SAF", "SWA"]
WORKBOOK_STATE_RESOURCE_ORDER = [("Arable", "Arable Land"), *BINARY_RESOURCES, *NUMERIC_RESOURCES[1:]]

FALLBACK_ANNUAL_GROWTH_RATES = {
    "Fisheries / Marine": 0.0080,
    "Forestry": 0.0050,
}
GROWTH_YEAR_MIN = 1820
GROWTH_YEAR_MAX = 2025

CONVERSION_FACTORS = [
    ("grain_t", "Bulk grain output in tons", 1.0, "Mass-equivalent baseline for wheat, maize, millet, and rice."),
    ("cashcrop_t", "Cash-crop output in tons", 1.0, "Mass-equivalent baseline for tobacco, sugar, bananas, coffee, tea, and similar tonnage rows."),
    ("cotton_bale", "Cotton bale to wheat-equivalent tons", 0.217724, "Mass-equivalent conversion using a standard 480 lb lint bale."),
    ("wine_l", "Wine litres to grape-mass-equivalent tons", 0.0013, "Proxy conversion from finished wine litres to grape output mass."),
    ("head_cattle", "One head of cattle to wheat-equivalent tons", 0.05, "Coarse public livestock conversion used for combined land-economy comparisons."),
    ("head_sheep_goat", "One head of sheep/goat to wheat-equivalent tons", 0.005, "Coarse small-stock conversion used for combined land-economy comparisons."),
    ("head_livestock_generic", "Generic livestock head to wheat-equivalent tons", 0.02, "Fallback for livestock rows that do not distinguish cattle from sheep/goats."),
    ("skins_pelts", "One skin or pelt to wheat-equivalent tons", 0.001, "Low-weight pastoral proxy for karakul or skin-count series."),
    ("fiber_t", "Fiber output in tons", 1.0, "Mass-equivalent baseline for wool, mohair, and similar fiber outputs."),
    ("acres_farmland", "Cultivated acreage to wheat-equivalent tons", 0.0, "Acreage rows are metadata only and do not feed combined land-economy maxima."),
]
CONVERSION_FACTOR_MAP = {key: factor for key, _desc, factor, _note in CONVERSION_FACTORS}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ADJUST_FILL = PatternFill("solid", fgColor="FCE4D6")
EXCEPTION_FILL = PatternFill("solid", fgColor="EADCF8")

START_DATE_SPECIALS = {"Gold Fields (discovered)", "Oil (discovered)", "Rubber (discovered)"}
PROXY_NOTES = {
    ("Northern Cape", "Lead Mine"): ("copper output proxy", "t output", "Public target quantity is copper-tonnage evidence mapped into the lead slot as a base-metal abstraction."),
    ("Hereroland", "Lead Mine"): ("polymetal base-metal proxy", "t output", "Public target evidence is Tsumeb/Berg Aukas polymetal output rather than a clean lead-metal series."),
    ("Namaqualand", "Lead Mine"): ("latent base-metal proxy", "t output", "Public target evidence is a small latent base-metal proxy rather than directly comparable lead-metal output."),
}
GDP_SELECTION_FIELDS = [
    "gdp_geography_used",
    "gdp_geography_level",
    "gdp_anchor_year",
    "gdp_value_at_anchor_year",
    "selected_resource_year",
    "selection_mode",
    "gdp_distance_ratio",
    "gdp_comparability_status",
    "target_estimation_method",
    "interpolation_used",
    "comparator_included_in_denominator",
    "comparator_basket_coverage_share",
    "selection_note",
]
TARGET_VALIDATION_FIELDS = [
    "evidence_scope",
    "target_match_status",
    "slot_support_status",
    "localization_discount",
    "validation_note",
]
TARGET_OBSERVATION_RAW_FIELDNAMES = [
    "sheet",
    "geography",
    "resource",
    "year",
    "normalized_quantity",
    "normalized_unit",
    "source_title",
    "source_url",
    "citation_locator",
    "note",
    *TARGET_VALIDATION_FIELDS,
]
TARGET_DATA_VALIDATION_FIELDNAMES = [
    "input_family",
    "source_file",
    "state",
    "resource",
    "land_class",
    "year",
    "evidence_scope",
    "target_match_status",
    "slot_support_status",
    "localization_discount",
    "validation_note",
    "drives_x",
    "discounted_effective_value",
    "review_action",
    "source_title",
    "source_url",
    "citation_locator",
]
REGIONAL_ADVANTAGES_FIELDNAMES = [
    "state",
    "advantage_type",
    "primary_resources",
    "preferred_layer",
    "keep_out_of_caps",
    "note",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
]
STATE_REVIEW_STATUS_FIELDNAMES = [
    "state",
    "review_status",
    "largest_remaining_issue",
    "summary_note",
]
RESEARCH_DIR = REPO.parent / "References/research"
AGRI_RESEARCH_PATH = RESEARCH_DIR / "agriculture and fisheries research.md"
COMPARATOR_RESEARCH_PATH = RESEARCH_DIR / "southern_africa_comparator_states_ranked_per_region.md"


def validation_profile(
    evidence_scope: str,
    target_match_status: str,
    slot_support_status: str,
    localization_discount: float,
    validation_note: str,
) -> dict[str, Any]:
    return {
        "evidence_scope": evidence_scope,
        "target_match_status": target_match_status,
        "slot_support_status": slot_support_status,
        "localization_discount": localization_discount,
        "validation_note": validation_note,
    }


OBSERVATION_TARGET_VALIDATION_DEFAULTS_BY_STATE = {
    "Cape Colony": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Western Cape target observations are treated as state-localized evidence."),
    "Northern Cape": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.95, "Northern Cape target observations rely on a bounded Northern Cape plus North West proxy rather than a perfect state-only series."),
    "Eastern Cape": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Eastern Cape target observations are treated as state-localized evidence."),
    "West Transvaal": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.95, "West Transvaal target observations rely on a bounded Gauteng/interior plateau proxy rather than a perfect state-only series."),
    "Eastern Transvaal": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.95, "Eastern Transvaal target observations rely on a bounded Mpumalanga plus Eswatini proxy rather than a perfect state-only series."),
    "Northern Transvaal": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.90, "Northern Transvaal target observations use a bounded Limpopo proxy for the split northern state footprint."),
    "Transorangia": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Transorangia target observations are treated as state-localized evidence."),
    "Drakensberg": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Drakensberg target observations are treated as state-localized evidence."),
    "Botswana": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Botswana target observations are treated as state-localized evidence."),
    "Lourenço Marques": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.90, "Lourenço Marques target observations use a bounded southern Mozambique proxy rather than a perfect state-only series."),
    "Zambezi": validation_profile("national_fallback", "partial_overlap", "distinct_slot_supported", 0.85, "Zambezi target observations use a bounded Zimbabwe-wide fallback because many published series are national rather than split-state."),
    "Hereroland": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.90, "Hereroland target observations use a bounded northern Namibia proxy rather than a perfect state-only series."),
    "Namaqualand": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.90, "Namaqualand target observations use a bounded southern Namibia proxy rather than a perfect state-only series."),
}

ARABLE_TARGET_VALIDATION_DEFAULTS_BY_STATE = {
    "Cape Colony": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Cape Colony arable capacity rows are treated as state-localized effective commercial land."),
    "Northern Cape": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Northern Cape arable capacity is localized to Orange-Vaal irrigation corridors and extensive Karoo grazing; North West cereal intensity is excluded."),
    "Eastern Cape": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Eastern Cape arable capacity rows are treated as state-localized effective commercial land."),
    "West Transvaal": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.95, "West Transvaal arable capacity uses a bounded Gauteng/interior plateau proxy."),
    "Eastern Transvaal": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.95, "Eastern Transvaal arable capacity uses a bounded Mpumalanga plus Eswatini land-potential proxy."),
    "Northern Transvaal": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Northern Transvaal arable capacity is localized to the Limpopo mixed-farming belt and supporting northern pasture country."),
    "Transorangia": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Transorangia arable capacity rows are treated as state-localized effective commercial land."),
    "Drakensberg": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Drakensberg arable capacity rows are treated as state-localized effective commercial land."),
    "Botswana": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Botswana arable capacity rows are treated as state-localized effective commercial land."),
    "Lourenço Marques": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.85, "Lourenço Marques arable capacity is bounded to the Maputo-Gaza littoral and lower Limpopo lowland rather than to whole-Mozambique potential."),
    "Zambezi": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.85, "Zambezi arable capacity is bounded to the Eastern Highlands-Mashonaland plateau and connected lowveld belts rather than to a whole-Zimbabwe fallback."),
    "Hereroland": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Hereroland arable capacity is localized to the northern communal belt, Otavi crop pocket, and central plateau grazing country."),
    "Namaqualand": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.85, "Namaqualand arable capacity uses a bounded southern Namibia proxy."),
}

WOOD_TARGET_VALIDATION_DEFAULTS_BY_STATE = {
    "Cape Colony": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.85, "Cape Colony wood capacity is bounded by a Western Cape plantation/restoration proxy rather than a perfectly localized surviving estate series."),
    "Northern Cape": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "Northern Cape has no distinct commercial forestry slot; broad dryland woodedness does not drive x."),
    "Eastern Cape": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Eastern Cape wood capacity is treated as a localized commercial forestry belt."),
    "West Transvaal": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "West Transvaal has no distinct commercial forestry slot; wooded land does not drive x."),
    "Eastern Transvaal": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Eastern Transvaal wood capacity is treated as a localized commercial forestry belt."),
    "Northern Transvaal": validation_profile("state_localized", "direct", "distinct_slot_supported", 1.0, "Northern Transvaal wood capacity is localized to a small Limpopo plantation fringe around the eastern escarpment; it is not a broad forestry-state claim."),
    "Transorangia": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "Transorangia has no distinct commercial forestry slot; woodland does not drive x."),
    "Drakensberg": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "Drakensberg remains a defended forestry zero; small woodlots and shelterbelts do not justify a distinct commercial slot."),
    "Botswana": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "Botswana has no distinct commercial forestry slot; wooded land does not drive x."),
    "Lourenço Marques": validation_profile("regional_proxy", "partial_overlap", "broad_potential_only", 0.35, "Mozambique-wide forestry claims remain non-driving for Lourenço Marques; the reviewed Maputo-Gaza littoral does not show a distinct commercial forestry slot."),
    "Zambezi": validation_profile("regional_proxy", "partial_overlap", "distinct_slot_supported", 0.80, "Zambezi wood capacity is bounded to the Eastern Highlands plantation belt inside the broader state footprint rather than to a whole-Zimbabwe forestry fallback."),
    "Hereroland": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "Hereroland has no distinct commercial forestry slot; wooded savanna does not drive x."),
    "Namaqualand": validation_profile("state_localized", "direct", "broad_potential_only", 1.0, "Namaqualand has no distinct commercial forestry slot; arid wooded pockets do not drive x."),
}

REGIONAL_ADVANTAGE_SEEDS = [
    {
        "state": "Cape Colony",
        "advantage_type": "export_orientation",
        "primary_resources": "Arable Land; Fishing; Wood",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Cape Colony's port access and export orientation should stay outside raw caps; the cap model should not carry maritime-commercial advantage by itself.",
        "citation_1_title": "cape_economics.md",
        "citation_1_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_1_locator": "Cape economy combines wheat, wine, wool, fishing, and shipping-facing commerce.",
        "citation_2_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_2_url": str(AGRI_RESEARCH_PATH),
        "citation_2_locator": "Western Cape profile is diversified and export-oriented, not just a raw land-capacity story.",
    },
    {
        "state": "Northern Cape",
        "advantage_type": "discoverability_unlock",
        "primary_resources": "Lead Mine; Gold Fields; Arable Land",
        "preferred_layer": "discoverability_unlock",
        "keep_out_of_caps": "yes",
        "note": "Northern Cape's regional advantage is patchy discovery and enclave development rather than uniformly high raw capacity.",
        "citation_1_title": "cape_economics.md",
        "citation_1_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_1_locator": "Namaqualand mining is real but enclave-like rather than a broad province-wide economy.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Northern Cape analogues are arid, sparse, and unevenly developed.",
    },
    {
        "state": "Eastern Cape",
        "advantage_type": "resource_diversity",
        "primary_resources": "Arable Land; Wood; Fishing",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Eastern Cape's value is mixed-farming breadth plus coastal access; that breadth should be reinforced outside raw cap inflation.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Eastern Cape combines grain, stock, timber, and coastal production.",
        "citation_2_title": "cape_economics.md",
        "citation_2_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_2_locator": "Eastern Cape has a broader agrarian and timber mix than a single cap family can express.",
    },
    {
        "state": "West Transvaal",
        "advantage_type": "infrastructure_access",
        "primary_resources": "Gold Mine; Coal Mine; Arable Land",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Implement West Transvaal as a future state modifier in common/static_modifiers/sb_modifiers.txt, applied from Transvaal ownership logic in common/on_actions/sb_on_actions.txt once the west-plateau industrial core is consolidated; do not raise raw caps.",
        "citation_1_title": "Southern Africa comparator states ranked per region",
        "citation_1_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_1_locator": "Gauteng and the western plateau are urban-industrial and mining-linked.",
        "citation_2_title": "mining research.md",
        "citation_2_url": str(RESEARCH_DIR / "mining research.md"),
        "citation_2_locator": "Transvaal mineral development is tightly linked to industrial concentration and infrastructure.",
    },
    {
        "state": "Eastern Transvaal",
        "advantage_type": "climate_reliability",
        "primary_resources": "Arable Land; Wood; Coal Mine",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Eastern Transvaal's humid escarpment reliability and multiproduct structure should not be carried only by raw cap differences.",
        "citation_1_title": "Southern Africa comparator states ranked per region",
        "citation_1_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_1_locator": "Mpumalanga and Eswatini combine forestry, sugar, maize, and citrus across a humid gradient.",
        "citation_2_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_2_url": str(AGRI_RESEARCH_PATH),
        "citation_2_locator": "Eastern Transvaal profile is broad and climate-favored rather than just larger-capacity land.",
    },
    {
        "state": "Northern Transvaal",
        "advantage_type": "crop_diversity",
        "primary_resources": "Arable Land; Wood",
        "preferred_layer": "journal_flavour",
        "keep_out_of_caps": "yes",
        "note": "Northern Transvaal's warm subtropical mix matters more as crop diversity and frontier character than as large cap inflation.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Limpopo combines maize, tobacco, cotton, bananas, and livestock.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Limpopo is mixed-farming first rather than a single dominant resource belt.",
    },
    {
        "state": "Transorangia",
        "advantage_type": "grain_heartland",
        "primary_resources": "Arable Land",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Transorangia's interior grain-heartland identity should be expressed as efficiency/reliability rather than ever-higher cap values.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Free State is a wheat-maize-livestock heartland.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Free State analogues emphasize scale and consistency more than special crops.",
    },
    {
        "state": "Drakensberg",
        "advantage_type": "niche_pastoral_quality",
        "primary_resources": "Arable Land",
        "preferred_layer": "journal_flavour",
        "keep_out_of_caps": "yes",
        "note": "Implement Drakensberg through journal flavour on je_sb_settle_drakensberg / je_sb_bst_ora_frontier in common/journal_entries/1-07_sb_bst_frontier.txt, with mountain-pastoral flavour and any niche reward attached there rather than through higher raw caps.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Lesotho is a mountain grain-and-pastoral system with strong wool/mohair identity.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Drakensberg profile is constrained but distinctive.",
    },
    {
        "state": "Botswana",
        "advantage_type": "ranching_quality",
        "primary_resources": "Arable Land",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Implement Botswana as a future ranching-quality state modifier in common/static_modifiers/sb_modifiers.txt, applied from BST/Tswana content in common/on_actions/sb_bst_on_actions.txt or equivalent Botswana ownership logic; do not inflate raw caps.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Botswana is cattle-first with narrow cereal corridors.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Botswana analogues are semi-arid agropastoral systems.",
    },
    {
        "state": "Lourenço Marques",
        "advantage_type": "port_littoral_access",
        "primary_resources": "Arable Land; Fishing",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Lourenço Marques should benefit from littoral trade and port access outside the raw-cap model.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Southern Mozambique is coastal, riverine, and plantation-linked rather than just a land-capacity story.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Lourenço Marques analogues emphasize coastal mixed-crop access.",
    },
    {
        "state": "Zambezi",
        "advantage_type": "estate_crop_quality",
        "primary_resources": "Arable Land; Wood; Gold Fields (undiscovered)",
        "preferred_layer": "state_modifier",
        "keep_out_of_caps": "yes",
        "note": "Zambezi's plateau-and-highlands estate quality should not be represented only as more cap volume.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Zimbabwe combines tobacco, wheat, sugar, tea, coffee, cotton, maize, and livestock across distinct belts.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Zimbabwe analogues are broad and diversified rather than just cap-heavy.",
    },
    {
        "state": "Hereroland",
        "advantage_type": "enclave_mineral_quality",
        "primary_resources": "Lead Mine; Arable Land",
        "preferred_layer": "journal_flavour",
        "keep_out_of_caps": "yes",
        "note": "Hereroland's mining significance is enclave quality and district identity, not a uniformly rich state-wide endowment.",
        "citation_1_title": "mining research.md",
        "citation_1_url": str(RESEARCH_DIR / "mining research.md"),
        "citation_1_locator": "Tsumeb/Otavi and Berg Aukas are high-quality but concentrated districts.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Northern Namibia remains agropastoral outside concentrated mining nodes.",
    },
    {
        "state": "Namaqualand",
        "advantage_type": "enclave_irrigation",
        "primary_resources": "Arable Land; Lead Mine",
        "preferred_layer": "journal_flavour",
        "keep_out_of_caps": "yes",
        "note": "Namaqualand's advantage is enclave irrigation and niche mining, not broad land or forestry capacity.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Southern Namibia is sheep-and-goat country with narrow irrigated river pockets.",
        "citation_2_title": "cape_economics.md",
        "citation_2_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_2_locator": "Namaqualand-style economies are enclave and corridor based rather than broad-capacity systems.",
    },
]

STATE_REVIEW_STATUS_SEEDS = {
    "Cape Colony": ("accepted", "No major model blocker remains; only minor wording polish is still open.", "Core land-family rows are on a defensible model and the remaining work is limited to light presentation cleanup."),
    "Northern Cape": ("accepted", "Gameplay distinctiveness now sits outside the cap provenance pass.", "Northern Cape land rows are now localized to Orange-Vaal irrigation and Karoo grazing; remaining distinctiveness belongs to enclave/discoverability flavor rather than another cap audit pass."),
    "Eastern Cape": ("accepted", "No major model blocker remains; only minor wording polish is still open.", "Eastern Cape no longer fails basic plausibility and the remaining work is limited to light presentation cleanup."),
    "West Transvaal": ("needs_gameplay_compensation", "Compensation is assigned to a future state modifier in sb_modifiers.txt routed through sb_on_actions.txt.", "West Transvaal keeps its current caps; any extra industrial/infrastructure edge should be implemented as a state modifier applied through the existing Transvaal ownership/on-action flow."),
    "Eastern Transvaal": ("accepted", "Comparator and docs polish remain.", "Eastern Transvaal remains directionally strong and the remaining issues are mostly presentation and consistency."),
    "Northern Transvaal": ("accepted", "Gameplay distinctiveness now sits outside the cap provenance pass.", "Northern Transvaal land rows now rest on localized Limpopo evidence; the remaining difference is crop-diversity and frontier flavor outside raw caps."),
    "Transorangia": ("accepted", "No major model blocker remains; only minor wording polish is still open.", "Transorangia's current package is directionally coherent."),
    "Drakensberg": ("needs_gameplay_compensation", "Compensation is assigned to BST frontier journal flavour in 1-07_sb_bst_frontier.txt.", "Drakensberg stays constrained in caps; its mountain-pastoral niche should be expressed through the existing BST frontier / settlement journal flow rather than through more cap volume."),
    "Botswana": ("needs_gameplay_compensation", "Compensation is assigned to a future ranching-quality state modifier in sb_modifiers.txt.", "Botswana keeps its current caps; any extra cattle-first identity should be implemented through a state modifier applied from the BST/Tswana on-action path rather than through cap inflation."),
    "Lourenço Marques": ("accepted", "Port-littoral identity remains better expressed outside caps.", "Lourenço Marques now uses a bounded Maputo-Gaza lowland proxy for arable land and a defended forestry zero; remaining distinctiveness is littoral trade flavor outside the cap model."),
    "Zambezi": ("accepted", "Estate-quality flavor remains better expressed outside caps.", "Zambezi now uses Eastern Highlands and plateau-localized land evidence rather than a whole-Zimbabwe fallback; the remaining differentiation is estate-quality flavor outside raw caps."),
    "Hereroland": ("accepted", "Enclave-mining flavor remains better expressed outside caps.", "Hereroland now rests on localized northern Namibia land evidence and a defended forestry zero; the remaining distinction is concentrated mineral identity outside raw caps."),
    "Namaqualand": ("accepted", "No major model blocker remains; only minor wording polish is still open.", "Namaqualand's low-cap enclave profile is now directionally coherent."),
}
ARABLE_BASKET_FIELDNAMES = [
    "state",
    "resource",
    "researched_plausible",
    "basket_reason",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
]
ARABLE_BASKET_SEEDS = {
    "Cape Colony": {
        "yes_resources": {"Wheat Farm", "Vineyard", "Livestock Ranch", "Tobacco Plantation"},
        "include_summary": "Western Cape is the region's clearest wheat-and-wine state with a real livestock base; tobacco exists only in pockets.",
        "exclude_summary": "Western Cape is not part of the broad eastern cereal or subtropical plantation belts.",
        "citation_1_locator": "Western Cape row: wheat, wine, livestock strongly; tobacco only in pockets.",
        "citation_2_locator": "[^wc]: Western Cape is Mediterranean with strong wine, wheat, wool, lucerne, and fruit production.",
    },
    "Northern Cape": {
        "yes_resources": {"Cotton Plantation", "Livestock Ranch", "Maize Farm", "Millet Farm", "Tobacco Plantation", "Vineyard", "Wheat Farm"},
        "include_summary": "Northern Cape supports sheep and cattle, Orange-Vaal irrigation wheat and vineyards, and only narrow interior maize-millet-cotton pockets.",
        "exclude_summary": "Northern Cape is too dry for humid plantation crops and does not inherit the full North West cereal belt.",
        "citation_1_locator": "Northern Cape row: livestock, fishing, and whaling are strongest, with wheat and wine along irrigated rivers later.",
        "citation_2_locator": "Northern Cape is arid, with Orange irrigation, sheep/cattle, grapes, lucerne, cotton, citrus, and wheat concentrated in river corridors.",
    },
    "Eastern Cape": {
        "yes_resources": {"Livestock Ranch", "Maize Farm", "Tobacco Plantation", "Wheat Farm"},
        "include_summary": "Eastern Cape is a mixed grain-and-livestock state centered on maize, wheat, and stock, with only weaker tobacco pockets.",
        "exclude_summary": "Eastern Cape is not a Kaapse wine state or a full subtropical plantation belt in the 1836-1936 window.",
        "citation_1_locator": "Eastern Cape row: maize, wheat, livestock strong; tobacco weaker.",
        "citation_2_locator": "[^ec]: Eastern Cape has merino sheep, Angora goats, dairy cattle, inland wheat/maize/sorghum, and coastal tobacco/potatoes.",
    },
    "West Transvaal": {
        "yes_resources": {"Livestock Ranch", "Maize Farm", "Millet Farm", "Tobacco Plantation"},
        "include_summary": "The western Transvaal/Highveld grain belt is best read as maize, millet, livestock, with only a smaller tobacco allowance.",
        "exclude_summary": "This inland plateau is not a humid plantation or vineyard zone.",
        "citation_1_locator": "North West row: maize, millet, livestock strong; tobacco weaker. Gauteng row: maize and livestock strong; tobacco only small.",
        "citation_2_locator": "[^gau]: Gauteng is mainly urban but retains maize/sorghum/groundnut and dairy farming; [^ncnw] gives the adjacent North West maize-and-stock profile.",
    },
    "Eastern Transvaal": {
        "yes_resources": {"Banana Plantation", "Cotton Plantation", "Livestock Ranch", "Maize Farm", "Sugar Plantation", "Tea Plantation", "Tobacco Plantation"},
        "include_summary": "The Mpumalanga and Eswatini-facing belt supports maize, sugar, tobacco, bananas, livestock, and some cotton, with tea as a weaker pocket crop.",
        "exclude_summary": "This reviewed basket excludes cooler-grain and Mediterranean crops that are not central to the eastern lowveld-escarpment profile.",
        "citation_1_locator": "Mpumalanga row: maize, sugar, tobacco, bananas, livestock strong; tea weaker. Eswatini section: maize/cattle/cotton/tobacco profile is safest before 1936.",
        "citation_2_locator": "[^mpu]: Mpumalanga produces maize, cotton, sugarcane, tobacco, citrus, and subtropical fruits with beef/dairy/sheep.",
    },
    "Northern Transvaal": {
        "yes_resources": {"Banana Plantation", "Coffee Plantation", "Cotton Plantation", "Livestock Ranch", "Maize Farm", "Tobacco Plantation"},
        "include_summary": "Limpopo is a warm northern state where maize, tobacco, cotton, bananas, livestock, and some coffee fit better than cooler plateau crops.",
        "exclude_summary": "The reviewed northern basket excludes sugar-heavy and Mediterranean profiles outside Limpopo's core evidence.",
        "citation_1_locator": "Limpopo row: maize, tobacco, cotton, bananas, livestock strong; coffee weaker.",
        "citation_2_locator": "[^lim]: Limpopo combines citrus, maize, tea/tobacco/groundnuts, cattle, goats, and sheep in a warm subtropical-to-montane gradient.",
    },
    "Transorangia": {
        "yes_resources": {"Livestock Ranch", "Maize Farm", "Tobacco Plantation", "Wheat Farm"},
        "include_summary": "Free State is a classic interior maize-wheat-livestock state with only a weak tobacco allowance.",
        "exclude_summary": "The reviewed Free State basket excludes humid plantation crops and Mediterranean specials.",
        "citation_1_locator": "Free State row: maize, wheat, livestock strong; tobacco weaker.",
        "citation_2_locator": "[^fs]: Free State is a major maize and wheat province with strong cattle and sheep grazing.",
    },
    "Drakensberg": {
        "yes_resources": {"Livestock Ranch", "Maize Farm", "Millet Farm", "Wheat Farm"},
        "include_summary": "Lesotho is a grain-and-livestock highland where maize, sorghum, wheat, and stock are all plausible inside a mountain-constrained basket.",
        "exclude_summary": "The reviewed Drakensberg basket excludes plantation crops with no serious 1836-1936 tradition in Lesotho.",
        "citation_1_locator": "Lesotho section: maize, sorghum, wheat, and livestock define the state; mountains are far better as grazing land than plantation land.",
        "citation_2_locator": "[^les]: Lesotho has maize/wheat/sorghum and pastoral sheep/cattle/goats, especially wool and mohair.",
    },
    "Botswana": {
        "yes_resources": {"Livestock Ranch", "Maize Farm", "Millet Farm"},
        "include_summary": "Botswana is livestock-first, with only a narrow maize-and-millet cereal bonus in the northeast and eastern corridor.",
        "exclude_summary": "Botswana is not a breadbasket and does not justify wetter plantation or cool-climate crops as part of its main basket.",
        "citation_1_locator": "Botswana section: cattle first, with maize, sorghum, and millet only in the northeast and eastern corridor.",
        "citation_2_locator": "[^bot]: Botswana is drought-prone and dominated by cattle, goats, and crops such as maize and sorghum.",
    },
    "Lourenço Marques": {
        "yes_resources": {"Banana Plantation", "Cotton Plantation", "Livestock Ranch", "Maize Farm", "Millet Farm", "Rice Farm", "Sugar Plantation"},
        "include_summary": "The Maputo-Gaza littoral and lower Limpopo support maize, millet, rice, bananas, sugar, some cotton, and a smaller livestock component.",
        "exclude_summary": "This lowland belt excludes the highland tea/coffee pockets and does not justify a whole-Mozambique crop suite.",
        "citation_1_locator": "Mozambique section: the south coast and lower Limpopo suit maize, millet/sorghum, late-colonial rice, bananas, and limited sugar; colonial cotton also appears in parts of the lower Limpopo.",
        "citation_2_locator": "[^smz]: Mozambican family farming includes maize and rice, with plantation sugar and southward irrigation especially along the Limpopo.",
    },
    "Zambezi": {
        "yes_resources": {"Coffee Plantation", "Cotton Plantation", "Livestock Ranch", "Maize Farm", "Millet Farm", "Sugar Plantation", "Tea Plantation", "Tobacco Plantation", "Wheat Farm"},
        "include_summary": "Eastern Highlands and plateau Zimbabwe justify maize, tobacco, cotton, wheat, cattle, sugar, and pocket tea/coffee in a diversified but bounded basket.",
        "exclude_summary": "The reviewed basket is narrower than a whole-Zimbabwe all-belt claim and keeps estate crops to the humid east and lowveld pockets.",
        "citation_1_locator": "Zimbabwe section: Eastern Highlands, Mashonaland plateau, and southern lowveld belts separate tea/coffee, maize/tobacco, and sugar systems inside one reviewed state footprint.",
        "citation_2_locator": "Zimbabwe's elevation and climate support maize, tobacco, cattle, cotton, wheat, and sugarcane, with the wettest east carrying the strongest high-value estate pockets.",
    },
    "Hereroland": {
        "yes_resources": {"Livestock Ranch", "Maize Farm", "Millet Farm", "Wheat Farm"},
        "include_summary": "Northern and central Namibia are agropastoral rather than plantation-oriented: millet, maize, livestock, and a smaller wheat allowance in the Otavi-Grootfontein crop pocket.",
        "exclude_summary": "The reviewed Hereroland basket excludes wine, cotton, and humid plantation crops as non-core identities.",
        "citation_1_locator": "Namibia section: the northern communal belt is strongest in millet, sorghum, and maize; Otavi-Grootfontein allows some commercial maize/wheat; central highlands are mainly cattle.",
        "citation_2_locator": "Northern Namibia is a maize-millet-plus-cattle savanna belt with a small Otavi crop pocket rather than a broad cereal province.",
    },
    "Namaqualand": {
        "yes_resources": {"Livestock Ranch", "Vineyard"},
        "include_summary": "Southern Namibia is an arid sheep-and-goat state with only a narrow irrigated grape belt along the Orange system.",
        "exclude_summary": "The reviewed Namaqualand basket excludes broad cereals and humid plantation crops outside the irrigated river fringe.",
        "citation_1_locator": "Namibia section: southern and western desert-karoo zones are sheep/goat country, with agriculture surviving mainly in river-irrigated pockets.",
        "citation_2_locator": "[^snam]: Arid southern Namibia supports extensive livestock and lower Orange irrigation belts with fruit and grapes.",
    },
}

ARABLE_LAND_CLASSES = [
    ("irrigated_perennial", 1.25, "Irrigated or perennial commercial land."),
    ("reliable_rainfed_crop", 1.00, "Reliable rainfed crop land."),
    ("mixed_farming", 0.70, "Mixed crop-and-stock land."),
    ("commercial_pasture", 0.35, "Commercial grazing and ranch land."),
    ("marginal_grazing", 0.10, "Marginal semi-arid grazing land."),
    ("desert_or_unusable", 0.00, "Desert or commercially unusable land."),
]
ARABLE_LAND_CLASS_WEIGHT_MAP = {name: weight for name, weight, _note in ARABLE_LAND_CLASSES}
ARABLE_LAND_CLASS_FIELDNAMES = [
    "land_class",
    "default_weight",
    "weight_note",
]
ARABLE_TARGET_CAPACITY_FIELDNAMES = [
    "state",
    "land_class",
    "representative_year",
    "raw_area_ha",
    "effective_weight",
    "effective_area_ha",
    "year_selection_reason",
    "capacity_note",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
    *TARGET_VALIDATION_FIELDS,
]
ARABLE_COMPARATOR_CAPACITY_FIELDNAMES = [
    "target_state",
    "comparator_geography",
    "benchmark_state_id",
    "rank",
    "representative_year",
    "land_class",
    "raw_area_ha",
    "effective_weight",
    "effective_area_ha",
    "benchmark_vanilla_arable_cap",
    "year_selection_reason",
    "capacity_note",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
]

ARABLE_TARGET_CAPACITY_SEEDS = {
    "Cape Colony": {
        "classes": {
            "irrigated_perennial": 220_000,
            "reliable_rainfed_crop": 950_000,
            "mixed_farming": 1_100_000,
            "commercial_pasture": 700_000,
            "marginal_grazing": 600_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Broad Western Cape commercial potential: Mediterranean grain-and-vine core, strong mixed farming, and a moderate pastoral fringe.",
    },
    "Northern Cape": {
        "classes": {
            "irrigated_perennial": 54_000,
            "reliable_rainfed_crop": 108_000,
            "mixed_farming": 225_000,
            "commercial_pasture": 675_000,
            "marginal_grazing": 247_500,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Northern Cape is anchored on Orange-Vaal irrigation strips and extensive Karoo grazing; sparse interior mixed farming matters, but North West cereal intensity is excluded.",
    },
    "Eastern Cape": {
        "classes": {
            "irrigated_perennial": 0,
            "reliable_rainfed_crop": 1_250_000,
            "mixed_farming": 1_400_000,
            "commercial_pasture": 600_000,
            "marginal_grazing": 600_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Broad Eastern Cape and Transkei potential: major mixed-farming and rainfed grain land with strong grazing support, not a desert-like state.",
    },
    "West Transvaal": {
        "classes": {
            "irrigated_perennial": 0,
            "reliable_rainfed_crop": 350_000,
            "mixed_farming": 500_000,
            "commercial_pasture": 400_000,
            "marginal_grazing": 400_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Interior plateau with real cereal-and-livestock potential, but less broad than the eastern escarpment or the Free State grain belt.",
    },
    "Eastern Transvaal": {
        "classes": {
            "irrigated_perennial": 180_000,
            "reliable_rainfed_crop": 1_000_000,
            "mixed_farming": 900_000,
            "commercial_pasture": 600_000,
            "marginal_grazing": 150_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Escarpment-and-lowveld commercial belt with strong reliable crop land, sugar/tobacco/fruit pockets, and substantial mixed-farming acreage.",
    },
    "Northern Transvaal": {
        "classes": {
            "irrigated_perennial": 72_000,
            "reliable_rainfed_crop": 360_000,
            "mixed_farming": 450_000,
            "commercial_pasture": 450_000,
            "marginal_grazing": 675_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Northern Transvaal is anchored on the Limpopo mixed-farming belt: warm maize, tobacco, cotton, and banana country with broad supporting pasture, but not eastern-escarpment intensity throughout.",
    },
    "Transorangia": {
        "classes": {
            "irrigated_perennial": 0,
            "reliable_rainfed_crop": 700_000,
            "mixed_farming": 900_000,
            "commercial_pasture": 700_000,
            "marginal_grazing": 750_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Classic interior cereal-and-stock heartland with broad effective commercial land but less perennial intensity than the Cape or eastern lowveld.",
    },
    "Drakensberg": {
        "classes": {
            "irrigated_perennial": 0,
            "reliable_rainfed_crop": 120_000,
            "mixed_farming": 250_000,
            "commercial_pasture": 500_000,
            "marginal_grazing": 250_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Mountain-constrained grain-and-grazing state: real agricultural capacity, but strongly limited by terrain.",
    },
    "Botswana": {
        "classes": {
            "irrigated_perennial": 0,
            "reliable_rainfed_crop": 50_000,
            "mixed_farming": 50_000,
            "commercial_pasture": 500_000,
            "marginal_grazing": 150_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Livestock-first dryland state with only narrow cereal corridors; low cap is driven by land quality, not output poverty alone.",
    },
    "Lourenço Marques": {
        "classes": {
            "irrigated_perennial": 180_000,
            "reliable_rainfed_crop": 850_000,
            "mixed_farming": 1_000_000,
            "commercial_pasture": 450_000,
            "marginal_grazing": 175_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Maputo-Gaza littoral and lower-Limpopo lowland with strong mixed commercial potential across irrigated, rainfed, and coastal systems; this is not a whole-Mozambique land claim.",
    },
    "Zambezi": {
        "classes": {
            "irrigated_perennial": 120_000,
            "reliable_rainfed_crop": 650_000,
            "mixed_farming": 900_000,
            "commercial_pasture": 500_000,
            "marginal_grazing": 450_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Eastern Highlands and plateau-centered Zambezi mix with connected lowveld belts; the row is narrower than a whole-Zimbabwe fallback but still supports broad commercial agricultural land.",
    },
    "Hereroland": {
        "classes": {
            "irrigated_perennial": 0,
            "reliable_rainfed_crop": 36_000,
            "mixed_farming": 45_000,
            "commercial_pasture": 360_000,
            "marginal_grazing": 45_000,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Hereroland is anchored on the northern communal belt, the Otavi-Grootfontein crop pocket, and broad central-plateau cattle land; it is not a high-cap cereal state.",
    },
    "Namaqualand": {
        "classes": {
            "irrigated_perennial": 5_000,
            "reliable_rainfed_crop": 0,
            "mixed_farming": 10_000,
            "commercial_pasture": 50_000,
            "marginal_grazing": 242_500,
            "desert_or_unusable": 0,
        },
        "capacity_note": "Arid southern Namibian belt with only tiny irrigated river pockets and minimal broad commercial agricultural potential.",
    },
}

ARABLE_COMPARATOR_UNITS_PER_CAP_BY_STATE = {
    "Cape Colony": 53_000.0,
    "Northern Cape": 58_000.0,
    "Eastern Cape": 55_000.0,
    "West Transvaal": 55_000.0,
    "Eastern Transvaal": 54_000.0,
    "Northern Transvaal": 55_000.0,
    "Transorangia": 55_000.0,
    "Drakensberg": 50_000.0,
    "Botswana": 60_000.0,
    "Lourenço Marques": 54_000.0,
    "Zambezi": 55_000.0,
    "Hereroland": 60_000.0,
    "Namaqualand": 62_000.0,
}

WOOD_LAND_CLASSES = [
    ("high_suitability_plantation", 1.00, "High-suitability commercial plantation land."),
    ("moderate_suitability_plantation", 0.65, "Moderate-suitability commercial plantation land."),
    ("restorable_commercial_forest", 0.40, "Historically depleted but commercially restorable forestry land."),
    ("marginal_forestry", 0.15, "Marginal forestry land that can support only thin commercial activity."),
    ("noncommercial_wooded_land", 0.00, "Wooded land that is not meaningfully commercial forestry."),
    ("arid_or_unusable", 0.00, "Arid or otherwise unusable forestry land."),
]
WOOD_LAND_CLASS_WEIGHT_MAP = {name: weight for name, weight, _note in WOOD_LAND_CLASSES}
WOOD_LAND_CLASS_FIELDNAMES = [
    "land_class",
    "default_weight",
    "weight_note",
]
WOOD_TARGET_CAPACITY_FIELDNAMES = [
    "state",
    "land_class",
    "representative_year",
    "raw_area_ha",
    "effective_weight",
    "effective_area_ha",
    "year_selection_reason",
    "capacity_note",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
    *TARGET_VALIDATION_FIELDS,
]
WOOD_COMPARATOR_CAPACITY_FIELDNAMES = [
    "comparator_geography",
    "representative_year",
    "land_class",
    "raw_area_ha",
    "effective_weight",
    "effective_area_ha",
    "benchmark_vanilla_wood_cap",
    "year_selection_reason",
    "capacity_note",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
]
WOOD_TARGET_CAPACITY_SEEDS = {
    "Cape Colony": {
        "classes": {
            "high_suitability_plantation": 35_000,
            "moderate_suitability_plantation": 65_000,
            "restorable_commercial_forest": 90_000,
            "marginal_forestry": 40_000,
            "noncommercial_wooded_land": 0,
            "arid_or_unusable": 0,
        },
        "capacity_note": "Western Cape remains plantation-first, but depleted southern coastal forest belts justify a capped restoration allowance above the surviving estate footprint.",
        "citation_1_title": "Trade & Industrial Policy Strategies: South Africa’s Forestry Value Chain",
        "citation_1_url": "https://www.tips.org.za/policy-briefs/item/download/1951_b81dc50ab22218561eec628aa634d782",
        "citation_1_locator": "Western Cape afforested area = 43,146 ha in 2017; used here as the surviving managed-estate floor, not the full potential ceiling.",
        "citation_2_title": "cape_economics.md",
        "citation_2_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_2_locator": "Timber was scarce in the arid Cape interior, but the southern coastal mountains retained exploitable forest belts and sawmilling activity.",
    },
    "Northern Cape": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Dry interior and riverine fringes do not justify a meaningful commercial forestry belt in this audit.",
        "citation_1_title": "cape_economics.md",
        "citation_1_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_1_locator": "Timber was scarce in the arid parts of the colony; exploitable forests sat in the southern coastal mountains rather than the dry interior.",
        "citation_2_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_2_url": str(AGRI_RESEARCH_PATH),
        "citation_2_locator": "Northern Cape and adjacent dry belts are arid-to-semi-arid systems where irrigation and grazing matter far more than forestry.",
    },
    "Eastern Cape": {
        "classes": {
            "high_suitability_plantation": 95_000,
            "moderate_suitability_plantation": 140_000,
            "restorable_commercial_forest": 180_000,
            "marginal_forestry": 55_000,
            "noncommercial_wooded_land": 0,
            "arid_or_unusable": 0,
        },
        "capacity_note": "Eastern Cape combines the surviving timber belt with a capped restoration allowance for the historically depleted Tsitsikamma-Outeniqua forestry zone.",
        "citation_1_title": "Trade & Industrial Policy Strategies: South Africa’s Forestry Value Chain",
        "citation_1_url": "https://www.tips.org.za/policy-briefs/item/download/1951_b81dc50ab22218561eec628aa634d782",
        "citation_1_locator": "Eastern Cape afforested area = 141,812 ha in 2017; treated here as the surviving estate floor beneath a broader potential-capacity estimate.",
        "citation_2_title": "cape_economics.md",
        "citation_2_url": str(RESEARCH_DIR / "cape_economics.md"),
        "citation_2_locator": "The southern coastal mountains retained exploitable forests; the Eastern Cape timber economy is stronger than the western Cape interior.",
    },
    "West Transvaal": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Interior plateau grain-and-mining country without a serious commercial forestry identity in this audit.",
        "citation_1_title": "Southern Africa comparator states ranked per region",
        "citation_1_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_1_locator": "Gauteng and North West are defined primarily by maize, livestock, mining, and interior plateau farming rather than forestry belts.",
        "citation_2_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_2_url": str(AGRI_RESEARCH_PATH),
        "citation_2_locator": "North West and Gauteng profiles are inland cereal-and-stock systems, not plantation or forest economies.",
    },
    "Eastern Transvaal": {
        "classes": {
            "high_suitability_plantation": 180_000,
            "moderate_suitability_plantation": 220_000,
            "restorable_commercial_forest": 60_000,
            "marginal_forestry": 40_000,
            "noncommercial_wooded_land": 0,
            "arid_or_unusable": 0,
        },
        "capacity_note": "The wetter eastern escarpment remains the strongest commercial forestry state in-scope, with only a small restoration allowance above the surviving plantation estate.",
        "citation_1_title": "Report on Commercial Timber Resources and Primary Roundwood Processing in South Africa 2019/2020",
        "citation_1_url": "https://www.dffe.gov.za/sites/default/files/reports/research/forestry/annualreport_commercialtimberresources_primaryroundwoodprocessing2019to2020.pdf",
        "citation_1_locator": "Table 20: Mpumalanga plantation area in 2019/20 = 348,111 ha; used here as the surviving estate floor.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "[^mpu]: Mpumalanga and Eswatini combine sugarcane, citrus, forestry, maize, and cattle across a humid escarpment-to-lowveld gradient.",
    },
    "Northern Transvaal": {
        "classes": {
            "high_suitability_plantation": 7_500,
            "moderate_suitability_plantation": 13_500,
            "restorable_commercial_forest": 0,
            "marginal_forestry": 11_250,
            "noncommercial_wooded_land": 0,
            "arid_or_unusable": 0,
        },
        "capacity_note": "Limpopo's small plantation fringe around the eastern escarpment is enough to justify a token wood row, but not a broad forestry-state cap.",
        "citation_1_title": "Southern Africa comparator states ranked per region",
        "citation_1_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_1_locator": "Limpopo is a warm mixed-farming zone, not a major forestry province; any forestry row should stay small and escarpment-fringe rather than statewide.",
        "citation_2_title": "Report on Commercial Timber Resources and Primary Roundwood Processing in South Africa 2019/2020",
        "citation_2_url": "https://www.dffe.gov.za/sites/default/files/reports/research/forestry/annualreport_commercialtimberresources_primaryroundwoodprocessing2019to2020.pdf",
        "citation_2_locator": "Table 20 and Limpopo section: 32,589 ha of commercial timber plantations were recorded in Limpopo in 2019/20.",
    },
    "Transorangia": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Interior cereal-and-stock heartland without a reviewed commercial forestry block in this audit.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Free State is a major maize, wheat, cattle, and sheep province rather than a commercial forestry region.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Free State comparator profile emphasizes cereal-and-stock land use, not a forestry belt.",
    },
    "Drakensberg": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Mountain grazing country with patchy woodland, but no audited commercial forestry block strong enough to exit explicit-zero handling.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Lesotho is reviewed as a mountain grain-and-pastoral system rather than a forestry economy.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Lesotho’s reviewed profile is highland grazing and grain, not a timber belt.",
    },
    "Botswana": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Dry livestock-first country without a meaningful audited commercial forestry belt.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Botswana is drought-prone and cattle-first, with only narrow cereal corridors rather than plantation or forest systems.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Botswana’s reviewed profile is a semi-arid agropastoral belt, not a forestry state.",
    },
    "Lourenço Marques": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Maputo-Gaza and the lower Limpopo are audited here as crop-and-riverine lowland country rather than as a distinct commercial forestry province.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Southern Mozambique and the lower Limpopo are reviewed through maize, millet, rice, bananas, sugar, and cotton rather than forestry.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Lourenço Marques analogues are coastal mixed-crop and riverine systems; Mozambique-wide timber claims do not create a localized forestry slot here.",
    },
    "Zambezi": {
        "classes": {
            "high_suitability_plantation": 50_000,
            "moderate_suitability_plantation": 70_000,
            "restorable_commercial_forest": 50_000,
            "marginal_forestry": 20_000,
            "noncommercial_wooded_land": 0,
            "arid_or_unusable": 0,
        },
        "capacity_note": "Zambezi carries a moderate forestry row anchored on the Eastern Highlands plantation belt, not on a generic whole-Zimbabwe forestry claim.",
        "citation_1_title": "Forests and Climate Change Working Paper 12",
        "citation_1_url": "https://www.fao.org/4/i2970e/i2970e.pdf",
        "citation_1_locator": "Zimbabwe had about 108,000 ha of commercial plantations, concentrated in the wetter east.",
        "citation_2_title": "Zimbabwe Biodiversity Report",
        "citation_2_url": "https://www.awf.org/sites/default/files/2023-09/ZBE%20Report%20Final%20Copy%20-%20080923.pdf",
        "citation_2_locator": "About 90% of Zimbabwe's planted or exotic forests are in the Eastern Highlands, which localizes the forestry slot to the eastern state footprint rather than the whole country.",
    },
    "Hereroland": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Northern and central Namibian savanna-woodland does not translate into a commercial forestry belt in this audit.",
        "citation_1_title": "Southern Africa comparator states ranked per region",
        "citation_1_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_1_locator": "Northern Namibia is reviewed as a savanna-woodland agropastoral belt where cattle and grains dominate rather than forestry.",
        "citation_2_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_2_url": str(AGRI_RESEARCH_PATH),
        "citation_2_locator": "Northern Namibia is a maize/millet-plus-cattle savanna belt, not a commercial timber belt.",
    },
    "Namaqualand": {
        "classes": {land_class: 0.0 for land_class, _weight, _note in WOOD_LAND_CLASSES},
        "capacity_note": "Arid southern Namibian terrain and riverine pockets do not justify commercial forestry land in this pass.",
        "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
        "citation_1_url": str(AGRI_RESEARCH_PATH),
        "citation_1_locator": "Southern Namibia is reviewed as sheep-and-goat country with tiny irrigated pockets, not a forestry zone.",
        "citation_2_title": "Southern Africa comparator states ranked per region",
        "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
        "citation_2_locator": "Southern Namibia analogues are arid grazing systems, not commercial forestry states.",
    },
}
WOOD_COMPARATOR_PROFILE_SHARES = {
    "plantation_core": {
        "high_suitability_plantation": 0.55,
        "moderate_suitability_plantation": 0.30,
        "restorable_commercial_forest": 0.10,
        "marginal_forestry": 0.05,
        "noncommercial_wooded_land": 0.0,
        "arid_or_unusable": 0.0,
    },
    "mixed_plantation": {
        "high_suitability_plantation": 0.35,
        "moderate_suitability_plantation": 0.40,
        "restorable_commercial_forest": 0.15,
        "marginal_forestry": 0.10,
        "noncommercial_wooded_land": 0.0,
        "arid_or_unusable": 0.0,
    },
    "marginal_plantation": {
        "high_suitability_plantation": 0.20,
        "moderate_suitability_plantation": 0.35,
        "restorable_commercial_forest": 0.15,
        "marginal_forestry": 0.30,
        "noncommercial_wooded_land": 0.0,
        "arid_or_unusable": 0.0,
    },
}
WOOD_COMPARATOR_PROFILE_BY_BENCHMARK = {
    "Minas Gerais": "plantation_core",
    "Sao Paulo": "plantation_core",
    "Parana": "plantation_core",
    "Santa Catarina": "plantation_core",
    "Bahia": "mixed_plantation",
    "Maranhao": "marginal_plantation",
    "Mato Grosso": "marginal_plantation",
    "Para": "marginal_plantation",
    "Goias": "mixed_plantation",
    "Rio de Janeiro": "marginal_plantation",
    "Piaui": "marginal_plantation",
    "Araucania": "mixed_plantation",
    "Los Rios": "mixed_plantation",
    "New South Wales": "mixed_plantation",
    "Victoria": "mixed_plantation",
    "Queensland": "mixed_plantation",
    "Tasmania": "mixed_plantation",
    "Western Australia": "marginal_plantation",
    "Rio Grande do Sul": "mixed_plantation",
    "North Island": "plantation_core",
}


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))

def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def ensure_csv_schema(path: Path, fieldnames: list[str], seed_rows: list[dict[str, Any]] | None = None) -> None:
    if not path.exists():
        write_csv(path, seed_rows or [], fieldnames)
        return
    rows = read_csv_rows(path)
    current_fields = rows[0].keys() if rows else []
    if list(current_fields) == fieldnames:
        return
    write_csv(path, rows, fieldnames)


def parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(float(str(value)))


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(str(value))


def parse_numeric_or_text(value: Any) -> Any:
    if value in (None, ""):
        return ""
    text = str(value)
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return float(text)
    return text


def normalize_validation_fields(row: dict[str, Any], default_profile: dict[str, Any]) -> dict[str, Any]:
    evidence_scope = str(row.get("evidence_scope") or default_profile["evidence_scope"])
    target_match_status = str(row.get("target_match_status") or default_profile["target_match_status"])
    slot_support_status = str(row.get("slot_support_status") or default_profile["slot_support_status"])
    localization_discount = parse_float(row.get("localization_discount"))
    if localization_discount is None:
        localization_discount = float(default_profile["localization_discount"])
    validation_note = str(row.get("validation_note") or default_profile["validation_note"])
    return {
        "evidence_scope": evidence_scope,
        "target_match_status": target_match_status,
        "slot_support_status": slot_support_status,
        "localization_discount": localization_discount,
        "validation_note": validation_note,
    }


def validation_drives_x(validation: dict[str, Any], corroboration_count: int = 1) -> bool:
    if validation["slot_support_status"] != "distinct_slot_supported":
        return False
    if validation["evidence_scope"] == "state_localized":
        return True
    if validation["target_match_status"] == "indirect":
        return corroboration_count >= 2
    return True


def validation_review_action(validation: dict[str, Any], drives_x: bool, corroboration_count: int = 1) -> str:
    if validation["slot_support_status"] == "broad_potential_only":
        return "re-audit_or_keep_out_of_x"
    if validation["target_match_status"] == "indirect" and corroboration_count < 2:
        return "needs_corroboration"
    if validation["evidence_scope"] == "state_localized":
        return "accepted_direct"
    return "accepted_with_bounded_fallback" if drives_x else "review_required"


def state_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def resource_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def short_geography_name(value: str) -> str:
    return str(value).split("\n", 1)[0].strip()


def resource_family(resource: str) -> str:
    if resource.startswith("Gold Fields"):
        return "Gold Fields"
    if resource.startswith("Oil"):
        return "Oil"
    if resource.startswith("Rubber"):
        return "Rubber"
    return resource


def growth_family_for_resource(resource: str) -> str:
    if resource in LAND_ECONOMY_RESOURCES or resource == "Arable Land":
        return "Land Economy"
    if resource in {"Fishing", "Whaling"}:
        return "Fisheries / Marine"
    if resource == "Wood":
        return "Managed Forestry Area"
    if resource == "Gold Mine" or resource.startswith("Gold Fields"):
        return "Gold Fields"
    return resource_family(resource)


def canonical_unit(unit: str | None) -> str:
    if not unit:
        return ""
    lower = unit.lower()
    if lower.startswith("mt ") or lower.startswith("kt ") or lower.startswith("kg ") or "kg output" in lower:
        return "t output"
    if "tons" in lower or "t output" in lower:
        return "t output" if lower != "t gold" else "t gold"
    if lower.startswith("head"):
        return "head livestock"
    if "bales" in lower:
        return "bales"
    if "wine output" in lower or lower.startswith("l "):
        return "l output"
    if "acres" in lower:
        return "acres"
    if "ha managed forestry" in lower or lower == "ha managed forestry" or lower == "ha":
        return "ha managed forestry"
    if "skins" in lower or "pelts" in lower:
        return "skins"
    return unit


def standardized_quantity(quantity: float | int, unit: str | None) -> float:
    lower = (unit or "").lower()
    if lower.startswith("mt "):
        return float(quantity) * 1_000_000
    if lower.startswith("kt "):
        return float(quantity) * 1_000
    if lower.startswith("kg ") or "kg output" in lower:
        return float(quantity) / 1_000
    return float(quantity)


def conversion_key_for_row(resource: str, unit: str) -> str | None:
    lower = unit.lower()
    if resource in {"Wheat Farm", "Maize Farm", "Millet Farm", "Rice Farm"} and "t output" in lower:
        return "grain_t"
    if resource in {"Sugar Plantation", "Tobacco Plantation", "Banana Plantation", "Coffee Plantation", "Tea Plantation", "Dye Plantation", "Opium", "Silk Plantation"} and "t output" in lower:
        return "cashcrop_t"
    if resource == "Cotton Plantation":
        if "bales" in lower:
            return "cotton_bale"
        if "t output" in lower:
            return "cashcrop_t"
    if resource == "Vineyard":
        if "l wine output" in lower:
            return "wine_l"
        if "t output" in lower:
            return "cashcrop_t"
        if "acres" in lower:
            return "acres_farmland"
    if resource == "Livestock Ranch":
        if "head cattle" in lower:
            return "head_cattle"
        if "head sheep" in lower:
            return "head_sheep_goat"
        if "skins" in lower or "pelts" in lower:
            return "skins_pelts"
        if lower.startswith("head"):
            return "head_livestock_generic"
    return None


def normalize_to_1950(quantity: float, year: int, family: str, growth_index_lookup: dict[tuple[str, int], float]) -> float:
    return quantity * growth_index_lookup.get((family, year), 1.0)


def _interpolate_quantity(q1: float, y1: int, q2: float, y2: int, year: int) -> float:
    if y1 == y2:
        return q1
    if q1 > 0 and q2 > 0:
        annual_factor = (q2 / q1) ** (1 / (y2 - y1))
        return q1 * (annual_factor ** (year - y1))
    slope = (q2 - q1) / (y2 - y1)
    return q1 + slope * (year - y1)


def build_fallback_growth_series(rate: float, family: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for year in range(GROWTH_YEAR_MIN, GROWTH_YEAR_MAX + 1):
        quantity = (1 + rate) ** (year - 1950)
        rows.append(
            {
                "family": family,
                "year": year,
                "quantity": quantity,
                "unit": "index",
                "index_to_1950": 1 / quantity if quantity else "",
                "method": f"fallback constant annual rate {rate:.4f}",
                "source_title": "Fallback workbook assumption",
                "source_url": "",
                "locator": "Used only for families that still lack a published annual reference series.",
            }
        )
    return rows


def extract_state_block(text: str, state_id: str) -> str:
    match = re.search(rf"{state_id}\s*=\s*\{{", text)
    if not match:
        raise ValueError(f"Missing state block for {state_id}")
    start = match.end() - 1
    depth = 0
    for idx in range(start, len(text)):
        if text[idx] == "{":
            depth += 1
        elif text[idx] == "}":
            depth -= 1
            if depth == 0:
                return text[match.start():idx + 1]
    raise ValueError(f"Unclosed state block for {state_id}")


def parse_state_block_resources(block: str) -> dict[str, Any]:
    parsed: dict[str, Any] = {}
    arable_match = re.search(r"arable_land\s*=\s*(\d+)", block)
    parsed["Arable Land"] = int(arable_match.group(1)) if arable_match else 0
    arable_resources_match = re.search(r"arable_resources\s*=\s*\{([^}]*)\}", block, re.S)
    enabled = set(re.findall(r'"([^"]+)"', arable_resources_match.group(1))) if arable_resources_match else set()
    for resource, building in RESOURCE_TO_BUILDING.items():
        parsed[resource] = "yes" if building in enabled else "no"
    capped_match = re.search(r"capped_resources\s*=\s*\{(.*?)\n\s*\}", block, re.S)
    capped_block = capped_match.group(1) if capped_match else ""
    for resource, building in CAP_BUILDINGS.items():
        match = re.search(rf"{re.escape(building)}\s*=\s*(\d+)", capped_block)
        parsed[resource] = int(match.group(1)) if match else 0
    special_blocks = re.finditer(r"resource\s*=\s*\{(.*?)\n\s*\}", block, re.S)
    seen_specials = {name: {"discovered": 0, "undiscovered": 0} for name in SPECIAL_RESOURCE_CONFIG}
    for match in special_blocks:
        resource_block = match.group(1)
        type_match = re.search(r'type\s*=\s*"([^"]+)"', resource_block)
        if not type_match:
            continue
        block_type = type_match.group(1)
        for family, config in SPECIAL_RESOURCE_CONFIG.items():
            if config["type"] != block_type:
                continue
            discovered_match = re.search(r"(?<!un)discovered_amount\s*=\s*(\d+)", resource_block)
            undiscovered_match = re.search(r"undiscovered_amount\s*=\s*(\d+)", resource_block)
            seen_specials[family]["discovered"] = int(discovered_match.group(1)) if discovered_match else 0
            seen_specials[family]["undiscovered"] = int(undiscovered_match.group(1)) if undiscovered_match else 0
    for family, payload in seen_specials.items():
        parsed[f"{family} (discovered)"] = payload["discovered"]
        parsed[f"{family} (undiscovered)"] = payload["undiscovered"]
    return parsed


def parse_live_state_resources() -> dict[str, dict[str, Any]]:
    text = STATE_FILE.read_text(encoding="utf-8")
    return {row["official_name"]: parse_state_block_resources(extract_state_block(text, row["state_id"])) for row in STATE_INFO}


def load_raw_data() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    def read_anchor_csv(path: Path, observation_family: str) -> list[dict[str, Any]]:
        ensure_csv_schema(path, TARGET_OBSERVATION_RAW_FIELDNAMES)
        rows: list[dict[str, Any]] = []
        for row in read_csv_rows(path):
            state_info = STATE_INFO_BY_SHEET.get(row["sheet"])
            validation = (
                normalize_validation_fields(row, OBSERVATION_TARGET_VALIDATION_DEFAULTS_BY_STATE[state_info["official_name"]])
                if state_info is not None
                else {field: row.get(field, "") for field in TARGET_VALIDATION_FIELDS}
            )
            rows.append(
                {
                    "sheet": row["sheet"],
                    "geography": row["geography"],
                    "resource": row["resource"],
                    "year": parse_int(row["year"]),
                    "normalized_quantity": parse_float(row["normalized_quantity"]),
                    "normalized_unit": row["normalized_unit"],
                    "source_title": row["source_title"],
                    "source_url": row["source_url"],
                    "citation_locator": row["citation_locator"],
                    "note": row.get("note", ""),
                    "observation_family": observation_family,
                    **validation,
                }
            )
        write_csv(path, rows, TARGET_OBSERVATION_RAW_FIELDNAMES)
        return rows

    return read_anchor_csv(RAW_HISTORICAL_CSV, "historical"), read_anchor_csv(RAW_MODERN_CSV, "modern")


def load_growth_anchor_series() -> dict[str, list[dict[str, Any]]]:
    series: dict[str, list[dict[str, Any]]] = {}
    for row in read_csv_rows(RAW_GROWTH_CSV):
        family = str(row["family"])
        series.setdefault(family, []).append(
            {
                "year": parse_int(row["year"]),
                "quantity": parse_float(row["quantity"]),
                "unit": row["unit"],
                "source_title": row["source_title"],
                "source_url": row["source_url"],
                "locator": row["locator"],
            }
        )
    return series


def build_growth_rows(anchor_series: dict[str, list[dict[str, Any]]]) -> tuple[list[dict[str, Any]], dict[tuple[str, int], float]]:
    rows: list[dict[str, Any]] = []
    index_lookup: dict[tuple[str, int], float] = {}
    for family, family_rows in anchor_series.items():
        if not family_rows:
            continue
        anchors = sorted(family_rows, key=lambda row: int(row["year"]))
        by_year: dict[int, dict[str, Any]] = {}
        for row in anchors:
            year = int(row["year"])
            if year not in by_year or float(row["quantity"]) > float(by_year[year]["quantity"]):
                by_year[year] = row
        years = sorted(by_year)
        if not years:
            continue
        if len(years) == 1:
            q0 = float(by_year[years[0]]["quantity"])
            series = {year: q0 for year in range(GROWTH_YEAR_MIN, GROWTH_YEAR_MAX + 1)}
            seed_anchor = by_year[years[0]]
            for year in range(GROWTH_YEAR_MIN, GROWTH_YEAR_MAX + 1):
                by_year.setdefault(
                    year,
                    {
                        "year": year,
                        "quantity": q0,
                        "source_title": seed_anchor["source_title"],
                        "source_url": seed_anchor["source_url"],
                        "locator": "constant single-anchor series",
                        "unit": seed_anchor["unit"],
                    },
                )
        else:
            series: dict[int, float] = {}
            for year in range(GROWTH_YEAR_MIN, GROWTH_YEAR_MAX + 1):
                if year in by_year:
                    series[year] = float(by_year[year]["quantity"])
                    continue
                if year < years[0]:
                    hi = 1950 if 1950 in by_year and years[0] != 1950 else years[min(len(years) - 1, 1)]
                    lo = years[0]
                    method = f"geometric backcast from {lo}-{hi}"
                elif year > years[-1]:
                    lo = 1950 if 1950 in by_year and years[-1] != 1950 else years[max(0, len(years) - 2)]
                    hi = years[-1]
                    method = f"geometric forecast from {lo}-{hi}"
                else:
                    lo = max(anchor_year for anchor_year in years if anchor_year < year)
                    hi = min(anchor_year for anchor_year in years if anchor_year > year)
                    method = f"geometric interpolation between {lo}-{hi}"
                series[year] = _interpolate_quantity(float(by_year[lo]["quantity"]), lo, float(by_year[hi]["quantity"]), hi, year)
                by_year.setdefault(year, {"year": year, "quantity": series[year], "source_title": by_year[lo]["source_title"], "source_url": by_year[lo]["source_url"], "locator": method, "unit": by_year[lo]["unit"]})
        quantity_1950 = series.get(1950) or 1.0
        for year in range(GROWTH_YEAR_MIN, GROWTH_YEAR_MAX + 1):
            anchor = by_year[year]
            if year in years:
                method = "anchor"
            elif year < years[0]:
                hi = 1950 if 1950 in by_year and years[0] != 1950 else years[min(len(years) - 1, 1)]
                method = f"geometric backcast from {years[0]}-{hi}"
            elif year > years[-1]:
                lo = 1950 if 1950 in by_year and years[-1] != 1950 else years[max(0, len(years) - 2)]
                method = f"geometric forecast from {lo}-{years[-1]}"
            else:
                lo = max(anchor_year for anchor_year in years if anchor_year < year)
                hi = min(anchor_year for anchor_year in years if anchor_year > year)
                method = f"geometric interpolation between {lo}-{hi}"
            quantity = series[year]
            index = quantity_1950 / quantity if quantity else ""
            rows.append(
                {
                    "family": family,
                    "year": year,
                    "quantity": quantity,
                    "unit": anchor["unit"],
                    "index_to_1950": index,
                    "method": method,
                    "source_title": anchor["source_title"],
                    "source_url": anchor["source_url"],
                    "locator": anchor["locator"],
                }
            )
            if index != "":
                index_lookup[(family, year)] = float(index)
    for family, rate in FALLBACK_ANNUAL_GROWTH_RATES.items():
        for row in build_fallback_growth_series(rate, family):
            if (family, int(row["year"])) in index_lookup:
                continue
            rows.append(row)
            index_lookup[(family, int(row["year"]))] = float(row["index_to_1950"])
    rows.sort(key=lambda row: (row["family"], row["year"]))
    return rows, index_lookup


def load_benchmark_cases(path: Path = RAW_BENCHMARK_CSV) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(path):
        parsed = dict(row)
        for key in ["historical_year", "peak_year", "benchmark_vanilla_cap"]:
            parsed[key] = parse_int(parsed.get(key))
        for key in ["historical_quantity", "peak_quantity"]:
            parsed[key] = parse_float(parsed.get(key))
        rows.append(parsed)
    return rows


def parse_agricultural_rankings() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(RAW_RANKINGS_CSV):
        rows.append(
            {
                "sheet_key": row["sheet_key"],
                "official_state": row["official_state"],
                "rank": parse_int(row["rank"]),
                "band": row["band"],
                "weight": parse_int(row["weight"]),
                "comparator": row["comparator"],
                "country": row["country"],
                "why": row["why"],
                "matched_raw_geography": row["matched_raw_geography"],
                "proxy_state_id": row["proxy_state_id"],
                "vanilla_proxy_arable_land": parse_float(row["vanilla_proxy_arable_land"]),
            }
        )
    return rows


def load_vanilla_priors() -> dict[str, dict[str, Any]]:
    priors: dict[str, dict[str, Any]] = {row["official_name"]: {} for row in STATE_INFO}
    for row in read_csv_rows(RAW_VANILLA_PRIORS_CSV):
        if row["state"] not in priors:
            continue
        priors[row["state"]][row["resource"]] = parse_numeric_or_text(row["value"])
    return priors


def seed_arable_basket_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for state in [row["official_name"] for row in STATE_INFO]:
        profile = ARABLE_BASKET_SEEDS[state]
        yes_resources = profile["yes_resources"]
        for _category, resource in BINARY_RESOURCES:
            researched = "yes" if resource in yes_resources else "no"
            rows.append(
                {
                    "state": state,
                    "resource": resource,
                    "researched_plausible": researched,
                    "basket_reason": (
                        f"{resource} is inside the reviewed {state} basket: {profile['include_summary']}"
                        if researched == "yes"
                        else f"{resource} is outside the reviewed {state} basket: {profile['exclude_summary']}"
                    ),
                    "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
                    "citation_1_url": str(AGRI_RESEARCH_PATH),
                    "citation_1_locator": profile["citation_1_locator"],
                    "citation_2_title": "Southern Africa comparator states ranked per region",
                    "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
                    "citation_2_locator": profile["citation_2_locator"],
                }
            )
    return rows


def load_arable_baskets() -> dict[str, dict[str, str]]:
    ensure_csv_schema(RAW_ARABLE_BASKETS_CSV, ARABLE_BASKET_FIELDNAMES, seed_arable_basket_rows())
    baskets: dict[str, dict[str, str]] = {row["official_name"]: {} for row in STATE_INFO}
    for row in read_csv_rows(RAW_ARABLE_BASKETS_CSV):
        if row["state"] not in baskets:
            continue
        baskets[row["state"]][row["resource"]] = row["researched_plausible"]
    return baskets


def seed_arable_land_class_weights_rows() -> list[dict[str, Any]]:
    return [
        {
            "land_class": land_class,
            "default_weight": weight,
            "weight_note": note,
        }
        for land_class, weight, note in ARABLE_LAND_CLASSES
    ]


def load_arable_land_class_weights() -> dict[str, float]:
    ensure_csv_schema(RAW_ARABLE_LAND_WEIGHTS_CSV, ARABLE_LAND_CLASS_FIELDNAMES, seed_arable_land_class_weights_rows())
    return {
        row["land_class"]: parse_float(row["default_weight"]) or 0.0
        for row in read_csv_rows(RAW_ARABLE_LAND_WEIGHTS_CSV)
    }


def seed_arable_target_capacity_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for state, payload in ARABLE_TARGET_CAPACITY_SEEDS.items():
        basket_profile = ARABLE_BASKET_SEEDS[state]
        validation = ARABLE_TARGET_VALIDATION_DEFAULTS_BY_STATE[state]
        for land_class, weight, _note in ARABLE_LAND_CLASSES:
            raw_area = float(payload["classes"].get(land_class, 0.0))
            localization_discount = float(validation["localization_discount"])
            rows.append(
                {
                    "state": state,
                    "land_class": land_class,
                    "representative_year": 1936,
                    "raw_area_ha": raw_area,
                    "effective_weight": weight,
                    "effective_area_ha": raw_area * weight * localization_discount,
                    "year_selection_reason": "Arable land capacity is audited as broad plausible 1936 commercial potential rather than GDP-selected realized output.",
                    "capacity_note": payload["capacity_note"],
                    "citation_1_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
                    "citation_1_url": str(AGRI_RESEARCH_PATH),
                    "citation_1_locator": basket_profile["citation_1_locator"],
                    "citation_2_title": "Southern Africa comparator states ranked per region",
                    "citation_2_url": str(COMPARATOR_RESEARCH_PATH),
                    "citation_2_locator": basket_profile["citation_2_locator"],
                    **validation,
                }
            )
    return rows


def load_arable_target_capacity_rows() -> list[dict[str, Any]]:
    ensure_csv_schema(RAW_ARABLE_TARGET_CAPACITY_CSV, ARABLE_TARGET_CAPACITY_FIELDNAMES, seed_arable_target_capacity_rows())
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(RAW_ARABLE_TARGET_CAPACITY_CSV):
        validation = normalize_validation_fields(row, ARABLE_TARGET_VALIDATION_DEFAULTS_BY_STATE[row["state"]])
        raw_area = parse_float(row["raw_area_ha"]) or 0.0
        effective_weight = parse_float(row["effective_weight"]) or 0.0
        effective_area = raw_area * effective_weight * float(validation["localization_discount"])
        rows.append(
            {
                "state": row["state"],
                "land_class": row["land_class"],
                "representative_year": parse_int(row["representative_year"]),
                "raw_area_ha": raw_area,
                "effective_weight": effective_weight,
                "effective_area_ha": effective_area,
                "year_selection_reason": row["year_selection_reason"],
                "capacity_note": row["capacity_note"],
                "citation_1_title": row["citation_1_title"],
                "citation_1_url": row["citation_1_url"],
                "citation_1_locator": row["citation_1_locator"],
                "citation_2_title": row["citation_2_title"],
                "citation_2_url": row["citation_2_url"],
                "citation_2_locator": row["citation_2_locator"],
                **validation,
            }
        )
    write_csv(RAW_ARABLE_TARGET_CAPACITY_CSV, rows, ARABLE_TARGET_CAPACITY_FIELDNAMES)
    return rows


def _arable_target_effective_mix(state: str) -> dict[str, float]:
    payload = ARABLE_TARGET_CAPACITY_SEEDS[state]
    effective_by_class = {
        land_class: float(payload["classes"].get(land_class, 0.0)) * ARABLE_LAND_CLASS_WEIGHT_MAP[land_class]
        for land_class, _weight, _note in ARABLE_LAND_CLASSES
    }
    total = sum(effective_by_class.values())
    if total == 0:
        return {land_class: 0.0 for land_class, _weight, _note in ARABLE_LAND_CLASSES}
    return {land_class: value / total for land_class, value in effective_by_class.items()}


def seed_arable_comparator_capacity_rows(rankings: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    ranking_rows = rankings or parse_agricultural_rankings()
    rows: list[dict[str, Any]] = []
    for rank_row in ranking_rows:
        target_state = rank_row["official_state"]
        comparator = rank_row["matched_raw_geography"]
        benchmark_cap = parse_float(rank_row["vanilla_proxy_arable_land"]) or 0.0
        if not comparator or benchmark_cap <= 0:
            continue
        target_mix = _arable_target_effective_mix(target_state)
        units_per_cap = ARABLE_COMPARATOR_UNITS_PER_CAP_BY_STATE[target_state]
        total_effective = benchmark_cap * units_per_cap
        for land_class, weight, _note in ARABLE_LAND_CLASSES:
            effective_area = total_effective * target_mix[land_class]
            raw_area = (effective_area / weight) if weight else 0.0
            rows.append(
                {
                    "target_state": target_state,
                    "comparator_geography": comparator,
                    "benchmark_state_id": rank_row["proxy_state_id"],
                    "rank": rank_row["rank"],
                    "representative_year": 1936,
                    "land_class": land_class,
                    "raw_area_ha": raw_area,
                    "effective_weight": weight,
                    "effective_area_ha": effective_area,
                    "benchmark_vanilla_arable_cap": benchmark_cap,
                    "year_selection_reason": "Comparator land capacity is benchmarked as broad potential effective hectares per vanilla arable cap rather than GDP-selected crop output.",
                    "capacity_note": f"Benchmark analogue for {target_state}; effective hectares are scaled from the reviewed target-state land profile and the comparator vanilla arable cap.",
                    "citation_1_title": "Southern Africa comparator states ranked per region",
                    "citation_1_url": str(COMPARATOR_RESEARCH_PATH),
                    "citation_1_locator": f"{target_state} comparator ranking row {rank_row['rank']}: {rank_row['why']}",
                    "citation_2_title": "Zuidelijk Afrika als landbouw- en visserijregio voor een Victoria 3-mod",
                    "citation_2_url": str(AGRI_RESEARCH_PATH),
                    "citation_2_locator": ARABLE_BASKET_SEEDS[target_state]["citation_1_locator"],
                }
            )
    return rows


def load_arable_comparator_capacity_rows(rankings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ensure_csv_schema(
        RAW_ARABLE_COMPARATOR_CAPACITY_CSV,
        ARABLE_COMPARATOR_CAPACITY_FIELDNAMES,
        seed_arable_comparator_capacity_rows(rankings),
    )
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(RAW_ARABLE_COMPARATOR_CAPACITY_CSV):
        rows.append(
            {
                "target_state": row["target_state"],
                "comparator_geography": row["comparator_geography"],
                "benchmark_state_id": row["benchmark_state_id"],
                "rank": parse_int(row["rank"]),
                "representative_year": parse_int(row["representative_year"]),
                "land_class": row["land_class"],
                "raw_area_ha": parse_float(row["raw_area_ha"]) or 0.0,
                "effective_weight": parse_float(row["effective_weight"]) or 0.0,
                "effective_area_ha": parse_float(row["effective_area_ha"]) or 0.0,
                "benchmark_vanilla_arable_cap": parse_float(row["benchmark_vanilla_arable_cap"]) or 0.0,
                "year_selection_reason": row["year_selection_reason"],
                "capacity_note": row["capacity_note"],
                "citation_1_title": row["citation_1_title"],
                "citation_1_url": row["citation_1_url"],
                "citation_1_locator": row["citation_1_locator"],
                "citation_2_title": row["citation_2_title"],
                "citation_2_url": row["citation_2_url"],
                "citation_2_locator": row["citation_2_locator"],
            }
        )
    return rows


def seed_wood_land_class_weights_rows() -> list[dict[str, Any]]:
    return [
        {
            "land_class": land_class,
            "default_weight": weight,
            "weight_note": note,
        }
        for land_class, weight, note in WOOD_LAND_CLASSES
    ]


def load_wood_land_class_weights() -> dict[str, float]:
    ensure_csv_schema(RAW_WOOD_LAND_WEIGHTS_CSV, WOOD_LAND_CLASS_FIELDNAMES, seed_wood_land_class_weights_rows())
    return {
        row["land_class"]: parse_float(row["default_weight"]) or 0.0
        for row in read_csv_rows(RAW_WOOD_LAND_WEIGHTS_CSV)
    }


def _validate_wood_restoration_cap(rows: list[dict[str, Any]], group_field: str) -> None:
    grouped: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in rows:
        grouped[str(row[group_field])][str(row["land_class"])] += float(row["effective_area_ha"])
    for group, totals in grouped.items():
        plantation_effective = (
            totals.get("high_suitability_plantation", 0.0)
            + totals.get("moderate_suitability_plantation", 0.0)
        )
        restoration_effective = totals.get("restorable_commercial_forest", 0.0)
        if restoration_effective > (plantation_effective * 0.5) + 1e-6:
            raise ValueError(
                f"Wood restoration allowance exceeds the 50% cap for {group}: "
                f"restoration={restoration_effective}, plantation={plantation_effective}"
            )


def seed_wood_target_capacity_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for state, payload in WOOD_TARGET_CAPACITY_SEEDS.items():
        validation = WOOD_TARGET_VALIDATION_DEFAULTS_BY_STATE[state]
        for land_class, weight, _note in WOOD_LAND_CLASSES:
            raw_area = float(payload["classes"].get(land_class, 0.0))
            localization_discount = float(validation["localization_discount"])
            rows.append(
                {
                    "state": state,
                    "land_class": land_class,
                    "representative_year": 1936,
                    "raw_area_ha": raw_area,
                    "effective_weight": weight,
                    "effective_area_ha": raw_area * weight * localization_discount,
                    "year_selection_reason": "Wood is audited as potential effective commercial forestry land; representative year is metadata only and not GDP-selected.",
                    "capacity_note": payload["capacity_note"],
                    "citation_1_title": payload["citation_1_title"],
                    "citation_1_url": payload["citation_1_url"],
                    "citation_1_locator": payload["citation_1_locator"],
                    "citation_2_title": payload["citation_2_title"],
                    "citation_2_url": payload["citation_2_url"],
                    "citation_2_locator": payload["citation_2_locator"],
                    **validation,
                }
            )
    _validate_wood_restoration_cap(rows, "state")
    return rows


def load_wood_target_capacity_rows() -> list[dict[str, Any]]:
    ensure_csv_schema(RAW_WOOD_TARGET_CAPACITY_CSV, WOOD_TARGET_CAPACITY_FIELDNAMES, seed_wood_target_capacity_rows())
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(RAW_WOOD_TARGET_CAPACITY_CSV):
        validation = normalize_validation_fields(row, WOOD_TARGET_VALIDATION_DEFAULTS_BY_STATE[row["state"]])
        raw_area = parse_float(row["raw_area_ha"]) or 0.0
        effective_weight = parse_float(row["effective_weight"]) or 0.0
        effective_area = raw_area * effective_weight * float(validation["localization_discount"])
        rows.append(
            {
                "state": row["state"],
                "land_class": row["land_class"],
                "representative_year": parse_int(row["representative_year"]),
                "raw_area_ha": raw_area,
                "effective_weight": effective_weight,
                "effective_area_ha": effective_area,
                "year_selection_reason": row["year_selection_reason"],
                "capacity_note": row["capacity_note"],
                "citation_1_title": row["citation_1_title"],
                "citation_1_url": row["citation_1_url"],
                "citation_1_locator": row["citation_1_locator"],
                "citation_2_title": row["citation_2_title"],
                "citation_2_url": row["citation_2_url"],
                "citation_2_locator": row["citation_2_locator"],
                **validation,
            }
        )
    _validate_wood_restoration_cap(rows, "state")
    write_csv(RAW_WOOD_TARGET_CAPACITY_CSV, rows, WOOD_TARGET_CAPACITY_FIELDNAMES)
    return rows


def seed_wood_comparator_capacity_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for case in read_csv_rows(RAW_WOOD_COMPARATOR_CSV):
        benchmark = case["benchmark_state"]
        profile_name = WOOD_COMPARATOR_PROFILE_BY_BENCHMARK.get(benchmark, "mixed_plantation")
        shares = WOOD_COMPARATOR_PROFILE_SHARES[profile_name]
        total_raw = parse_float(case["historical_quantity"]) or 0.0
        vanilla_cap = parse_float(case["benchmark_vanilla_cap"]) or 0.0
        for land_class, weight, _note in WOOD_LAND_CLASSES:
            raw_area = total_raw * shares.get(land_class, 0.0)
            rows.append(
                {
                    "comparator_geography": benchmark,
                    "representative_year": 1936,
                    "land_class": land_class,
                    "raw_area_ha": raw_area,
                    "effective_weight": weight,
                    "effective_area_ha": raw_area * weight,
                    "benchmark_vanilla_wood_cap": vanilla_cap,
                    "year_selection_reason": "Wood comparator capacity is benchmarked as potential effective commercial forestry land rather than GDP-selected plantation-estate observations.",
                    "capacity_note": f"{case['comment']} Profile `{profile_name}` converts the observed commercial estate into a plantation-first potential-capacity mix with a capped restoration allowance.",
                    "citation_1_title": case["source_a_title"],
                    "citation_1_url": case["source_a_url"],
                    "citation_1_locator": case["source_a_locator"],
                    "citation_2_title": case["source_b_title"],
                    "citation_2_url": case["source_b_url"],
                    "citation_2_locator": case["source_b_locator"],
                }
            )
    _validate_wood_restoration_cap(rows, "comparator_geography")
    return rows


def load_wood_comparator_capacity_rows() -> list[dict[str, Any]]:
    ensure_csv_schema(
        RAW_WOOD_COMPARATOR_CAPACITY_CSV,
        WOOD_COMPARATOR_CAPACITY_FIELDNAMES,
        seed_wood_comparator_capacity_rows(),
    )
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(RAW_WOOD_COMPARATOR_CAPACITY_CSV):
        rows.append(
            {
                "comparator_geography": row["comparator_geography"],
                "representative_year": parse_int(row["representative_year"]),
                "land_class": row["land_class"],
                "raw_area_ha": parse_float(row["raw_area_ha"]) or 0.0,
                "effective_weight": parse_float(row["effective_weight"]) or 0.0,
                "effective_area_ha": parse_float(row["effective_area_ha"]) or 0.0,
                "benchmark_vanilla_wood_cap": parse_float(row["benchmark_vanilla_wood_cap"]) or 0.0,
                "year_selection_reason": row["year_selection_reason"],
                "capacity_note": row["capacity_note"],
                "citation_1_title": row["citation_1_title"],
                "citation_1_url": row["citation_1_url"],
                "citation_1_locator": row["citation_1_locator"],
                "citation_2_title": row["citation_2_title"],
                "citation_2_url": row["citation_2_url"],
                "citation_2_locator": row["citation_2_locator"],
            }
        )
    _validate_wood_restoration_cap(rows, "comparator_geography")
    return rows


def compute_wood_state_payloads(target_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    payloads: dict[str, dict[str, Any]] = {
        info["official_name"]: {
            "observed_output_x": 0.0,
            "restoration_effective_ha": 0.0,
            "plantation_effective_ha": 0.0,
            "restoration_cap_effective_ha": 0.0,
            "capacity_note": "",
        }
        for info in STATE_INFO
    }
    for row in target_rows:
        payload = payloads[row["state"]]
        effective_area = float(row["effective_area_ha"])
        payload["observed_output_x"] += effective_area
        if row["land_class"] in {"high_suitability_plantation", "moderate_suitability_plantation"}:
            payload["plantation_effective_ha"] += effective_area
        elif row["land_class"] == "restorable_commercial_forest":
            payload["restoration_effective_ha"] += effective_area
        if row["capacity_note"] and not payload["capacity_note"]:
            payload["capacity_note"] = row["capacity_note"]
    for payload in payloads.values():
        payload["restoration_cap_effective_ha"] = payload["plantation_effective_ha"] * 0.5
    return payloads


def compute_wood_denominator(
    comparator_rows: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    totals: dict[str, dict[str, Any]] = {}
    for row in comparator_rows:
        comparator = row["comparator_geography"]
        payload = totals.setdefault(
            comparator,
            {
                "comparator_geography": comparator,
                "representative_year": row["representative_year"],
                "benchmark_vanilla_wood_cap": row["benchmark_vanilla_wood_cap"],
                "effective_commercial_forestry_hectares": 0.0,
                "year_selection_reason": row["year_selection_reason"],
                "capacity_note": row["capacity_note"],
            },
        )
        payload["effective_commercial_forestry_hectares"] += float(row["effective_area_ha"])
    summary_rows: list[dict[str, Any]] = []
    per_cap_values: list[float] = []
    for payload in totals.values():
        vanilla_cap = float(payload["benchmark_vanilla_wood_cap"])
        effective_per_cap = payload["effective_commercial_forestry_hectares"] / vanilla_cap if vanilla_cap else 0.0
        payload["effective_hectares_per_cap"] = effective_per_cap
        summary_rows.append(payload)
        if effective_per_cap:
            per_cap_values.append(effective_per_cap)
    denominator = (sum(per_cap_values) / len(per_cap_values)) if per_cap_values else None
    denominator_row = {
        "resource_family": "Wood",
        "valid_comparator_count": len(per_cap_values),
        "denominator_units_per_cap": denominator,
        "status": "formula-driven" if len(per_cap_values) >= 20 else "denominator_unavailable",
        "method": (
            "mean_of_effective_commercial_forestry_comparator_pool"
            if len(per_cap_values) >= 20
            else "insufficient_comparator_pool"
        ),
    }
    summary_rows.sort(key=lambda row: row["comparator_geography"])
    return denominator_row, summary_rows


def load_adjusted_vanilla_totals(vanilla_priors: dict[str, dict[str, Any]]) -> dict[str, dict[str, int]]:
    proxy_by_state = {row["official_name"]: row["vanilla_proxy_id"] for row in STATE_INFO}
    result: dict[str, dict[str, int]] = {}
    for region, states in REGIONS.items():
        seen_proxies: set[str] = set()
        unique_states: list[str] = []
        for state in states:
            proxy = proxy_by_state[state]
            if proxy in seen_proxies:
                continue
            seen_proxies.add(proxy)
            unique_states.append(state)
        result[region] = {
            resource: sum(int(vanilla_priors[state].get(resource, 0) or 0) for state in unique_states)
            for resource in SUMMARY_RESOURCES
        }
    return result


def load_counterevidence_cases() -> dict[tuple[str, str], list[dict[str, str]]]:
    cases: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in read_csv_rows(RAW_COUNTEREVIDENCE_CSV):
        cases[(row["state"], row["resource"])].append(row)
    return dict(cases)


def load_family_normalising_constants() -> list[dict[str, Any]]:
    fieldnames = [
        "family",
        "value",
        "purpose",
        "citation_1_title",
        "citation_1_url",
        "citation_1_locator",
        "citation_2_title",
        "citation_2_url",
        "citation_2_locator",
    ]
    ensure_csv_schema(RAW_FAMILY_CONSTANTS_CSV, fieldnames)
    rows: list[dict[str, Any]] = []
    for row in read_csv_rows(RAW_FAMILY_CONSTANTS_CSV):
        rows.append(
            {
                "family": row["family"],
                "value": parse_float(row.get("value")) if row.get("value") not in (None, "") else "",
                "purpose": row.get("purpose", ""),
                "citation_1_title": row.get("citation_1_title", ""),
                "citation_1_url": row.get("citation_1_url", ""),
                "citation_1_locator": row.get("citation_1_locator", ""),
                "citation_2_title": row.get("citation_2_title", ""),
                "citation_2_url": row.get("citation_2_url", ""),
                "citation_2_locator": row.get("citation_2_locator", ""),
            }
        )
    return rows


def load_gdp_reference_anchor() -> dict[str, Any]:
    rows = read_csv_rows(RAW_GDP_REFERENCE_CSV)
    if len(rows) != 1:
        raise ValueError(f"Expected exactly one GDP reference row in {RAW_GDP_REFERENCE_CSV}")
    row = rows[0]
    return {
        "reference_geography": row["reference_geography"],
        "reference_year": parse_int(row["reference_year"]),
        "gdp_per_capita": parse_float(row["gdp_per_capita"]),
        "unit": row["unit"],
        "source_title": row.get("source_title", ""),
        "source_url": row.get("source_url", ""),
        "locator": row.get("locator", ""),
    }


def load_gdp_series() -> dict[str, dict[str, Any]]:
    payloads: dict[str, dict[str, Any]] = {}
    for row in read_csv_rows(RAW_GDP_SERIES_CSV):
        geography = row["geography"]
        payload = payloads.setdefault(
            geography,
            {
                "geography_level": row["geography_level"],
                "parent_country": row["parent_country"],
                "series": {},
                "source_title": row.get("source_title", ""),
                "source_url": row.get("source_url", ""),
                "locator": row.get("locator", ""),
            },
        )
        year = parse_int(row["year"])
        value = parse_float(row["gdp_per_capita"])
        if year is None or value is None:
            continue
        payload["series"][year] = value
    return payloads


def load_gdp_geography_map() -> dict[tuple[str, str], dict[str, str]]:
    mapping: dict[tuple[str, str], dict[str, str]] = {}
    for row in read_csv_rows(RAW_GDP_MAP_CSV):
        mapping[(row["entity_type"], row["entity"])] = row
    return mapping


def _series_value_for_year(series: dict[int, float], year: int) -> float | None:
    return series.get(year)


def _anchor_year_for_series(series: dict[int, float], reference_value: float) -> tuple[int | None, float | None, str]:
    if not series:
        return None, None, "manual_override"
    year, value = min(
        series.items(),
        key=lambda item: (abs(item[1] - reference_value), abs(item[0] - 1950), item[0]),
    )
    if max(series.values()) < reference_value:
        peak_year, peak_value = max(series.items(), key=lambda item: (item[1], item[0]))
        return peak_year, peak_value, "economic_peak_fallback"
    return year, value, "gdp_match"


def _build_selection_stub(
    entity: str,
    resource: str,
    map_row: dict[str, str] | None,
    reference_value: float,
    selected_year: Any,
    selection_mode: str,
    selection_note: str,
) -> dict[str, Any]:
    return {
        "entity": entity,
        "resource": resource,
        "gdp_geography_used": (
            map_row.get("preferred_substate_geography")
            or map_row.get("fallback_national_geography")
            if map_row
            else ""
        ),
        "gdp_geography_level": "",
        "gdp_reference_value": reference_value,
        "gdp_anchor_year": "",
        "gdp_value_at_anchor_year": "",
        "selected_resource_year": selected_year,
        "selection_mode": selection_mode,
        "gdp_distance_ratio": "",
        "gdp_comparability_status": "",
        "target_estimation_method": "",
        "interpolation_used": "no",
        "comparator_included_in_denominator": "",
        "comparator_basket_coverage_share": "",
        "selection_note": selection_note,
    }


def _manual_candidate_fallback(candidates: list[dict[str, Any]], reference_year: int) -> dict[str, Any] | None:
    if not candidates:
        return None
    return min(candidates, key=lambda candidate: (abs(int(candidate["year"]) - reference_year), int(candidate["year"])))


def select_year_by_gdp(
    entity: str,
    entity_type: str,
    resource: str,
    candidates: list[dict[str, Any]],
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> dict[str, Any]:
    map_row = gdp_map.get((entity_type, entity))
    reference_value = float(gdp_reference["gdp_per_capita"] or 0.0)
    reference_year = int(gdp_reference.get("reference_year") or 1950)
    if map_row is None:
        fallback_candidate = _manual_candidate_fallback(candidates, reference_year)
        selection = _build_selection_stub(
            entity,
            resource,
            None,
            reference_value,
            fallback_candidate["year"] if fallback_candidate is not None else "",
            "manual_override",
            "No GDP geography mapping exists for this entity; representative year stayed manual.",
        )
        selection["candidate"] = fallback_candidate
        return selection

    preferred = map_row.get("preferred_substate_geography", "")
    fallback = map_row.get("fallback_national_geography", "")
    geography_used = ""
    geography_payload: dict[str, Any] | None = None
    if preferred and preferred in gdp_series:
        geography_used = preferred
        geography_payload = gdp_series[preferred]
    elif fallback and fallback in gdp_series:
        geography_used = fallback
        geography_payload = gdp_series[fallback]

    if geography_payload is None:
        fallback_candidate = _manual_candidate_fallback(candidates, reference_year)
        selection = _build_selection_stub(
            entity,
            resource,
            map_row,
            reference_value,
            fallback_candidate["year"] if fallback_candidate is not None else "",
            "manual_override",
            map_row.get("notes", "") or "No frozen GDP series exists for the mapped geography; representative year stayed manual.",
        )
        selection["candidate"] = fallback_candidate
        return selection

    series = geography_payload["series"]
    anchor_year, anchor_value, base_mode = _anchor_year_for_series(series, reference_value)
    candidate_rows = []
    for candidate in candidates:
        year = int(candidate["year"])
        gdp_value = _series_value_for_year(series, year)
        if gdp_value is None:
            continue
        candidate_rows.append(
            {
                **candidate,
                "gdp_value": gdp_value,
                "gdp_distance": abs(gdp_value - reference_value),
            }
        )
    if not candidate_rows:
        fallback_candidate = _manual_candidate_fallback(candidates, reference_year)
        selection = _build_selection_stub(
            entity,
            resource,
            map_row,
            reference_value,
            fallback_candidate["year"] if fallback_candidate is not None else "",
            "manual_override",
            "Mapped GDP geography lacks a value for the candidate resource years; representative year stayed manual.",
        )
        selection["candidate"] = fallback_candidate
        return selection

    if base_mode == "economic_peak_fallback":
        selected = max(candidate_rows, key=lambda row: (row["gdp_value"], row["year"]))
        selection_mode = "economic_peak_fallback"
        selection_note = (
            f"{geography_used} never reaches GBR 1950 GDP per capita in the frozen series; "
            f"candidate year {selected['year']} is the strongest available development match under that ceiling."
        )
    else:
        selected = min(
            candidate_rows,
            key=lambda row: (row["gdp_distance"], abs(int(row["year"]) - int(anchor_year)), int(row["year"])),
        )
        selection_mode = "gdp_match"
        selection_note = (
            f"Selected candidate year {selected['year']} because its GDP per capita is closest to the GBR 1950 reference anchor."
        )

    if any(candidate.get("label") == "peak" for candidate in candidate_rows):
        peak_row = next(candidate for candidate in candidate_rows if candidate.get("label") == "peak")
        current_output = float(selected.get("normalized_output", 0.0))
        peak_output = float(peak_row.get("normalized_output", 0.0))
        if peak_row["year"] < selected["year"] and peak_output > current_output * 1.15:
            selected = peak_row
            selection_mode = "resource_peak_override"
            selection_note = (
                f"Earlier peak year {peak_row['year']} retained because the GDP-matched candidate materially understates a depleted or exhausted field."
            )
        elif peak_row["year"] > selected["year"] and peak_output > current_output * 1.15:
            selected = peak_row
            selection_mode = "recovery_override"
            selection_note = (
                f"Later peak year {peak_row['year']} retained because the GDP-matched candidate appears depressed relative to the recovered field."
            )

    distance_ratio = abs(float(selected["gdp_value"]) - reference_value) / reference_value if reference_value else ""
    return {
        "entity": entity,
        "resource": resource,
        "gdp_geography_used": geography_used,
        "gdp_geography_level": geography_payload["geography_level"],
        "gdp_reference_value": reference_value,
        "gdp_anchor_year": anchor_year if anchor_year is not None else "",
        "gdp_value_at_anchor_year": anchor_value if anchor_value is not None else "",
        "selected_resource_year": selected["year"],
        "selection_mode": selection_mode,
        "gdp_distance_ratio": distance_ratio,
        "gdp_comparability_status": "",
        "target_estimation_method": "",
        "interpolation_used": "no",
        "comparator_included_in_denominator": "",
        "comparator_basket_coverage_share": "",
        "selection_note": selection_note or map_row.get("notes", ""),
        "candidate": selected,
    }


def gdp_profile_for_entity(
    entity: str,
    entity_type: str,
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> dict[str, Any]:
    map_row = gdp_map.get((entity_type, entity))
    reference_value = float(gdp_reference["gdp_per_capita"] or 0.0)
    if map_row is None:
        return {
            "gdp_geography_used": "",
            "gdp_geography_level": "",
            "gdp_anchor_year": "",
            "gdp_value_at_anchor_year": "",
            "selection_mode": "manual_override",
            "selection_note": "No GDP geography mapping exists for this entity.",
        }
    preferred = map_row.get("preferred_substate_geography", "")
    fallback = map_row.get("fallback_national_geography", "")
    geography_used = preferred if preferred in gdp_series else fallback
    geography_payload = gdp_series.get(geography_used)
    if geography_payload is None:
        return {
            "gdp_geography_used": geography_used,
            "gdp_geography_level": "",
            "gdp_anchor_year": "",
            "gdp_value_at_anchor_year": "",
            "selection_mode": "manual_override",
            "selection_note": map_row.get("notes", "") or "No frozen GDP series exists for the mapped geography.",
        }
    anchor_year, anchor_value, mode = _anchor_year_for_series(geography_payload["series"], reference_value)
    return {
        "gdp_geography_used": geography_used,
        "gdp_geography_level": geography_payload["geography_level"],
        "gdp_anchor_year": anchor_year if anchor_year is not None else "",
        "gdp_value_at_anchor_year": anchor_value if anchor_value is not None else "",
        "selection_mode": mode,
        "selection_note": map_row.get("notes", ""),
    }


def select_observation_rows(
    observation_rows: list[dict[str, Any]],
    entity_type: str,
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> tuple[dict[tuple[str, str], dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in observation_rows:
        grouped[(row["entity"], row["resource"])].append(row)
    selected_rows: dict[tuple[str, str], dict[str, Any]] = {}
    selection_rows: list[dict[str, Any]] = []
    for (entity, resource), rows in grouped.items():
        filtered_rows = [row for row in rows if resource not in START_DATE_SPECIALS or int(row["year"]) <= 1836]
        if not filtered_rows:
            filtered_rows = rows
        candidates = [
            {
                "year": int(row["year"]),
                "normalized_output": float(row["normalized_1950_output"]),
                "row": row,
                "observation_family": row.get("observation_family", ""),
            }
            for row in filtered_rows
        ]
        selection = select_year_by_gdp(entity, entity_type, resource, candidates, gdp_reference, gdp_series, gdp_map)
        candidate = selection.pop("candidate")
        if candidate is None:
            continue
        selected_rows[(entity, resource)] = candidate["row"]
        selection_rows.append(selection)
    return selected_rows, selection_rows


def linear_interpolate_value(year_a: int, value_a: float, year_b: int, value_b: float, target_year: int) -> float:
    if year_a == year_b:
        return value_a
    span = year_b - year_a
    fraction = (target_year - year_a) / span
    return value_a + (value_b - value_a) * fraction


def candidate_distance_ratio(candidate: dict[str, Any], reference_value: float) -> float | None:
    gdp_value = candidate.get("gdp_value")
    if gdp_value in (None, "") or not reference_value:
        return None
    return abs(float(gdp_value) - float(reference_value)) / float(reference_value)


def build_interpolated_candidate(
    candidates: list[dict[str, Any]],
    anchor_year: int | None,
    anchor_value: float | None,
    reference_value: float,
    max_span: int,
) -> dict[str, Any] | None:
    if anchor_year in (None, "") or anchor_value in (None, ""):
        return None
    if not reference_value:
        return None
    anchor_distance_ratio = abs(float(anchor_value) - float(reference_value)) / float(reference_value)
    if anchor_distance_ratio > ARABLE_GDP_MAX_DISTANCE_RATIO:
        return None
    candidates_by_family: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for candidate in candidates:
        family = str(candidate.get("observation_family", ""))
        if family:
            candidates_by_family[family].append(candidate)
    best_candidate: dict[str, Any] | None = None
    best_key: tuple[float, int, int] | None = None
    for family_candidates in candidates_by_family.values():
        earlier_rows = [row for row in family_candidates if int(row["year"]) < int(anchor_year)]
        later_rows = [row for row in family_candidates if int(row["year"]) > int(anchor_year)]
        for earlier in earlier_rows:
            for later in later_rows:
                span = int(later["year"]) - int(earlier["year"])
                if span <= 0 or span > max_span:
                    continue
                if earlier.get("row", {}).get("wheat_equivalent_1950_output") in ("", None):
                    continue
                if later.get("row", {}).get("wheat_equivalent_1950_output") in ("", None):
                    continue
                interpolated_normalized = linear_interpolate_value(
                    int(earlier["year"]),
                    float(earlier["normalized_output"]),
                    int(later["year"]),
                    float(later["normalized_output"]),
                    int(anchor_year),
                )
                interpolated_wheat = linear_interpolate_value(
                    int(earlier["year"]),
                    float(earlier["row"]["wheat_equivalent_1950_output"]),
                    int(later["year"]),
                    float(later["row"]["wheat_equivalent_1950_output"]),
                    int(anchor_year),
                )
                interpolated_row = dict(earlier["row"])
                interpolated_row.update(
                    {
                        "year": int(anchor_year),
                        "normalized_1950_output": interpolated_normalized,
                        "wheat_equivalent_1950_output": interpolated_wheat,
                        "note": (
                            f"{earlier['row'].get('note', '')} | "
                            f"Interpolated to GDP anchor year {anchor_year} from {earlier['year']}-{later['year']}."
                        ).strip(" |"),
                    }
                )
                distance_key = (
                    abs(int(anchor_year) - int(earlier["year"])) + abs(int(later["year"]) - int(anchor_year)),
                    span,
                    int(later["year"]),
                )
                if best_key is None or distance_key < best_key:
                    best_key = distance_key
                    best_candidate = {
                        "year": int(anchor_year),
                        "normalized_output": interpolated_normalized,
                        "row": interpolated_row,
                        "observation_family": earlier.get("observation_family", ""),
                        "gdp_value": float(anchor_value),
                        "gdp_distance": abs(float(anchor_value) - float(reference_value)),
                        "interpolation_bounds": f"{earlier['year']}-{later['year']}",
                    }
    return best_candidate


def finalize_arable_selection(
    selection: dict[str, Any],
    entity_type: str,
    reference_value: float,
    filtered_rows: list[dict[str, Any]],
) -> tuple[dict[str, Any] | None, dict[str, Any]]:
    candidate = selection.get("candidate")
    selected_row = candidate["row"] if candidate is not None else None
    distance_ratio = selection.get("gdp_distance_ratio")
    anchor_year = parse_int(selection.get("gdp_anchor_year"))
    anchor_value = parse_float(selection.get("gdp_value_at_anchor_year"))
    interpolation_candidate = build_interpolated_candidate(
        [
            {
                "year": int(row["year"]),
                "normalized_output": float(row["normalized_1950_output"]),
                "row": row,
                "observation_family": row.get("observation_family", ""),
            }
            for row in filtered_rows
        ],
        anchor_year,
        anchor_value,
        reference_value,
        ARABLE_INTERPOLATION_MAX_SPAN,
    )

    comparable = (
        selection.get("selection_mode") not in {"manual_override", "economic_peak_fallback"}
        and distance_ratio not in ("", None)
        and float(distance_ratio) <= ARABLE_GDP_MAX_DISTANCE_RATIO
    )
    if comparable:
        selection["gdp_comparability_status"] = "comparable"
        selection["interpolation_used"] = "no"
        return selected_row, selection
    if interpolation_candidate is not None:
        selection["selected_resource_year"] = interpolation_candidate["year"]
        selection["selection_mode"] = "interpolated_gdp_match"
        selection["gdp_distance_ratio"] = candidate_distance_ratio(interpolation_candidate, reference_value) or 0.0
        selection["gdp_comparability_status"] = "interpolated_comparable"
        selection["interpolation_used"] = "yes"
        selection["selection_note"] = (
            f"Interpolated to GDP anchor year {anchor_year} from bounded {interpolation_candidate['interpolation_bounds']} observations."
        )
        selection["candidate"] = interpolation_candidate
        return interpolation_candidate["row"], selection
    selection["interpolation_used"] = "no"
    if entity_type == "target_state":
        selection["gdp_comparability_status"] = "resource_history_fallback"
        selection["selection_note"] = (
            f"{selection.get('selection_note', '')} GDP distance exceeds the arable comparability gate, so the best defended target observation is retained as local evidence."
        ).strip()
        return selected_row, selection
    selection["gdp_comparability_status"] = "noncomparable_excluded"
    selection["selection_note"] = (
        f"{selection.get('selection_note', '')} Comparator row fails the arable GDP comparability gate and is excluded from denominator arithmetic."
    ).strip()
    return selected_row, selection


def select_arable_observation_rows(
    observation_rows: list[dict[str, Any]],
    entity_type: str,
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> tuple[dict[tuple[str, str], dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    arable_resource_names = {resource for category, resource in BINARY_RESOURCES if category == "Arable Resource"}
    for row in observation_rows:
        if row["resource"] not in arable_resource_names:
            continue
        grouped[(row["entity"], row["resource"])].append(row)
    selected_rows: dict[tuple[str, str], dict[str, Any]] = {}
    selection_rows: list[dict[str, Any]] = []
    reference_value = float(gdp_reference["gdp_per_capita"] or 0.0)
    for (entity, resource), rows in grouped.items():
        filtered_rows = [row for row in rows if resource not in START_DATE_SPECIALS or int(row["year"]) <= 1836]
        if not filtered_rows:
            filtered_rows = rows
        candidates = [
            {
                "year": int(row["year"]),
                "normalized_output": float(row["normalized_1950_output"]),
                "row": row,
                "observation_family": row.get("observation_family", ""),
            }
            for row in filtered_rows
        ]
        selection = select_year_by_gdp(entity, entity_type, resource, candidates, gdp_reference, gdp_series, gdp_map)
        finalized_row, finalized_selection = finalize_arable_selection(selection, entity_type, reference_value, filtered_rows)
        finalized_selection.pop("candidate", None)
        if finalized_row is not None:
            selected_rows[(entity, resource)] = finalized_row
        selection_rows.append(finalized_selection)
    return selected_rows, selection_rows


def estimate_target_bidirectional_x(
    arable_target_selection_map: dict[tuple[str, str], dict[str, Any]],
    target_observations: list[dict[str, Any]],
    preliminary_expectation_rows: list[dict[str, Any]],
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
) -> tuple[dict[tuple[str, str], float], dict[tuple[str, str], dict[str, Any]]]:
    overrides: dict[tuple[str, str], float] = {}
    grouped_rows: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in target_observations:
        if row["resource"] not in LAND_ECONOMY_RESOURCES:
            continue
        grouped_rows[(row["entity"], row["resource"])].append(row)
    reference_value = float(gdp_reference["gdp_per_capita"] or 0.0)
    eligible_keys = {
        (row["state"], row["resource"])
        for row in preliminary_expectation_rows
        if row["target_output_status"] == "partial" and row["researched_plausible"] == "yes"
    }
    for key, selection in arable_target_selection_map.items():
        if not selection.get("target_estimation_method"):
            selection["target_estimation_method"] = "selected_observation"
        if key not in eligible_keys:
            continue
        anchor_year = parse_int(selection.get("gdp_anchor_year"))
        geography_used = selection.get("gdp_geography_used", "")
        if anchor_year is None or not geography_used or geography_used not in gdp_series:
            continue
        series = gdp_series[geography_used]["series"]
        observation_rows = grouped_rows.get(key, [])
        if len(observation_rows) < 2:
            continue
        best_pair: tuple[dict[str, Any], dict[str, Any], tuple[int, int, int]] | None = None
        families = {row.get("observation_family", "") for row in observation_rows if row.get("observation_family")}
        for family in families:
            family_rows = [row for row in observation_rows if row.get("observation_family", "") == family]
            earlier_rows = [row for row in family_rows if int(row["year"]) < anchor_year and row.get("wheat_equivalent_1950_output") not in ("", None)]
            later_rows = [row for row in family_rows if int(row["year"]) > anchor_year and row.get("wheat_equivalent_1950_output") not in ("", None)]
            for earlier in earlier_rows:
                for later in later_rows:
                    span = int(later["year"]) - int(earlier["year"])
                    if span <= 0 or span > ARABLE_TARGET_BIDIRECTIONAL_MAX_SPAN:
                        continue
                    if _series_value_for_year(series, int(earlier["year"])) is None or _series_value_for_year(series, int(later["year"])) is None:
                        continue
                    score = (
                        span,
                        abs(anchor_year - int(earlier["year"])) + abs(int(later["year"]) - anchor_year),
                        int(later["year"]),
                    )
                    if best_pair is None or score < best_pair[2]:
                        best_pair = (earlier, later, score)
        if best_pair is None:
            continue
        earlier, later, _score = best_pair
        gdp_earlier = _series_value_for_year(series, int(earlier["year"]))
        gdp_later = _series_value_for_year(series, int(later["year"]))
        if gdp_earlier is None or gdp_later is None:
            continue
        earlier_distance = abs(float(gdp_earlier) - reference_value)
        later_distance = abs(float(gdp_later) - reference_value)
        if earlier_distance == 0:
            estimate = float(earlier["wheat_equivalent_1950_output"])
        elif later_distance == 0:
            estimate = float(later["wheat_equivalent_1950_output"])
        else:
            earlier_weight = 1.0 / earlier_distance
            later_weight = 1.0 / later_distance
            estimate = (
                (earlier_weight * float(earlier["wheat_equivalent_1950_output"]))
                + (later_weight * float(later["wheat_equivalent_1950_output"]))
            ) / (earlier_weight + later_weight)
        overrides[key] = estimate
        selection["target_estimation_method"] = "bidirectional_target_estimate"
        selection["selection_note"] = (
            f"{selection.get('selection_note', '')} Target x is replaced by a GDP-distance-weighted bidirectional estimate from {earlier['year']}-{later['year']}."
        ).strip()
    return overrides, arable_target_selection_map


def enrich_observation_rows(rows: list[dict[str, Any]], include_target: bool, growth_index_lookup: dict[tuple[str, int], float]) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    corroboration_counts: dict[tuple[str, str], int] = defaultdict(int)
    if include_target:
        for row in rows:
            state_info = STATE_INFO_BY_SHEET.get(str(row["sheet"]))
            if state_info is None:
                continue
            corroboration_counts[(state_info["official_name"], str(row["resource"]))] += 1
    for row in rows:
        geography = str(row["geography"])
        is_target = "target" in geography.lower()
        if is_target != include_target:
            continue
        resource = str(row["resource"])
        standardized = standardized_quantity(float(row["normalized_quantity"]), str(row["normalized_unit"]))
        standardized_unit = canonical_unit(str(row["normalized_unit"]))
        growth_family = growth_family_for_resource(resource)
        index_to_1950 = growth_index_lookup.get((growth_family, int(row["year"])), 1.0)
        normalized_1950 = standardized * index_to_1950
        conversion_key = conversion_key_for_row(resource, standardized_unit)
        conversion_factor = CONVERSION_FACTOR_MAP.get(conversion_key) if conversion_key else None
        wheat_equivalent = normalized_1950 * conversion_factor if conversion_factor is not None else None
        short_name = short_geography_name(geography)
        validation = {field: row.get(field, "") for field in TARGET_VALIDATION_FIELDS}
        corroboration_count = corroboration_counts.get((STATE_INFO_BY_SHEET[row["sheet"]]["official_name"], resource), 1) if include_target and row["sheet"] in STATE_INFO_BY_SHEET else 1
        drives_x = validation_drives_x(validation, corroboration_count) if include_target else False
        localization_discount = float(validation.get("localization_discount") or 1.0) if include_target else 1.0
        discounted_normalized_1950 = normalized_1950 * localization_discount if include_target and drives_x else ""
        discounted_wheat_equivalent = (
            wheat_equivalent * localization_discount
            if include_target and drives_x and wheat_equivalent is not None
            else ""
        )
        review_action = validation_review_action(validation, drives_x, corroboration_count) if include_target else ""
        enriched.append(
            {
                "sheet_key": row["sheet"],
                "official_state": STATE_INFO_BY_SHEET[row["sheet"]]["official_name"],
                "entity": STATE_INFO_BY_SHEET[row["sheet"]]["official_name"] if include_target else short_name,
                "geography": geography,
                "short_geography": short_name,
                "resource": resource,
                "resource_family": resource_family(resource),
                "observation_family": row.get("observation_family", ""),
                "year": int(row["year"]),
                "raw_quantity": row["normalized_quantity"],
                "raw_unit": row["normalized_unit"],
                "standardized_quantity": standardized,
                "standardized_unit": standardized_unit,
                "sector": "Land Economy" if resource in LAND_ECONOMY_RESOURCES else ("Fisheries / Marine" if resource in {"Fishing", "Whaling"} else ("Forestry" if resource == "Wood" else "Mining")),
                "growth_family": growth_family,
                "index_to_1950": index_to_1950,
                "conversion_key": conversion_key or "",
                "conversion_factor": conversion_factor if conversion_factor is not None else "",
                "normalized_1950_output": normalized_1950,
                "wheat_equivalent_1950_output": wheat_equivalent if wheat_equivalent is not None else "",
                "discounted_normalized_1950_output": discounted_normalized_1950,
                "discounted_wheat_equivalent_1950_output": discounted_wheat_equivalent,
                "drives_x": "yes" if drives_x else "no",
                "review_action": review_action,
                "corroboration_count": corroboration_count if include_target else "",
                "proxy_state_id": row["sheet"] if include_target else "",
                "source_title": row["source_title"],
                "source_url": row["source_url"],
                "locator": row["citation_locator"],
                "note": row.get("note", ""),
                **validation,
            }
        )
    return enriched


def build_selected_target_map(selected_target_rows: dict[tuple[str, str], dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    by_state_resource: dict[tuple[str, str], dict[str, Any]] = {}
    for (entity, resource), row in selected_target_rows.items():
        by_state_resource[(entity, resource)] = row
    return by_state_resource


def summarize_counterevidence(
    cases: list[dict[str, str]],
    final_cap: int,
    exception_status: str,
) -> tuple[str, str]:
    if exception_status:
        if cases:
            note = " | ".join(case.get("decision", "") for case in cases if case.get("decision"))
            return "constrains", note or "Explicit exception row."
        return "constrains", "Explicit exception row."
    if not cases:
        return "not_reviewed", "No linked counterevidence case in the frozen research registry."
    decision_blob = " | ".join(case.get("decision", "") for case in cases if case.get("decision"))
    lower = decision_blob.lower()
    if final_cap == 0 and any(token in lower for token in ["remains disabled", "keeps no", "set to 0", "rejected", "unsupported", "keeps only undiscovered"]):
        return "supports", decision_blob
    if final_cap > 0 and any(token in lower for token in ["enabled", "retain", "keeps only undiscovered", "keeps a large latent", "remains enabled", "belongs in"]):
        return "supports", decision_blob
    if any(token in lower for token in ["rejected", "unsupported", "too small", "not enough", "constrained"]):
        return "constrains", decision_blob
    return "constrains", decision_blob


def build_target_observed_maps(
    selected_target_rows: dict[tuple[str, str], dict[str, Any]],
    wheat_equivalent_overrides: dict[tuple[str, str], float] | None = None,
) -> tuple[dict[tuple[str, str], float], dict[tuple[str, str], float]]:
    normalized: dict[tuple[str, str], float] = {}
    wheat_eq: dict[tuple[str, str], float] = {}
    overrides = wheat_equivalent_overrides or {}
    for row in selected_target_rows.values():
        key = (row["official_state"], row["resource"])
        if row.get("drives_x") != "yes":
            continue
        normalized_value = row.get("discounted_normalized_1950_output", row["normalized_1950_output"])
        normalized[key] = float(normalized_value)
        if key in overrides:
            wheat_eq[key] = float(overrides[key])
        else:
            discounted_wheat = row.get("discounted_wheat_equivalent_1950_output", row["wheat_equivalent_1950_output"])
            if discounted_wheat not in ("", None):
                wheat_eq[key] = float(discounted_wheat)
    return normalized, wheat_eq


def build_comparator_maxima(selected_comparator_rows: dict[tuple[str, str], dict[str, Any]]) -> tuple[dict[tuple[str, str], float], dict[str, float]]:
    maxima: dict[tuple[str, str], float] = {}
    combined: dict[str, float] = defaultdict(float)
    for row in selected_comparator_rows.values():
        if row["resource"] not in LAND_ECONOMY_RESOURCES or row["wheat_equivalent_1950_output"] in ("", None):
            continue
        key = (row["entity"], row["resource"])
        maxima[key] = float(row["wheat_equivalent_1950_output"])
    for (entity, _resource), value in maxima.items():
        combined[entity] += value
    return maxima, dict(combined)


def compute_benchmark_comparator_rows(
    cases: list[dict[str, Any]],
    growth_index_lookup: dict[tuple[str, int], float],
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
    entity_type: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    selection_rows: list[dict[str, Any]] = []
    for case in cases:
        family = resource_family(str(case["resource"]))
        growth_family = growth_family_for_resource(str(case["resource"]))
        historical_standardized = standardized_quantity(float(case["historical_quantity"]), str(case["historical_unit"]))
        historical_1950 = normalize_to_1950(historical_standardized, int(case["historical_year"]), growth_family, growth_index_lookup)
        peak_1950 = None
        if case.get("peak_year") not in (None, "") and case.get("peak_quantity") not in (None, ""):
            peak_standardized = standardized_quantity(float(case["peak_quantity"]), str(case["peak_unit"]))
            peak_1950 = normalize_to_1950(peak_standardized, int(case["peak_year"]), growth_family, growth_index_lookup)
        candidate_rows = [
            {
                "year": int(case["historical_year"]),
                "label": "historical",
                "normalized_output": historical_1950,
            }
        ]
        if peak_1950 is not None:
            candidate_rows.append(
                {
                    "year": int(case["peak_year"]),
                    "label": "peak",
                    "normalized_output": peak_1950,
                }
            )
        selection = select_year_by_gdp(
            str(case["benchmark_state"]),
            entity_type,
            str(case["resource"]),
            candidate_rows,
            gdp_reference,
            gdp_series,
            gdp_map,
        )
        selected_candidate = selection.pop("candidate")
        selected_year = int(selected_candidate["year"]) if selected_candidate is not None else int(case["historical_year"])
        selected_label = selected_candidate.get("label", "historical") if selected_candidate is not None else "historical"
        maximum = historical_1950 if selected_label == "historical" else (peak_1950 if peak_1950 is not None else historical_1950)
        vanilla_cap = parse_float(case.get("benchmark_vanilla_cap")) or 0.0
        rows.append(
            {
                "resource_family": family,
                "comparator_geography": case["benchmark_state"],
                "vanilla_proxy_state_id": case["benchmark_state_id"],
                "historical_year": case["historical_year"],
                "historical_output": historical_standardized,
                "historical_unit": canonical_unit(str(case["historical_unit"])),
                "peak_year": case.get("peak_year", ""),
                "peak_output": standardized_quantity(float(case["peak_quantity"]), str(case["peak_unit"])) if case.get("peak_quantity") not in (None, "") else "",
                "peak_unit": canonical_unit(str(case.get("peak_unit", ""))) if case.get("peak_quantity") not in (None, "") else "",
                "growth_family": growth_family,
                "historical_index_to_1950": growth_index_lookup.get((growth_family, int(case["historical_year"])), ""),
                "peak_index_to_1950": growth_index_lookup.get((growth_family, int(case["peak_year"])), "") if case.get("peak_year") not in (None, "") else "",
                "historical_1950_equivalent_output": historical_1950,
                "peak_1950_equivalent_output": peak_1950 if peak_1950 is not None else "",
                "normalized_1950_equivalent_maximum": maximum,
                "selected_resource_year": selected_year,
                "selection_mode": selection["selection_mode"],
                "gdp_geography_used": selection["gdp_geography_used"],
                "gdp_geography_level": selection["gdp_geography_level"],
                "gdp_anchor_year": selection["gdp_anchor_year"],
                "gdp_value_at_anchor_year": selection["gdp_value_at_anchor_year"],
                "gdp_distance_ratio": selection["gdp_distance_ratio"],
                "selection_note": selection["selection_note"],
                "vanilla_cap": vanilla_cap,
                "max_output_units_per_cap": maximum / vanilla_cap if vanilla_cap else "",
                "source_a_title": case["source_a_title"],
                "source_a_url": case["source_a_url"],
                "source_a_locator": case["source_a_locator"],
                "source_b_title": case["source_b_title"],
                "source_b_url": case["source_b_url"],
                "source_b_locator": case["source_b_locator"],
                "comment": case["comment"],
            }
        )
        selection_rows.append(selection)
    return rows, selection_rows


def compute_mining_comparator_rows(
    cases: list[dict[str, Any]],
    growth_index_lookup: dict[tuple[str, int], float],
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    return compute_benchmark_comparator_rows(cases, growth_index_lookup, gdp_reference, gdp_series, gdp_map, "mining_comparator")


def compute_wood_comparator_rows(
    cases: list[dict[str, Any]],
    growth_index_lookup: dict[tuple[str, int], float],
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    return compute_benchmark_comparator_rows(cases, growth_index_lookup, gdp_reference, gdp_series, gdp_map, "wood_comparator")


def compute_resource_denominators(
    mining_rows: list[dict[str, Any]],
    wood_denominator_row: dict[str, Any],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    families = ["Coal Mine", "Iron Mine", "Gold Fields", "Gold Mine", "Lead Mine", "Sulfur Mine", "Fishing", "Whaling", "Oil", "Rubber", "Wood"]
    payloads: dict[str, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []
    for family in families:
        if family == "Wood":
            payloads[family] = dict(wood_denominator_row)
        else:
            source_family = "Gold Fields" if family == "Gold Mine" else family
            values = [
                float(row["max_output_units_per_cap"])
                for row in mining_rows
                if row["resource_family"] == source_family and row["max_output_units_per_cap"] not in ("", None)
            ]
            valid_count = len(values)
            average = (sum(values) / valid_count) if values else None
            status = "formula-driven" if valid_count >= 20 else "denominator_unavailable"
            payloads[family] = {
                "resource_family": family,
                "valid_comparator_count": valid_count,
                "denominator_units_per_cap": average,
                "status": status,
                "method": (
                    "simple_mean_of_20_comparator_pool"
                    if status == "formula-driven"
                    else "insufficient_comparator_pool"
                ),
            }
        rows.append(payloads[family])
    return payloads, rows


def weighted_mean_and_population_stdev(values: list[tuple[float, float]]) -> tuple[float | None, float | None]:
    active = [(value, weight) for value, weight in values if weight > 0]
    if not active:
        return None, None
    total_weight = sum(weight for _, weight in active)
    mean = sum(value * weight for value, weight in active) / total_weight
    variance = sum(weight * ((value - mean) ** 2) for value, weight in active) / total_weight
    return mean, math.sqrt(variance)


def gaussian_penalty(z_score: float, k_value: float) -> float:
    return math.exp(-0.5 * ((z_score / k_value) ** 2))


def compute_arable_state_means(
    rankings: list[dict[str, Any]],
    target_capacity_rows: list[dict[str, Any]],
    comparator_capacity_rows: list[dict[str, Any]],
    arable_baskets: dict[str, dict[str, str]],
    live_values: dict[str, dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    state_payloads: dict[str, dict[str, Any]] = {}
    state_rows: list[dict[str, Any]] = []
    comparator_diagnostics: list[dict[str, Any]] = []
    target_totals: dict[str, float] = defaultdict(float)
    for row in target_capacity_rows:
        target_totals[row["state"]] += float(row["effective_area_ha"])
    comparator_totals: dict[tuple[str, str], dict[str, Any]] = {}
    for row in comparator_capacity_rows:
        key = (row["target_state"], row["comparator_geography"])
        payload = comparator_totals.setdefault(
            key,
            {
                "benchmark_state_id": row["benchmark_state_id"],
                "rank": row["rank"],
                "benchmark_vanilla_arable_cap": row["benchmark_vanilla_arable_cap"],
                "effective_area_ha": 0.0,
                "year_selection_reason": row["year_selection_reason"],
                "capacity_note": row["capacity_note"],
            },
        )
        payload["effective_area_ha"] += float(row["effective_area_ha"])
    for info in STATE_INFO:
        state = info["official_name"]
        basket_resources = [resource for _, resource in BINARY_RESOURCES if arable_baskets[state].get(resource) == "yes"]
        live_resources = [resource for _, resource in BINARY_RESOURCES if live_values[state].get(resource) == "yes"]
        researched_only_count = len([resource for resource in basket_resources if resource not in live_resources])
        live_only_count = len([resource for resource in live_resources if resource not in basket_resources])
        overlap_count = len([resource for resource in basket_resources if resource in live_resources])
        researched_basket_size = len(basket_resources)
        researched_only_share = (researched_only_count / researched_basket_size) if researched_basket_size else 0.0
        weighted_units = 0.0
        active_weight_sum = 0.0
        valid_count = 0
        for rank_row in [row for row in rankings if row["official_state"] == state]:
            entity = rank_row["matched_raw_geography"]
            weight = int(rank_row["weight"] or 0)
            comparator_payload = comparator_totals.get((state, entity))
            included = bool(entity and weight and comparator_payload and comparator_payload["benchmark_vanilla_arable_cap"])
            effective_area = float(comparator_payload["effective_area_ha"]) if comparator_payload else 0.0
            benchmark_cap = float(comparator_payload["benchmark_vanilla_arable_cap"]) if comparator_payload else 0.0
            effective_hectares_per_cap = (effective_area / benchmark_cap) if benchmark_cap else ""
            comparator_diagnostics.append(
                {
                    "target_state": state,
                    "comparator": entity,
                    "rank": rank_row["rank"],
                    "weight": weight,
                    "effective_commercial_hectares": effective_area if included else "",
                    "effective_hectares_per_cap": effective_hectares_per_cap,
                    "comparator_included_in_denominator": "yes" if included else "no",
                    "included_reason": "Comparator contributes a land-capacity benchmark row." if included else "Missing comparator land-capacity row or comparator proxy cap.",
                    "capacity_note": comparator_payload["capacity_note"] if comparator_payload else "",
                }
            )
            if not included:
                continue
            valid_count += 1
            active_weight_sum += weight
            weighted_units += float(effective_hectares_per_cap) * weight
        weighted_mean = (weighted_units / active_weight_sum) if active_weight_sum else None
        payload = {
            "state": state,
            "observed_output_x": target_totals.get(state, 0.0),
            "target_effective_commercial_hectares": target_totals.get(state, 0.0),
            "weighted_state_effective_hectares_per_cap": weighted_mean,
            "valid_comparator_count": valid_count,
            "active_weight_sum": active_weight_sum,
            "researched_basket_size": researched_basket_size,
            "live_enabled_count": len(live_resources),
            "overlap_count": overlap_count,
            "researched_only_count": researched_only_count,
            "live_only_count": live_only_count,
            "researched_only_resource_share": researched_only_share,
            "mismatch_status": (
                "large_mismatch"
                if researched_only_share >= 0.30 or researched_only_count >= 3 or live_only_count >= 2
                else "mismatch"
                if researched_only_count or live_only_count
                else "aligned"
            ),
            "status": "formula-driven" if valid_count >= 5 else "denominator_unavailable",
        }
        state_payloads[state] = payload
        state_rows.append(payload)
    return state_payloads, state_rows, comparator_diagnostics


def compute_shared_arable_denominator(state_payloads: dict[str, dict[str, Any]]) -> dict[str, Any]:
    values = [
        payload["weighted_state_effective_hectares_per_cap"]
        for payload in state_payloads.values()
        if payload["weighted_state_effective_hectares_per_cap"] not in (None, "")
    ]
    shared = (sum(values) / len(values)) if values else None
    return {
        "shared_sb_arable_effective_hectares_per_cap": shared,
        "state_mean_count": len(values),
        "method": "simple_mean_of_13_effective_land_state_means",
    }


def build_arable_resource_expectations(
    live_values: dict[str, dict[str, Any]],
    arable_baskets: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for state in [row["official_name"] for row in STATE_INFO]:
        live_resources = [resource for _, resource in BINARY_RESOURCES if live_values[state].get(resource) == "yes"]
        for _category, resource in BINARY_RESOURCES:
            researched_plausible = arable_baskets[state].get(resource) == "yes"
            live_enabled = live_values[state].get(resource) == "yes"
            if researched_plausible and live_enabled:
                membership_status = "researched_and_live"
            elif researched_plausible:
                membership_status = "researched_only"
            elif live_enabled:
                membership_status = "live_only"
            else:
                membership_status = "excluded"
            rows.append(
                {
                    "state": state,
                    "resource": resource,
                    "researched_plausible": "yes" if researched_plausible else "no",
                    "live_enabled_in_state": "yes" if live_enabled else "no",
                    "basket_membership_status": membership_status,
                    "audit_relevance": "gameplay_or_crop_audit_only",
                    "gameplay_note": (
                        "Researched plausible crop is absent from live gameplay resources."
                        if membership_status == "researched_only"
                        else "Live gameplay resource sits outside the researched crop-suitability basket."
                        if membership_status == "live_only"
                        else "Live gameplay and researched crop basket are aligned for this resource."
                        if membership_status == "researched_and_live"
                        else "Resource is outside both the researched basket and live gameplay set."
                    ),
                }
            )
    return rows


def legacy_override_rows() -> list[dict[str, str]]:
    return read_csv_rows(LEGACY_OVERRIDES_CSV)


ADJUSTMENT_INPUT_FIELDNAMES = [
    "state",
    "resource",
    "problem_type",
    "output_addition_y",
    "plausibility_haircut_z",
    "minimum_operating_floor_cap",
    "exception_status",
    "exception_final_cap",
    "representative_max_year",
    "representative_max_output",
    "year_selection_reason",
    "counterevidence_status",
    "counterevidence_note",
    "plausibility_1936_status",
    "observed_output_x_reason",
    "y_basis",
    "y_reason",
    "y_counterevidence_trigger",
    "y_quantification_method",
    "z_reason",
    "audit_class",
    "adjustment_reason",
    "proxy_kind",
    "proxy_native_unit",
    "proxy_to_slot_note",
    "citation_1_title",
    "citation_1_url",
    "citation_1_locator",
    "citation_2_title",
    "citation_2_url",
    "citation_2_locator",
    "calculation_note",
]


def seed_adjustment_inputs_if_needed(
    base_rows: dict[tuple[str, str], dict[str, Any]],
    arable_state_payloads: dict[str, dict[str, Any]],
) -> None:
    if RAW_ADJUSTMENT_INPUTS_CSV.exists():
        return
    seeded_rows: list[dict[str, Any]] = []
    for row in legacy_override_rows():
        state = row["state"]
        resource = row["resource"]
        key = (state, resource)
        base = base_rows.get(key, {})
        denominator = parse_float(base.get("denominator_units_per_cap"))
        observed = parse_float(base.get("observed_output_x")) or 0.0
        final_cap = parse_int(row["final_audited_cap"]) or 0
        proxy_kind, proxy_native_unit, proxy_to_slot_note = PROXY_NOTES.get(key, ("", "", ""))
        seeded = {
            "state": state,
            "resource": resource,
            "problem_type": row["problem_type"],
            "output_addition_y": 0.0,
            "plausibility_haircut_z": 0.0,
            "minimum_operating_floor_cap": "",
            "exception_status": "",
            "exception_final_cap": "",
            "representative_max_year": "",
            "representative_max_output": "",
            "year_selection_reason": "",
            "counterevidence_status": "",
            "counterevidence_note": "",
            "plausibility_1936_status": "",
            "observed_output_x_reason": "",
            "y_basis": "",
            "y_reason": "",
            "y_counterevidence_trigger": "",
            "y_quantification_method": "",
            "z_reason": "",
            "audit_class": "",
            "adjustment_reason": row["override_reason"],
            "proxy_kind": proxy_kind,
            "proxy_native_unit": proxy_native_unit,
            "proxy_to_slot_note": proxy_to_slot_note,
            "citation_1_title": row["citation_1_title"],
            "citation_1_url": row["citation_1_url"],
            "citation_1_locator": row["citation_1_locator"],
            "citation_2_title": row["citation_2_title"],
            "citation_2_url": row["citation_2_url"],
            "citation_2_locator": row["citation_2_locator"],
            "calculation_note": "",
        }
        if row["problem_type"] == "rounding floor":
            seeded["minimum_operating_floor_cap"] = final_cap
            seeded["calculation_note"] = "Legacy rounding-floor row migrated without nonzero output adjustment terms."
        elif denominator in (None, 0.0) or row["problem_type"].startswith("manual family"):
            seeded["exception_status"] = "denominator_unavailable"
            seeded["exception_final_cap"] = final_cap
            seeded["calculation_note"] = "Legacy manual family row migrated as an explicit denominator-unavailable exception."
        else:
            desired_output = final_cap * denominator
            if resource == "Arable Land":
                seeded["output_addition_y"] = 0.0
                seeded["plausibility_haircut_z"] = 0.0
                seeded["y_basis"] = ""
                seeded["y_reason"] = "No upward addition."
                seeded["y_counterevidence_trigger"] = "No counterevidence-triggered addition."
                seeded["y_quantification_method"] = "No upward addition."
                seeded["z_reason"] = "No downward plausibility haircut."
                seeded["audit_class"] = "direct"
                seeded["adjustment_reason"] = "Land-capacity model uses direct effective commercial hectares; legacy output-gap uplift is retired."
                seeded["calculation_note"] = "Seeded for the land-capacity model: arable no longer uses legacy output-gap Y terms."
            elif row["problem_type"] in {"overstatement", "chronology"}:
                seeded["plausibility_haircut_z"] = round(max(0.0, observed - desired_output), 6)
                seeded["calculation_note"] = "Seeded from legacy audited cap as a downward plausibility haircut in output units."
            elif resource == "Lead Mine":
                seeded["output_addition_y"] = round(max(0.0, desired_output - observed), 6)
                seeded["y_basis"] = "proxy_target_gap"
                seeded["y_reason"] = "Legacy seeded proxy uplift on a base-metal abstraction row."
                seeded["y_counterevidence_trigger"] = "Direct target evidence survives only through a proxy abstraction."
                seeded["y_quantification_method"] = "Difference between audited output target and direct observed output."
                seeded["calculation_note"] = "Seeded from legacy audited cap as an explicit proxy-output uplift on a base-metal abstraction row."
            else:
                seeded["output_addition_y"] = round(max(0.0, desired_output - observed), 6)
                seeded["y_basis"] = "counterevidence_bounded_gap"
                seeded["y_reason"] = "Legacy seeded uplift for an incomplete direct target chain."
                seeded["y_counterevidence_trigger"] = "Direct target observations understate the audited row."
                seeded["y_quantification_method"] = "Difference between audited output target and direct observed output."
                seeded["calculation_note"] = "Seeded from legacy audited cap as one Y term justified by counterevidence and comparator context."
        seeded_rows.append(seeded)
    write_csv(RAW_ADJUSTMENT_INPUTS_CSV, seeded_rows, ADJUSTMENT_INPUT_FIELDNAMES)


def load_adjustment_inputs() -> dict[tuple[str, str], dict[str, Any]]:
    ensure_csv_schema(RAW_ADJUSTMENT_INPUTS_CSV, ADJUSTMENT_INPUT_FIELDNAMES)
    inputs: dict[tuple[str, str], dict[str, Any]] = {}
    wood_capacity_by_state: dict[str, dict[str, Any]] = {}
    for wood_row in load_wood_target_capacity_rows():
        payload = wood_capacity_by_state.setdefault(
            wood_row["state"],
            {
                "effective_area_total": 0.0,
                "citation_1_title": "",
                "citation_1_url": "",
                "citation_1_locator": "",
                "citation_2_title": "",
                "citation_2_url": "",
                "citation_2_locator": "",
            },
        )
        payload["effective_area_total"] += float(wood_row["effective_area_ha"])
        if not payload["citation_1_title"]:
            payload["citation_1_title"] = wood_row["citation_1_title"]
            payload["citation_1_url"] = wood_row["citation_1_url"]
            payload["citation_1_locator"] = wood_row["citation_1_locator"]
        if not payload["citation_2_title"]:
            payload["citation_2_title"] = wood_row["citation_2_title"]
            payload["citation_2_url"] = wood_row["citation_2_url"]
            payload["citation_2_locator"] = wood_row["citation_2_locator"]
    for row in read_csv_rows(RAW_ADJUSTMENT_INPUTS_CSV):
        resource = row["resource"]
        parsed = {
            "problem_type": row["problem_type"],
            "output_addition_y": parse_float(row["output_addition_y"]) or 0.0,
            "plausibility_haircut_z": parse_float(row["plausibility_haircut_z"]) or 0.0,
            "minimum_operating_floor_cap": parse_int(row["minimum_operating_floor_cap"]),
            "exception_status": row["exception_status"],
            "exception_final_cap": parse_int(row["exception_final_cap"]),
            "representative_max_year": row.get("representative_max_year", ""),
            "representative_max_output": parse_float(row.get("representative_max_output")) if row.get("representative_max_output") not in (None, "") else "",
            "year_selection_reason": row.get("year_selection_reason", ""),
            "counterevidence_status": row.get("counterevidence_status", ""),
            "counterevidence_note": row.get("counterevidence_note", ""),
            "plausibility_1936_status": row.get("plausibility_1936_status", ""),
            "observed_output_x_reason": row.get("observed_output_x_reason", ""),
            "y_basis": row.get("y_basis", ""),
            "y_reason": row.get("y_reason", ""),
            "y_counterevidence_trigger": row.get("y_counterevidence_trigger", ""),
            "y_quantification_method": row.get("y_quantification_method", ""),
            "z_reason": row.get("z_reason", ""),
            "audit_class": row.get("audit_class", ""),
            "adjustment_reason": row["adjustment_reason"],
            "proxy_kind": row.get("proxy_kind", ""),
            "proxy_native_unit": row.get("proxy_native_unit", ""),
            "proxy_to_slot_note": row.get("proxy_to_slot_note", ""),
            "citation_1_title": row["citation_1_title"],
            "citation_1_url": row["citation_1_url"],
            "citation_1_locator": row["citation_1_locator"],
            "citation_2_title": row["citation_2_title"],
            "citation_2_url": row["citation_2_url"],
            "citation_2_locator": row["citation_2_locator"],
            "calculation_note": row["calculation_note"],
        }
        if resource == "Arable Land":
            parsed.update(
                {
                    "problem_type": "",
                    "output_addition_y": 0.0,
                    "plausibility_haircut_z": 0.0,
                    "minimum_operating_floor_cap": None,
                    "exception_status": "",
                    "exception_final_cap": None,
                    "y_basis": "",
                    "y_reason": "No upward addition.",
                    "y_counterevidence_trigger": "No counterevidence-triggered addition.",
                    "y_quantification_method": "No upward addition.",
                    "z_reason": "No downward plausibility haircut.",
                    "audit_class": "direct",
                    "adjustment_reason": "Land-capacity model uses direct effective commercial hectares; legacy output-gap uplift is retired.",
                    "calculation_note": "Arable now uses direct land-capacity x with no legacy output-gap Y or GDP/output smoothing.",
                }
            )
        elif resource == "Wood":
            wood_seed = wood_capacity_by_state.get(row["state"], {})
            if not parsed["exception_status"] and float(wood_seed.get("effective_area_total", 0.0)) == 0.0:
                parsed["exception_status"] = "manual_zero_exception"
                parsed["exception_final_cap"] = 0
            if not parsed["citation_1_title"]:
                parsed["citation_1_title"] = wood_seed.get("citation_1_title", "")
                parsed["citation_1_url"] = wood_seed.get("citation_1_url", "")
                parsed["citation_1_locator"] = wood_seed.get("citation_1_locator", "")
            if not parsed["citation_2_title"]:
                parsed["citation_2_title"] = wood_seed.get("citation_2_title", "")
                parsed["citation_2_url"] = wood_seed.get("citation_2_url", "")
                parsed["citation_2_locator"] = wood_seed.get("citation_2_locator", "")
            parsed.update(
                {
                    "problem_type": "" if parsed["exception_status"] != "manual_zero_exception" else parsed["problem_type"],
                    "output_addition_y": parsed["output_addition_y"],
                    "plausibility_haircut_z": parsed["plausibility_haircut_z"],
                    "representative_max_year": 1936,
                    "year_selection_reason": "Representative year is audited metadata only for the forestry-capacity model; wood no longer uses GDP-selected plantation-estate rows.",
                    "observed_output_x_reason": "Wood now uses direct effective commercial forestry hectares from audited wood land-class rows rather than observed plantation-estate area.",
                    "y_basis": parsed["y_basis"],
                    "y_reason": parsed["y_reason"] or "No upward addition.",
                    "y_counterevidence_trigger": parsed["y_counterevidence_trigger"] or "No counterevidence-triggered addition.",
                    "y_quantification_method": parsed["y_quantification_method"] or "No upward addition.",
                    "z_reason": parsed["z_reason"] or "No downward plausibility haircut.",
                    "audit_class": parsed["audit_class"] or ("exception" if parsed["exception_status"] else "direct"),
                    "adjustment_reason": (
                        "Potential-forestry model keeps this row as an explicit zero exception; wooded land alone does not count as commercial forestry."
                        if parsed["exception_status"]
                        else "Potential-forestry model replaces observed plantation estate area with direct effective commercial forestry hectares."
                    ),
                    "calculation_note": (
                        "Explicit zero retained under the potential-forestry model because the audited footprint lacks a defensible commercial forestry belt."
                        if parsed["exception_status"]
                        else "Wood now uses direct effective commercial forestry hectares with plantation-first potential and capped restoration allowances."
                    ),
                }
            )
        inputs[(row["state"], resource)] = parsed
    return inputs


def build_base_state_resource_rows(
    live_values: dict[str, dict[str, Any]],
    vanilla_priors: dict[str, dict[str, Any]],
    arable_baskets: dict[str, dict[str, str]],
    selected_target_rows: dict[tuple[str, str], dict[str, Any]],
    target_selection_map: dict[tuple[str, str], dict[str, Any]],
    resource_denominators: dict[str, dict[str, Any]],
    shared_arable_denominator: dict[str, Any],
    arable_state_payloads: dict[str, dict[str, Any]],
    wood_state_payloads: dict[str, dict[str, Any]],
    gdp_reference: dict[str, Any],
    gdp_series: dict[str, dict[str, Any]],
    gdp_map: dict[tuple[str, str], dict[str, str]],
) -> dict[tuple[str, str], dict[str, Any]]:
    rows: dict[tuple[str, str], dict[str, Any]] = {}
    shared_arable = parse_float(shared_arable_denominator["shared_sb_arable_effective_hectares_per_cap"])
    for state_info in STATE_INFO:
        state = state_info["official_name"]
        state_gdp_profile = gdp_profile_for_entity(state, "target_state", gdp_reference, gdp_series, gdp_map)
        for _category, resource in NUMERIC_RESOURCES:
            family = resource_family(resource)
            current_live = live_values[state].get(resource, 0)
            vanilla = vanilla_priors[state].get(resource, 0)
            if resource == "Arable Land":
                observed_x = arable_state_payloads[state]["observed_output_x"]
                denominator = shared_arable
                denominator_status = "formula-driven" if denominator not in (None, 0) else "denominator_unavailable"
                gdp_fields = {
                    "gdp_geography_used": "",
                    "gdp_geography_level": "",
                    "gdp_anchor_year": "",
                    "gdp_value_at_anchor_year": "",
                    "selected_resource_year": "",
                    "selection_mode": "",
                    "gdp_distance_ratio": "",
                    "gdp_comparability_status": "",
                    "target_estimation_method": "",
                    "interpolation_used": "",
                    "comparator_included_in_denominator": arable_state_payloads[state].get("valid_comparator_count", ""),
                    "comparator_basket_coverage_share": "",
                    "selection_note": "Arable land is audited from direct effective commercial agricultural hectares rather than GDP-selected crop output.",
                }
            elif resource == "Wood":
                observed_x = wood_state_payloads[state]["observed_output_x"]
                denom_payload = resource_denominators.get(family, {})
                denominator = parse_float(denom_payload.get("denominator_units_per_cap"))
                denominator_status = denom_payload.get("status", "denominator_unavailable")
                gdp_fields = {
                    "gdp_geography_used": "",
                    "gdp_geography_level": "",
                    "gdp_anchor_year": "",
                    "gdp_value_at_anchor_year": "",
                    "selected_resource_year": "",
                    "selection_mode": "",
                    "gdp_distance_ratio": "",
                    "gdp_comparability_status": "",
                    "target_estimation_method": "",
                    "interpolation_used": "",
                    "comparator_included_in_denominator": resource_denominators.get("Wood", {}).get("valid_comparator_count", ""),
                    "comparator_basket_coverage_share": "",
                    "selection_note": "Wood is audited from direct effective commercial forestry hectares rather than GDP-selected plantation-estate observations.",
                }
            else:
                selected_target = selected_target_rows.get((state, resource))
                selection_meta = target_selection_map.get((state, resource), {})
                if selected_target is not None and selected_target.get("drives_x") == "yes":
                    observed_x = float(selected_target.get("discounted_normalized_1950_output", selected_target["normalized_1950_output"]))
                else:
                    observed_x = None
                denom_payload = resource_denominators.get(family, {})
                denominator = parse_float(denom_payload.get("denominator_units_per_cap"))
                denominator_status = denom_payload.get("status", "denominator_unavailable")
                gdp_fields = {field: selection_meta.get(field, "") for field in GDP_SELECTION_FIELDS}
                if not gdp_fields.get("selection_mode"):
                    gdp_fields = {
                        "gdp_geography_used": state_gdp_profile["gdp_geography_used"],
                        "gdp_geography_level": state_gdp_profile["gdp_geography_level"],
                        "gdp_anchor_year": state_gdp_profile["gdp_anchor_year"],
                        "gdp_value_at_anchor_year": state_gdp_profile["gdp_value_at_anchor_year"],
                        "selected_resource_year": "",
                        "selection_mode": state_gdp_profile["selection_mode"],
                        "gdp_distance_ratio": "",
                        "gdp_comparability_status": "",
                        "target_estimation_method": "",
                        "interpolation_used": "no",
                        "comparator_included_in_denominator": "",
                        "comparator_basket_coverage_share": "",
                        "selection_note": "No direct target observation is frozen for this row; the state GDP anchor is recorded for audit context only.",
                    }
            base_cap = (observed_x / denominator) if (observed_x not in (None, "") and denominator not in (None, 0)) else None
            rows[(state, resource)] = {
                "state": state,
                "state_id": state_info["state_id"],
                "resource": resource,
                "resource_family": family if resource != "Arable Land" else "Arable Land",
                "vanilla_proxy_state": state_info["vanilla_proxy_name"],
                "vanilla_proxy_id": state_info["vanilla_proxy_id"],
                "vanilla_cap": vanilla if vanilla != "" else 0,
                "current_live_cap": current_live,
                "observed_output_x": observed_x if observed_x is not None else "",
                "denominator_units_per_cap": denominator if denominator is not None else "",
                "denominator_status": denominator_status,
                "base_cap": base_cap if base_cap is not None else "",
                "adjustment_key": "",
                "output_addition_y": 0.0,
                "plausibility_haircut_z": 0.0,
                "adjusted_output": observed_x if observed_x is not None else "",
                "adjusted_cap": base_cap if base_cap is not None else "",
                "minimum_operating_floor_cap": "",
                "final_audited_cap": current_live,
                "status": "formula-driven" if denominator_status == "formula-driven" else "denominator_unavailable",
                "exception_status": "",
                "problem_type": "",
                "adjustment_reason": "",
                "citation_1_title": "",
                "citation_1_url": "",
                "citation_1_locator": "",
                "citation_2_title": "",
                "citation_2_url": "",
                "citation_2_locator": "",
                "calculation_note": "",
                "proxy_kind": "",
                "proxy_native_unit": "",
                "proxy_to_slot_note": "",
                "y_basis": "",
                "y_reason": "",
                "y_counterevidence_trigger": "",
                "y_quantification_method": "",
                "researched_basket_size": arable_state_payloads[state].get("researched_basket_size", "") if resource == "Arable Land" else "",
                "live_enabled_count": arable_state_payloads[state].get("live_enabled_count", "") if resource == "Arable Land" else "",
                "overlap_count": arable_state_payloads[state].get("overlap_count", "") if resource == "Arable Land" else "",
                "researched_only_count": arable_state_payloads[state].get("researched_only_count", "") if resource == "Arable Land" else "",
                "live_only_count": arable_state_payloads[state].get("live_only_count", "") if resource == "Arable Land" else "",
                "researched_only_output_share": arable_state_payloads[state].get("researched_only_resource_share", "") if resource == "Arable Land" else "",
                "mismatch_status": arable_state_payloads[state].get("mismatch_status", "") if resource == "Arable Land" else "",
                **gdp_fields,
            }
    return rows


def apply_adjustments(
    base_rows: dict[tuple[str, str], dict[str, Any]],
    adjustment_inputs: dict[tuple[str, str], dict[str, Any]],
    counterevidence_cases: dict[tuple[str, str], list[dict[str, str]]],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, int]], list[dict[str, Any]]]:
    resource_adjustment_rows: list[dict[str, Any]] = []
    final_caps: dict[str, dict[str, int]] = {row["official_name"]: {} for row in STATE_INFO}
    audit_rows: list[dict[str, Any]] = []
    for state_info in STATE_INFO:
        state = state_info["official_name"]
        for _category, resource in NUMERIC_RESOURCES:
            base = dict(base_rows[(state, resource)])
            input_row = adjustment_inputs.get((state, resource), {})
            row_cases = counterevidence_cases.get((state, resource), [])
            y_addition = input_row.get("output_addition_y", 0.0)
            z = input_row.get("plausibility_haircut_z", 0.0)
            minimum_floor = input_row.get("minimum_operating_floor_cap")
            exception_status = input_row.get("exception_status", "")
            denominator = parse_float(base["denominator_units_per_cap"])
            observed = parse_float(base["observed_output_x"]) or 0.0
            base_cap = parse_float(base["base_cap"])
            proxy_kind = input_row.get("proxy_kind", "")
            proxy_native_unit = input_row.get("proxy_native_unit", "")
            proxy_to_slot_note = input_row.get("proxy_to_slot_note", "")
            representative_max_year = input_row.get("representative_max_year", "")
            representative_max_output = input_row.get("representative_max_output", "")
            year_selection_reason = input_row.get("year_selection_reason", "")
            observed_output_x_reason = input_row.get("observed_output_x_reason", "")
            if resource == "Wood":
                representative_max_year = 1936
                representative_max_output = observed
                year_selection_reason = input_row.get("year_selection_reason", "") or "Representative year is audited metadata only for the forestry-capacity model; wood no longer uses GDP-selected plantation-estate rows."
                observed_output_x_reason = input_row.get("observed_output_x_reason", "") or "Wood now uses direct effective commercial forestry hectares from audited wood land-class rows."
            if not proxy_kind and (state, resource) in PROXY_NOTES:
                proxy_kind, proxy_native_unit, proxy_to_slot_note = PROXY_NOTES[(state, resource)]
            adjustment_key = ""
            has_adjustment_terms = any(abs(value) > 0 for value in [y_addition, z]) or bool(minimum_floor)
            if denominator in (None, 0.0) and not exception_status:
                exception_status = "denominator_unavailable"
            if exception_status:
                final_cap = input_row.get("exception_final_cap")
                if final_cap is None:
                    final_cap = int(base["current_live_cap"])
                status = "denominator unavailable" if exception_status == "denominator_unavailable" else "explicit exception"
                adjusted_output = ""
                adjusted_cap = ""
            else:
                adjusted_output = observed + y_addition - z
                adjusted_cap = adjusted_output / denominator
                rounded_cap = round(adjusted_cap)
                final_cap = max(rounded_cap, minimum_floor) if minimum_floor else rounded_cap
                if final_cap == 0 and observed == 0 and not has_adjustment_terms:
                    counterevidence_status, _counterevidence_note = summarize_counterevidence(row_cases, final_cap, exception_status)
                    status = "constrained zero" if counterevidence_status == "supports" else "review required"
                    if status == "review required" and int(base["current_live_cap"]) == 0:
                        exception_status = "manual_zero_exception"
                        status = "explicit exception"
                elif has_adjustment_terms:
                    status = "quantified adjustment"
                else:
                    status = "formula-driven"
            if has_adjustment_terms or exception_status or status in {"constrained zero", "review required"}:
                adjustment_key = f"{state_slug(state)}::{resource_slug(resource)}"
                audit_row = {
                    "adjustment_key": adjustment_key,
                    "state": state,
                    "resource": resource,
                    "family": base["resource_family"],
                    "observed_output_x": observed,
                    "output_addition_y": y_addition,
                    "plausibility_haircut_z": z,
                    "adjusted_output": adjusted_output,
                    "denominator_units_per_cap": denominator if denominator is not None else "",
                    "base_cap": base_cap if base_cap is not None else "",
                    "adjusted_cap": adjusted_cap if adjusted_cap != "" else "",
                    "minimum_operating_floor_cap": minimum_floor or "",
                    "final_audited_cap": final_cap,
                    "exception_status": exception_status,
                    "problem_type": input_row.get("problem_type", ""),
                    "adjustment_reason": input_row.get("adjustment_reason", ""),
                    "representative_max_year": representative_max_year,
                    "representative_max_output": representative_max_output,
                    "year_selection_reason": year_selection_reason,
                    "counterevidence_status": input_row.get("counterevidence_status", ""),
                    "counterevidence_note": input_row.get("counterevidence_note", ""),
                    "plausibility_1936_status": input_row.get("plausibility_1936_status", ""),
                    "observed_output_x_reason": observed_output_x_reason,
                    "y_basis": input_row.get("y_basis", ""),
                    "y_reason": input_row.get("y_reason", ""),
                    "y_counterevidence_trigger": input_row.get("y_counterevidence_trigger", ""),
                    "y_quantification_method": input_row.get("y_quantification_method", ""),
                    "z_reason": input_row.get("z_reason", ""),
                    "audit_class": input_row.get("audit_class", ""),
                    "proxy_kind": proxy_kind,
                    "proxy_native_unit": proxy_native_unit,
                    "proxy_to_slot_note": proxy_to_slot_note,
                    "citation_1_title": input_row.get("citation_1_title", ""),
                    "citation_1_url": input_row.get("citation_1_url", ""),
                    "citation_1_locator": input_row.get("citation_1_locator", ""),
                    "citation_2_title": input_row.get("citation_2_title", ""),
                    "citation_2_url": input_row.get("citation_2_url", ""),
                    "citation_2_locator": input_row.get("citation_2_locator", ""),
                    "calculation_note": input_row.get("calculation_note", ""),
                    **{field: base.get(field, "") for field in GDP_SELECTION_FIELDS},
                }
                audit_rows.append(audit_row)
            base.update(
                {
                    "adjustment_key": adjustment_key,
                    "output_addition_y": y_addition,
                    "plausibility_haircut_z": z,
                    "adjusted_output": adjusted_output,
                    "adjusted_cap": adjusted_cap,
                    "minimum_operating_floor_cap": minimum_floor or "",
                    "final_audited_cap": final_cap,
                    "status": status,
                    "exception_status": exception_status,
                    "problem_type": input_row.get("problem_type", ""),
                    "adjustment_reason": input_row.get("adjustment_reason", ""),
                    "representative_max_year": representative_max_year,
                    "representative_max_output": representative_max_output,
                    "year_selection_reason": year_selection_reason,
                    "counterevidence_status": input_row.get("counterevidence_status", ""),
                    "counterevidence_note": input_row.get("counterevidence_note", ""),
                    "plausibility_1936_status": input_row.get("plausibility_1936_status", ""),
                    "observed_output_x_reason": observed_output_x_reason,
                    "y_basis": input_row.get("y_basis", ""),
                    "y_reason": input_row.get("y_reason", ""),
                    "y_counterevidence_trigger": input_row.get("y_counterevidence_trigger", ""),
                    "y_quantification_method": input_row.get("y_quantification_method", ""),
                    "z_reason": input_row.get("z_reason", ""),
                    "audit_class": input_row.get("audit_class", ""),
                    "calculation_note": input_row.get("calculation_note", ""),
                    "proxy_kind": proxy_kind,
                    "proxy_native_unit": proxy_native_unit,
                    "proxy_to_slot_note": proxy_to_slot_note,
                    "citation_1_title": input_row.get("citation_1_title", ""),
                    "citation_1_url": input_row.get("citation_1_url", ""),
                    "citation_1_locator": input_row.get("citation_1_locator", ""),
                    "citation_2_title": input_row.get("citation_2_title", ""),
                    "citation_2_url": input_row.get("citation_2_url", ""),
                    "citation_2_locator": input_row.get("citation_2_locator", ""),
                    **{field: base.get(field, "") for field in GDP_SELECTION_FIELDS},
                }
            )
            resource_adjustment_rows.append(base)
            final_caps[state][resource] = int(final_cap)
    return resource_adjustment_rows, final_caps, audit_rows


def compute_priority_rows(resource_adjustment_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    priorities: list[dict[str, Any]] = []
    for row in resource_adjustment_rows:
        issue_type = ""
        priority = ""
        base_cap = parse_float(row["base_cap"])
        adjusted_cap = parse_float(row["adjusted_cap"])
        live_delta = int(row["final_audited_cap"]) - int(row["current_live_cap"])
        manual_gdp = row.get("selection_mode") == "manual_override"
        weak_gdp = parse_float(row.get("gdp_distance_ratio")) not in (None, "") and float(row["gdp_distance_ratio"]) > 0.35
        national_fallback = row.get("gdp_geography_level") == "national" and row["state"] != row.get("gdp_geography_used", "")
        basket_mismatch = row["resource"] == "Arable Land" and row.get("mismatch_status") in {"mismatch", "large_mismatch"}
        large_basket_mismatch = row["resource"] == "Arable Land" and row.get("mismatch_status") == "large_mismatch"
        if row["status"] == "review required":
            priority = "P1"
            issue_type = "unsupported zero or missing target audit"
        elif row["status"] == "explicit exception":
            if int(row["final_audited_cap"]) == 0 and int(row["current_live_cap"]) == 0 and not (manual_gdp or weak_gdp or national_fallback):
                continue
            priority = "P1"
            issue_type = "explicit audit exception"
        elif row["exception_status"] == "denominator_unavailable":
            priority = "P1"
            issue_type = "denominator unavailable"
        elif row["status"] == "constrained zero":
            if int(row["current_live_cap"]) == 0 and not (manual_gdp or weak_gdp or national_fallback):
                continue
            priority = "P2"
            issue_type = "constrained zero"
        elif parse_float(row["plausibility_haircut_z"]) not in (None, 0.0):
            priority = "P1"
            issue_type = "downward plausibility haircut"
        elif large_basket_mismatch:
            priority = "P1"
            issue_type = "arable audit/gameplay mismatch"
        elif abs(live_delta) >= 5:
            priority = "P1"
            issue_type = "material cap change vs live"
        elif parse_float(row["output_addition_y"]) not in (None, 0.0) and parse_float(row["plausibility_haircut_z"]) not in (None, 0.0):
            priority = "P1"
            issue_type = "upward addition plus haircut"
        elif row.get("y_basis") == "mixed":
            priority = "P1"
            issue_type = "mixed Y basis"
        elif parse_float(row["output_addition_y"]) not in (None, 0.0):
            priority = "P2"
            issue_type = "output addition"
        elif basket_mismatch:
            priority = "P2"
            issue_type = "arable audit/gameplay mismatch"
        elif manual_gdp:
            priority = "P2"
            issue_type = "manual GDP selection override"
        elif weak_gdp:
            priority = "P2"
            issue_type = "weak GDP match"
        elif national_fallback:
            priority = "P3"
            issue_type = "national GDP fallback on substate row"
        elif row["status"] != "formula-driven":
            priority = "P3"
            issue_type = row["status"]
        if not priority:
            continue
        ratio = ""
        if base_cap not in (None, 0.0) and adjusted_cap not in (None, 0.0, ""):
            ratio = adjusted_cap / base_cap
        priorities.append(
            {
                "priority": priority,
                "state": row["state"],
                "resource": row["resource"],
                "resource_family": row["resource_family"],
                "status": row["status"],
                "base_cap": base_cap if base_cap is not None else "",
                "final_audited_cap": row["final_audited_cap"],
                "adjusted_to_base_ratio": ratio,
                "issue_type": issue_type,
                "adjustment_key": row["adjustment_key"],
                "note": row["adjustment_reason"] or row["calculation_note"] or row["status"],
            }
        )
    order = {"P1": 0, "P2": 1, "P3": 2}
    priorities.sort(key=lambda row: (order[row["priority"]], row["state"], row["resource"]))
    return priorities


def build_row_audit_rows(
    resource_adjustment_rows: list[dict[str, Any]],
    selected_target_rows: dict[tuple[str, str], dict[str, Any]],
    counterevidence_cases: dict[tuple[str, str], list[dict[str, str]]],
    family_constants: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    constant_map = {row["family"]: row for row in family_constants}
    rows: list[dict[str, Any]] = []
    for row in resource_adjustment_rows:
        state = row["state"]
        resource = row["resource"]
        cases = counterevidence_cases.get((state, resource), [])
        representative = selected_target_rows.get((state, resource))
        if resource == "Arable Land":
            representative_year = row.get("representative_max_year", "") or 1936
            representative_output: Any = row["observed_output_x"]
            year_reason = row.get("year_selection_reason") or row.get("selection_note") or "Representative year is audited metadata only for the land-capacity model; arable no longer uses GDP-selected crop-output rows."
            observed_reason = row.get("observed_output_x_reason") or "Observed x is the summed effective commercial agricultural hectares from audited land-class rows."
        elif resource == "Wood":
            representative_year = row.get("representative_max_year", "") or 1936
            representative_output = row["observed_output_x"]
            year_reason = row.get("year_selection_reason") or row.get("selection_note") or "Representative year is audited metadata only for the forestry-capacity model; wood no longer uses GDP-selected plantation-estate rows."
            observed_reason = row.get("observed_output_x_reason") or "Observed x is the summed effective commercial forestry hectares from audited wood land-class rows."
        elif representative is not None:
            representative_year = representative["year"]
            representative_output = representative.get("discounted_normalized_1950_output", representative["normalized_1950_output"])
            year_reason = row.get("year_selection_reason") or row.get("selection_note") or "Representative year selected through the GDP-anchored audit rule."
            observed_reason = row.get("observed_output_x_reason") or representative.get("note") or representative.get("locator") or ""
        else:
            representative_year = row.get("representative_max_year", "") or "none"
            representative_output = row.get("representative_max_output", "") or row["observed_output_x"]
            year_reason = row.get("year_selection_reason") or row.get("selection_note") or "No direct target observation is currently frozen for this row; the cap remains under explicit audit handling."
            observed_reason = row.get("observed_output_x_reason") or "No direct target observation is currently frozen for this row."

        counterevidence_status = row.get("counterevidence_status", "")
        counterevidence_note = row.get("counterevidence_note", "")
        if not counterevidence_status:
            counterevidence_status, counterevidence_note = summarize_counterevidence(
                cases,
                int(row["final_audited_cap"]),
                str(row.get("exception_status", "")),
            )

        if row.get("audit_class"):
            audit_class = row["audit_class"]
        elif row.get("exception_status"):
            audit_class = "exception"
        elif row.get("proxy_kind"):
            audit_class = "proxy"
        elif parse_float(row.get("plausibility_haircut_z")) not in (None, 0.0):
            audit_class = "plausibility_clamped"
        elif parse_float(row.get("output_addition_y")) not in (None, 0.0):
            audit_class = "direct_plus_gapfill"
        else:
            audit_class = "direct"

        if row.get("plausibility_1936_status"):
            plausibility_status = row["plausibility_1936_status"]
        elif parse_float(row.get("plausibility_haircut_z")) not in (None, 0.0):
            plausibility_status = "clamped"
        elif row["status"] in {"review required", "explicit exception"}:
            plausibility_status = "review_required"
        elif row["status"] == "constrained zero":
            plausibility_status = "constrained_zero"
        else:
            plausibility_status = "pass"

        constant_row = constant_map.get(row["resource_family"])
        rows.append(
            {
                "state": state,
                "resource": resource,
                "resource_family": row["resource_family"],
                "status": row["status"],
                "final_audited_cap": row["final_audited_cap"],
                "representative_max_year": representative_year,
                "representative_max_output": representative_output,
                "year_selection_reason": year_reason,
                "counterevidence_status": counterevidence_status,
                "counterevidence_note": counterevidence_note,
                "plausibility_1936_status": plausibility_status,
                "observed_output_x_reason": observed_reason,
                "y_basis": row.get("y_basis", ""),
                "y_reason": row.get("y_reason") or ("No upward addition." if parse_float(row.get("output_addition_y")) in (None, 0.0) else row.get("adjustment_reason", "")),
                "y_counterevidence_trigger": row.get("y_counterevidence_trigger") or ("No counterevidence-triggered addition." if parse_float(row.get("output_addition_y")) in (None, 0.0) else row.get("adjustment_reason", "")),
                "y_quantification_method": row.get("y_quantification_method") or ("No upward addition." if parse_float(row.get("output_addition_y")) in (None, 0.0) else row.get("calculation_note", "") or row.get("adjustment_reason", "")),
                "z_reason": row.get("z_reason") or ("No downward plausibility haircut." if parse_float(row.get("plausibility_haircut_z")) in (None, 0.0) else row.get("adjustment_reason", "")),
                "audit_class": audit_class,
                "family_normalising_constant": constant_row.get("value", "") if constant_row else "",
                "family_constant_purpose": constant_row.get("purpose", "") if constant_row else "",
                "adjustment_key": row["adjustment_key"],
                "exception_status": row.get("exception_status", ""),
                "problem_type": row.get("problem_type", ""),
                "adjustment_reason": row.get("adjustment_reason", ""),
                "citation_1_title": row.get("citation_1_title", ""),
                "citation_1_url": row.get("citation_1_url", ""),
                "citation_1_locator": row.get("citation_1_locator", ""),
                "citation_2_title": row.get("citation_2_title", ""),
                "citation_2_url": row.get("citation_2_url", ""),
                "citation_2_locator": row.get("citation_2_locator", ""),
                **{field: row.get(field, "") for field in GDP_SELECTION_FIELDS if field != "selected_resource_year"},
            }
        )
    return rows


def build_regional_totals(vanilla_totals: dict[str, dict[str, int]], final_caps: dict[str, dict[str, int]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for region, states in REGIONS.items():
        for resource in SUMMARY_RESOURCES:
            vanilla_total = vanilla_totals[region].get(resource, 0)
            final_total = sum(final_caps[state].get(resource, 0) for state in states)
            rows.append(
                {
                    "region": region,
                    "resource": resource,
                    "vanilla_total": vanilla_total,
                    "final_sb_total": final_total,
                    "absolute_delta": final_total - vanilla_total,
                    "percentage_delta": ((final_total - vanilla_total) / vanilla_total) if vanilla_total else "",
                }
            )
    return rows


def cleanup_legacy_outputs() -> None:
    for path in OLD_PUBLIC_OUTPUTS:
        if path.exists():
            path.unlink()
    for path in PUBLIC_ROOT.rglob(".DS_Store"):
        path.unlink()


def write_public_outputs(
    growth_rows: list[dict[str, Any]],
    target_observations: list[dict[str, Any]],
    comparator_observations: list[dict[str, Any]],
    target_data_validation_rows: list[dict[str, Any]],
    gdp_selection_rows: list[dict[str, Any]],
    family_normalising_constants: list[dict[str, Any]],
    resource_denominator_rows: list[dict[str, Any]],
    arable_target_capacity_rows: list[dict[str, Any]],
    arable_comparator_capacity_rows: list[dict[str, Any]],
    arable_state_mean_rows: list[dict[str, Any]],
    arable_shared_denominator: dict[str, Any],
    arable_resource_expectation_rows: list[dict[str, Any]],
    arable_comparator_diagnostics: list[dict[str, Any]],
    wood_target_capacity_rows: list[dict[str, Any]],
    wood_comparator_capacity_rows: list[dict[str, Any]],
    resource_adjustment_rows: list[dict[str, Any]],
    regional_total_rows: list[dict[str, Any]],
    audit_rows: list[dict[str, Any]],
    row_audit_rows: list[dict[str, Any]],
    priority_rows: list[dict[str, Any]],
    state_regional_advantage_rows: list[dict[str, Any]],
    state_review_status_rows: list[dict[str, Any]],
) -> None:
    cleanup_legacy_outputs()
    DERIVED_DIR.mkdir(parents=True, exist_ok=True)
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    state_summary_rows = build_state_summary_rows(resource_adjustment_rows)
    state_resource_delta_rows = build_state_resource_delta_rows(resource_adjustment_rows)
    write_csv(
        DERIVED_DIR / "target_observations.csv",
        target_observations,
        [
            "sheet_key",
            "official_state",
            "entity",
            "geography",
            "short_geography",
            "resource",
            "resource_family",
            "observation_family",
            "year",
            "raw_quantity",
            "raw_unit",
            "standardized_quantity",
            "standardized_unit",
            "sector",
            "growth_family",
            "index_to_1950",
            "conversion_key",
            "conversion_factor",
            "normalized_1950_output",
            "wheat_equivalent_1950_output",
            "discounted_normalized_1950_output",
            "discounted_wheat_equivalent_1950_output",
            "drives_x",
            "review_action",
            "corroboration_count",
            *TARGET_VALIDATION_FIELDS,
            "source_title",
            "source_url",
            "locator",
            "note",
        ],
    )
    write_csv(
        DERIVED_DIR / "comparator_observations.csv",
        comparator_observations,
        [
            "sheet_key",
            "official_state",
            "entity",
            "geography",
            "short_geography",
            "resource",
            "resource_family",
            "observation_family",
            "year",
            "raw_quantity",
            "raw_unit",
            "standardized_quantity",
            "standardized_unit",
            "sector",
            "growth_family",
            "index_to_1950",
            "conversion_key",
            "conversion_factor",
            "normalized_1950_output",
            "wheat_equivalent_1950_output",
            "discounted_normalized_1950_output",
            "discounted_wheat_equivalent_1950_output",
            "drives_x",
            "review_action",
            "corroboration_count",
            *TARGET_VALIDATION_FIELDS,
            "source_title",
            "source_url",
            "locator",
            "note",
        ],
    )
    write_csv(
        AUDIT_DIR / "target_data_validation.csv",
        target_data_validation_rows,
        TARGET_DATA_VALIDATION_FIELDNAMES,
    )
    write_csv(
        DERIVED_DIR / "gdp_selection_rows.csv",
        gdp_selection_rows,
        [
            "entity",
            "resource",
            "gdp_geography_used",
            "gdp_geography_level",
            "gdp_reference_value",
            "gdp_anchor_year",
            "gdp_value_at_anchor_year",
            "selected_resource_year",
            "selection_mode",
            "gdp_distance_ratio",
            "gdp_comparability_status",
            "target_estimation_method",
            "interpolation_used",
            "comparator_included_in_denominator",
            "comparator_basket_coverage_share",
            "selection_note",
        ],
    )
    write_csv(
        RAW_FAMILY_CONSTANTS_CSV,
        family_normalising_constants,
        [
            "family",
            "value",
            "purpose",
            "citation_1_title",
            "citation_1_url",
            "citation_1_locator",
            "citation_2_title",
            "citation_2_url",
            "citation_2_locator",
        ],
    )
    write_csv(
        DERIVED_DIR / "resource_denominators.csv",
        resource_denominator_rows,
        ["resource_family", "valid_comparator_count", "denominator_units_per_cap", "status", "method"],
    )
    write_csv(
        DERIVED_DIR / "arable_target_capacity_rows.csv",
        arable_target_capacity_rows,
        ARABLE_TARGET_CAPACITY_FIELDNAMES,
    )
    write_csv(
        DERIVED_DIR / "arable_comparator_capacity_rows.csv",
        arable_comparator_capacity_rows,
        ARABLE_COMPARATOR_CAPACITY_FIELDNAMES,
    )
    write_csv(
        DERIVED_DIR / "arable_state_means.csv",
        arable_state_mean_rows,
        [
            "state",
            "target_effective_commercial_hectares",
            "weighted_state_effective_hectares_per_cap",
            "valid_comparator_count",
            "active_weight_sum",
            "researched_basket_size",
            "live_enabled_count",
            "overlap_count",
            "researched_only_count",
            "live_only_count",
            "researched_only_resource_share",
            "mismatch_status",
            "status",
        ],
    )
    write_csv(
        DERIVED_DIR / "arable_shared_denominator.csv",
        [arable_shared_denominator],
        [
            "shared_sb_arable_effective_hectares_per_cap",
            "state_mean_count",
            "method",
        ],
    )
    write_csv(
        DERIVED_DIR / "arable_comparator_diagnostics.csv",
        arable_comparator_diagnostics,
        [
            "target_state",
            "comparator",
            "rank",
            "weight",
            "effective_commercial_hectares",
            "effective_hectares_per_cap",
            "comparator_included_in_denominator",
            "included_reason",
            "capacity_note",
        ],
    )
    write_csv(
        DERIVED_DIR / "arable_resource_expectations.csv",
        arable_resource_expectation_rows,
        [
            "state",
            "resource",
            "researched_plausible",
            "live_enabled_in_state",
            "basket_membership_status",
            "audit_relevance",
            "gameplay_note",
        ],
    )
    write_csv(
        DERIVED_DIR / "wood_target_capacity_rows.csv",
        wood_target_capacity_rows,
        WOOD_TARGET_CAPACITY_FIELDNAMES,
    )
    write_csv(
        DERIVED_DIR / "wood_comparator_capacity_rows.csv",
        wood_comparator_capacity_rows,
        WOOD_COMPARATOR_CAPACITY_FIELDNAMES,
    )
    write_csv(
        DERIVED_DIR / "resource_adjustments.csv",
        resource_adjustment_rows,
        [
            "state",
            "state_id",
            "resource",
            "resource_family",
            "vanilla_proxy_state",
            "vanilla_proxy_id",
            "vanilla_cap",
            "current_live_cap",
            "observed_output_x",
            "output_addition_y",
            "plausibility_haircut_z",
            "adjusted_output",
            "denominator_units_per_cap",
            "denominator_status",
            "base_cap",
            "adjusted_cap",
            "minimum_operating_floor_cap",
            "final_audited_cap",
            "status",
            "exception_status",
            "adjustment_key",
            "problem_type",
            "adjustment_reason",
            "representative_max_year",
            "representative_max_output",
            "year_selection_reason",
            *GDP_SELECTION_FIELDS,
            "counterevidence_status",
            "counterevidence_note",
            "plausibility_1936_status",
            "observed_output_x_reason",
            "y_basis",
            "y_reason",
            "y_counterevidence_trigger",
            "y_quantification_method",
            "z_reason",
            "audit_class",
            "proxy_kind",
            "proxy_native_unit",
            "proxy_to_slot_note",
            "citation_1_title",
            "citation_1_url",
            "citation_1_locator",
            "citation_2_title",
            "citation_2_url",
            "citation_2_locator",
            "calculation_note",
        ],
    )
    write_csv(
        DERIVED_DIR / "final_resource_caps.csv",
        resource_adjustment_rows,
        [
            "state",
            "state_id",
            "resource",
            "resource_family",
            "vanilla_proxy_state",
            "vanilla_proxy_id",
            "vanilla_cap",
            "current_live_cap",
            "observed_output_x",
            "output_addition_y",
            "plausibility_haircut_z",
            "adjusted_output",
            "denominator_units_per_cap",
            "base_cap",
            "adjusted_cap",
            "final_audited_cap",
            "status",
            "exception_status",
            "adjustment_key",
            "problem_type",
            "adjustment_reason",
            "representative_max_year",
            "representative_max_output",
            "year_selection_reason",
            *GDP_SELECTION_FIELDS,
            "counterevidence_status",
            "counterevidence_note",
            "plausibility_1936_status",
            "observed_output_x_reason",
            "y_basis",
            "y_reason",
            "y_counterevidence_trigger",
            "y_quantification_method",
            "z_reason",
            "audit_class",
        ],
    )
    write_csv(
        DERIVED_DIR / "regional_resource_totals.csv",
        regional_total_rows,
        ["region", "resource", "vanilla_total", "final_sb_total", "absolute_delta", "percentage_delta"],
    )
    write_csv(
        DERIVED_DIR / "state_delta_summary.csv",
        state_summary_rows,
        [
            "state",
            "vanilla_total",
            "current_live_total",
            "final_audited_total",
            "delta_vs_vanilla",
            "delta_vs_live",
            "formula_rows",
            "adjusted_rows",
            "exception_rows",
        ],
    )
    write_csv(
        DERIVED_DIR / "state_resource_deltas.csv",
        state_resource_delta_rows,
        [
            "state",
            "resource",
            "resource_family",
            "vanilla_cap",
            "current_live_cap",
            "final_audited_cap",
            "delta_vs_vanilla",
            "delta_vs_live",
            "status",
            "exception_status",
            "audit_class",
            "adjustment_reason",
        ],
    )
    write_csv(
        AUDIT_DIR / "adjustments.csv",
        audit_rows,
        [
            "adjustment_key",
            "state",
            "resource",
            "family",
            "observed_output_x",
            "output_addition_y",
            "plausibility_haircut_z",
            "adjusted_output",
            "denominator_units_per_cap",
            "base_cap",
            "adjusted_cap",
            "minimum_operating_floor_cap",
            "final_audited_cap",
            "exception_status",
            "problem_type",
            "adjustment_reason",
            "representative_max_year",
            "representative_max_output",
            "year_selection_reason",
            *GDP_SELECTION_FIELDS,
            "counterevidence_status",
            "counterevidence_note",
            "plausibility_1936_status",
            "observed_output_x_reason",
            "y_basis",
            "y_reason",
            "y_counterevidence_trigger",
            "y_quantification_method",
            "z_reason",
            "audit_class",
            "proxy_kind",
            "proxy_native_unit",
            "proxy_to_slot_note",
            "citation_1_title",
            "citation_1_url",
            "citation_1_locator",
            "citation_2_title",
            "citation_2_url",
            "citation_2_locator",
            "calculation_note",
        ],
    )
    write_csv(
        AUDIT_DIR / "row_audit.csv",
        row_audit_rows,
        [
            "state",
            "resource",
            "resource_family",
            "status",
            "final_audited_cap",
            "representative_max_year",
            "representative_max_output",
            "year_selection_reason",
            "counterevidence_status",
            "counterevidence_note",
            "plausibility_1936_status",
            "observed_output_x_reason",
            "y_basis",
            "y_reason",
            "y_counterevidence_trigger",
            "y_quantification_method",
            "z_reason",
            "audit_class",
            "gdp_geography_used",
            "gdp_geography_level",
            "gdp_anchor_year",
            "gdp_value_at_anchor_year",
            "selection_mode",
            "gdp_distance_ratio",
            "selection_note",
            "family_normalising_constant",
            "family_constant_purpose",
            "adjustment_key",
            "exception_status",
            "problem_type",
            "adjustment_reason",
            "citation_1_title",
            "citation_1_url",
            "citation_1_locator",
            "citation_2_title",
            "citation_2_url",
            "citation_2_locator",
        ],
    )
    write_csv(
        AUDIT_DIR / "priority_rows.csv",
        priority_rows,
        ["priority", "state", "resource", "resource_family", "status", "base_cap", "final_audited_cap", "adjusted_to_base_ratio", "issue_type", "adjustment_key", "note"],
    )
    write_csv(
        AUDIT_DIR / "state_regional_advantages.csv",
        state_regional_advantage_rows,
        REGIONAL_ADVANTAGES_FIELDNAMES,
    )
    write_csv(
        AUDIT_DIR / "state_review_status.csv",
        state_review_status_rows,
        STATE_REVIEW_STATUS_FIELDNAMES,
    )
    (AUDIT_DIR / "sb_state_delta_report.md").write_text(
        build_state_delta_report_markdown(state_summary_rows, state_resource_delta_rows),
        encoding="utf-8",
    )
    write_csv(
        DERIVED_DIR / "growth_series.csv",
        growth_rows,
        ["family", "year", "quantity", "unit", "index_to_1950", "method", "source_title", "source_url", "locator"],
    )


def style_header(ws, row: int) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_subheader(ws, row: int) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(str(cell.value)) + 2, 60))
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(12, width)


def add_table(ws, start_row: int, title: str, headers: list[str], rows: list[dict[str, Any]], fields: list[str]) -> int:
    ws.cell(start_row, 1, title)
    style_subheader(ws, start_row)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(start_row + 1, col_idx, header)
    style_header(ws, start_row + 1)
    row_idx = start_row + 2
    if not rows:
        ws.cell(row_idx, 1, "No rows")
        return row_idx + 2
    for row in rows:
        for col_idx, field in enumerate(fields, start=1):
            ws.cell(row_idx, col_idx, row.get(field, ""))
        row_idx += 1
    return row_idx + 1


def add_key_value_block(ws, start_row: int, title: str, items: list[tuple[str, Any]]) -> int:
    ws.cell(start_row, 1, title)
    style_subheader(ws, start_row)
    ws.cell(start_row + 1, 1, "Field")
    ws.cell(start_row + 1, 2, "Value")
    style_header(ws, start_row + 1)
    row = start_row + 2
    for key, value in items:
        ws.cell(row, 1, key)
        ws.cell(row, 2, value)
        row += 1
    return row + 1


def build_state_summary_rows(resource_adjustment_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_state: dict[str, dict[str, Any]] = {}
    for row in resource_adjustment_rows:
        state = row["state"]
        payload = by_state.setdefault(
            state,
            {
                "state": state,
                "vanilla_total": 0,
                "current_live_total": 0,
                "final_audited_total": 0,
                "formula_rows": 0,
                "adjusted_rows": 0,
                "exception_rows": 0,
            },
        )
        payload["vanilla_total"] += int(row["vanilla_cap"])
        payload["current_live_total"] += int(row["current_live_cap"])
        payload["final_audited_total"] += int(row["final_audited_cap"])
        if row["status"] == "formula-driven":
            payload["formula_rows"] += 1
        elif row["status"] in {"denominator unavailable", "explicit exception", "review required", "constrained zero"}:
            payload["exception_rows"] += 1
        else:
            payload["adjusted_rows"] += 1
    rows: list[dict[str, Any]] = []
    for payload in by_state.values():
        payload["delta_vs_vanilla"] = payload["final_audited_total"] - payload["vanilla_total"]
        payload["delta_vs_live"] = payload["final_audited_total"] - payload["current_live_total"]
        rows.append(payload)
    rows.sort(key=lambda row: row["state"])
    return rows


def build_state_resource_delta_rows(resource_adjustment_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    state_order = {row["official_name"]: idx for idx, row in enumerate(STATE_INFO)}
    resource_order = {resource: idx for idx, resource in enumerate(SUMMARY_RESOURCES)}
    rows: list[dict[str, Any]] = []
    for row in resource_adjustment_rows:
        vanilla_cap = int(row["vanilla_cap"])
        current_live_cap = int(row["current_live_cap"])
        final_audited_cap = int(row["final_audited_cap"])
        rows.append(
            {
                "state": row["state"],
                "resource": row["resource"],
                "resource_family": row["resource_family"],
                "vanilla_cap": vanilla_cap,
                "current_live_cap": current_live_cap,
                "final_audited_cap": final_audited_cap,
                "delta_vs_vanilla": final_audited_cap - vanilla_cap,
                "delta_vs_live": final_audited_cap - current_live_cap,
                "status": row["status"],
                "exception_status": row["exception_status"],
                "audit_class": row["audit_class"],
                "adjustment_reason": row["adjustment_reason"],
            }
        )
    rows.sort(key=lambda row: (state_order[row["state"]], resource_order[row["resource"]]))
    return rows


def build_state_delta_report_markdown(
    state_summary_rows: list[dict[str, Any]],
    state_resource_delta_rows: list[dict[str, Any]],
) -> str:
    grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in state_resource_delta_rows:
        grouped_rows[row["state"]].append(row)
    lines = [
        "# SB State Resource Delta Report",
        "",
        "Generated from `final_resource_caps.csv`.",
        "",
        "This report compares vanilla caps, the current live file, and the audited result for every SB state/resource row.",
        "",
    ]
    for summary in state_summary_rows:
        state = summary["state"]
        lines.extend(
            [
                f"## {state}",
                "",
                f"Totals: vanilla `{summary['vanilla_total']}`, live `{summary['current_live_total']}`, audited `{summary['final_audited_total']}`, delta vs vanilla `{summary['delta_vs_vanilla']}`, delta vs live `{summary['delta_vs_live']}`.",
                "",
                "| Resource | Vanilla | Live | Audited | Delta vs vanilla | Delta vs live | Status |",
                "|---|---:|---:|---:|---:|---:|---|",
            ]
        )
        for row in grouped_rows[state]:
            lines.append(
                f"| {row['resource']} | {row['vanilla_cap']} | {row['current_live_cap']} | {row['final_audited_cap']} | {row['delta_vs_vanilla']} | {row['delta_vs_live']} | {row['status']} |"
            )
        lines.append("")
    return "\n".join(lines) + "\n"


def seed_regional_advantages_rows() -> list[dict[str, Any]]:
    return list(REGIONAL_ADVANTAGE_SEEDS)


def build_state_review_status_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for info in STATE_INFO:
        state = info["official_name"]
        review_status, largest_remaining_issue, summary_note = STATE_REVIEW_STATUS_SEEDS[state]
        rows.append(
            {
                "state": state,
                "review_status": review_status,
                "largest_remaining_issue": largest_remaining_issue,
                "summary_note": summary_note,
            }
        )
    return rows


def build_target_data_validation_rows(
    target_observations: list[dict[str, Any]],
    arable_target_capacity_rows: list[dict[str, Any]],
    wood_target_capacity_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in target_observations:
        rows.append(
            {
                "input_family": "target_observation",
                "source_file": "historical_anchors.csv" if row.get("observation_family") == "historical" else "modern_maxima.csv",
                "state": row["official_state"],
                "resource": row["resource"],
                "land_class": "",
                "year": row["year"],
                "evidence_scope": row.get("evidence_scope", ""),
                "target_match_status": row.get("target_match_status", ""),
                "slot_support_status": row.get("slot_support_status", ""),
                "localization_discount": row.get("localization_discount", ""),
                "validation_note": row.get("validation_note", ""),
                "drives_x": row.get("drives_x", "no"),
                "discounted_effective_value": row.get("discounted_normalized_1950_output", ""),
                "review_action": row.get("review_action", ""),
                "source_title": row.get("source_title", ""),
                "source_url": row.get("source_url", ""),
                "citation_locator": row.get("locator", ""),
            }
        )
    for source_file, resource_name, capacity_rows in [
        ("arable_target_capacity_rows.csv", "Arable Land", arable_target_capacity_rows),
        ("wood_target_capacity_rows.csv", "Wood", wood_target_capacity_rows),
    ]:
        for row in capacity_rows:
            validation = {field: row.get(field, "") for field in TARGET_VALIDATION_FIELDS}
            drives_x = validation_drives_x(validation, 1)
            rows.append(
                {
                    "input_family": "target_capacity",
                    "source_file": source_file,
                    "state": row["state"],
                    "resource": resource_name,
                    "land_class": row["land_class"],
                    "year": row["representative_year"],
                    "evidence_scope": validation["evidence_scope"],
                    "target_match_status": validation["target_match_status"],
                    "slot_support_status": validation["slot_support_status"],
                    "localization_discount": validation["localization_discount"],
                    "validation_note": validation["validation_note"],
                    "drives_x": "yes" if drives_x else "no",
                    "discounted_effective_value": row["effective_area_ha"] if drives_x else "",
                    "review_action": validation_review_action(validation, drives_x, 1),
                    "source_title": row.get("citation_1_title", ""),
                    "source_url": row.get("citation_1_url", ""),
                    "citation_locator": row.get("citation_1_locator", ""),
                }
            )
    rows.sort(key=lambda item: (item["state"], item["resource"], item["source_file"], item["year"], item["land_class"]))
    return rows


def humanize_label(value: Any) -> str:
    text = str(value or "").replace("_", " ").replace("-", " ").strip()
    return text[:1].upper() + text[1:] if text else ""


def binary_change_label(vanilla_value: str, sb_value: str) -> str:
    if vanilla_value == sb_value:
        return "No change"
    if vanilla_value == "yes" and sb_value == "no":
        return "Removed"
    if vanilla_value == "no" and sb_value == "yes":
        return "Added"
    return "Changed"


def build_overview_resource_rows(regional_total_rows: list[dict[str, Any]], region: str) -> list[dict[str, Any]]:
    resource_order = {resource: idx for idx, resource in enumerate(SUMMARY_RESOURCES)}
    rows = [
        {
            "resource": row["resource"],
            "vanilla_baseline": int(float(row["vanilla_total"])),
            "sb_update": int(float(row["final_sb_total"])),
            "delta_vs_vanilla": int(float(row["absolute_delta"])),
        }
        for row in regional_total_rows
        if row["region"] == region
    ]
    rows.sort(key=lambda row: resource_order[row["resource"]])
    return rows


def build_workbook_tag_rows(regional_total_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    resource_order = {resource: idx for idx, resource in enumerate(SUMMARY_RESOURCES)}
    tag_order = {tag: idx for idx, tag in enumerate(WORKBOOK_TAG_REGIONS)}
    rows = [
        {
            "tag": row["region"],
            "resource": row["resource"],
            "vanilla_baseline": int(float(row["vanilla_total"])),
            "sb_update": int(float(row["final_sb_total"])),
            "delta_vs_vanilla": int(float(row["absolute_delta"])),
        }
        for row in regional_total_rows
        if row["region"] in WORKBOOK_TAG_REGIONS
    ]
    rows.sort(key=lambda row: (tag_order[row["tag"]], resource_order[row["resource"]]))
    return rows


def build_state_sheet_rows(
    resource_adjustment_rows: list[dict[str, Any]],
    arable_resource_expectation_rows: list[dict[str, Any]],
    vanilla_priors: dict[str, dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    adjustment_map = {(row["state"], row["resource"]): row for row in resource_adjustment_rows}
    expectation_map = {(row["state"], row["resource"]): row for row in arable_resource_expectation_rows}
    rows_by_state: dict[str, list[dict[str, Any]]] = {}
    for info in STATE_INFO:
        state = info["official_name"]
        rows: list[dict[str, Any]] = []
        for category, resource in WORKBOOK_STATE_RESOURCE_ORDER:
            if category == "Arable Resource":
                expectation = expectation_map[(state, resource)]
                vanilla_value = str(vanilla_priors[state].get(resource, "no")).lower()
                sb_value = str(expectation["researched_plausible"]).lower()
                rows.append(
                    {
                        "category": humanize_label(category),
                        "resource": resource,
                        "vanilla_baseline": vanilla_value,
                        "sb_update": sb_value,
                        "delta_or_change": binary_change_label(vanilla_value, sb_value),
                    }
                )
                continue
            adjustment = adjustment_map[(state, resource)]
            vanilla_cap = int(float(adjustment["vanilla_cap"]))
            final_cap = int(float(adjustment["final_audited_cap"]))
            rows.append(
                {
                    "category": humanize_label(category),
                    "resource": resource,
                    "vanilla_baseline": vanilla_cap,
                    "sb_update": final_cap,
                    "delta_or_change": final_cap - vanilla_cap,
                }
            )
        rows_by_state[state] = rows
    return rows_by_state


def build_workbook(
    resource_adjustment_rows: list[dict[str, Any]],
    regional_total_rows: list[dict[str, Any]],
    arable_resource_expectation_rows: list[dict[str, Any]],
    vanilla_priors: dict[str, dict[str, Any]],
    state_review_status_rows: list[dict[str, Any]],
) -> Workbook:
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"

    state_summary_rows = build_state_summary_rows(resource_adjustment_rows)
    state_summary_map = {row["state"]: row for row in state_summary_rows}
    state_sheet_rows = build_state_sheet_rows(resource_adjustment_rows, arable_resource_expectation_rows, vanilla_priors)
    sb_scope_rows = build_overview_resource_rows(regional_total_rows, "SB Scope")
    tag_rows = build_workbook_tag_rows(regional_total_rows)

    overview.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    overview["A1"] = "SB scope - vanilla baseline and proposed SB update"
    overview["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    overview["A1"].fill = TITLE_FILL
    overview["A1"].alignment = Alignment(horizontal="left", vertical="center")
    overview.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    overview["A2"] = (
        "Vanilla baseline is compared against the audited SB update. "
        "Tags: CAP = Cape states, TRN = Transvaal states, SAF = South African states, SWA = Namibian states."
    )
    overview["A2"].alignment = Alignment(wrap_text=True)
    row = 4
    row = add_table(
        overview,
        row,
        "SB totals before and after",
        ["Resource", "Vanilla baseline", "SB update", "Delta vs vanilla"],
        sb_scope_rows,
        ["resource", "vanilla_baseline", "sb_update", "delta_vs_vanilla"],
    )
    add_table(
        overview,
        row,
        "Major tag changes",
        ["Tag", "Resource", "Vanilla baseline", "SB update", "Delta vs vanilla"],
        tag_rows,
        ["tag", "resource", "vanilla_baseline", "sb_update", "delta_vs_vanilla"],
    )
    overview.freeze_panes = "A6"

    for info in STATE_INFO:
        state = info["official_name"]
        ws = wb.create_sheet(state)
        summary = state_summary_map[state]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws["A1"] = f"{state} - vanilla baseline and proposed SB update"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = TITLE_FILL
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        ws["A2"] = (
            f"SB target mapping: {state} -> {info['state_id']} | "
            f"Vanilla proxy: {info['vanilla_proxy_name']} -> {info['vanilla_proxy_id']}"
        )
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
        ws["A3"] = "Final column shows the proposed SB update. Arable-resource rows are yes/no gameplay availability, while capped rows are numeric caps."
        ws["A3"].alignment = Alignment(wrap_text=True)
        row = 5
        row = add_key_value_block(
            ws,
            row,
            "Totals",
            [
                ("Vanilla total", summary["vanilla_total"]),
                ("SB update total", summary["final_audited_total"]),
                ("Delta vs vanilla", summary["delta_vs_vanilla"]),
            ],
        )
        table_start = row
        add_table(
            ws,
            table_start,
            "Resource rows",
            ["Category", "Resource", "Vanilla baseline", "SB update", "Delta / change"],
            state_sheet_rows[state],
            ["category", "resource", "vanilla_baseline", "sb_update", "delta_or_change"],
        )
        ws.freeze_panes = f"A{table_start + 2}"

    for ws in wb.worksheets:
        autofit(ws)
    return wb


def rewrite_state_block(block: str, final_caps_for_state: dict[str, int]) -> str:
    arable = final_caps_for_state["Arable Land"]
    block = re.sub(r"(?m)^(\s*arable_land\s*=\s*)\d+", rf"\g<1>{arable}", block)
    capped_lines = [f"        {building} = {final_caps_for_state[resource]}" for resource, building in CAP_BUILDINGS.items() if final_caps_for_state.get(resource, 0)]
    new_capped_block = "    capped_resources = {\n" + ("\n".join(capped_lines) + "\n" if capped_lines else "") + "    }"
    if re.search(r"\n\s*capped_resources\s*=\s*\{.*?\n\s*\}", block, re.S):
        block = re.sub(r"\n\s*capped_resources\s*=\s*\{.*?\n\s*\}", "\n" + new_capped_block, block, flags=re.S)
    else:
        block = re.sub(r"\n\}", "\n" + new_capped_block + "\n}", block, count=1)

    def keep_resource_block(match: re.Match[str]) -> str:
        resource_block = match.group(0)
        type_match = re.search(r'type\s*=\s*"([^"]+)"', resource_block)
        if type_match and type_match.group(1) in {config["type"] for config in SPECIAL_RESOURCE_CONFIG.values()}:
            return ""
        return resource_block

    block = re.sub(r"\n\s*resource\s*=\s*\{.*?\n\s*\}", keep_resource_block, block, flags=re.S)
    special_blocks: list[str] = []
    for family, config in SPECIAL_RESOURCE_CONFIG.items():
        discovered = final_caps_for_state.get(f"{family} (discovered)", 0)
        undiscovered = final_caps_for_state.get(f"{family} (undiscovered)", 0)
        if not discovered and not undiscovered:
            continue
        lines = ["    resource = {", f'        type = "{config["type"]}"']
        if config.get("depleted_type"):
            lines.append(f'        depleted_type = "{config["depleted_type"]}"')
        if undiscovered:
            lines.append(f"        undiscovered_amount = {undiscovered}")
        if discovered:
            lines.append(f"        discovered_amount = {discovered}")
        lines.append("    }")
        special_blocks.append("\n".join(lines))
    if special_blocks:
        block = re.sub(r"\n\}", "\n" + "\n".join(special_blocks) + "\n}", block, count=1)
    block = re.sub(r"\n{3,}", "\n\n", block)
    return block


def sync_live_state_file(final_caps: dict[str, dict[str, int]]) -> None:
    text = STATE_FILE.read_text(encoding="utf-8")
    for info in STATE_INFO:
        old_block = extract_state_block(text, info["state_id"])
        new_block = rewrite_state_block(old_block, final_caps[info["official_name"]])
        text = text.replace(old_block, new_block)
    STATE_FILE.write_text(text, encoding="utf-8")


def build_public_workbook() -> dict[str, Any]:
    growth_rows, growth_index_lookup = build_growth_rows(load_growth_anchor_series())
    historical_rows, modern_rows = load_raw_data()
    gdp_reference = load_gdp_reference_anchor()
    gdp_series = load_gdp_series()
    gdp_map = load_gdp_geography_map()
    target_observations = enrich_observation_rows(historical_rows + modern_rows, include_target=True, growth_index_lookup=growth_index_lookup)
    comparator_observations = enrich_observation_rows(historical_rows + modern_rows, include_target=False, growth_index_lookup=growth_index_lookup)
    selected_target_rows, target_gdp_selection_rows = select_observation_rows(target_observations, "target_state", gdp_reference, gdp_series, gdp_map)
    selected_comparator_rows, comparator_gdp_selection_rows = select_observation_rows(comparator_observations, "agri_comparator", gdp_reference, gdp_series, gdp_map)
    target_selection_map = {
        (entity, resource): selection
        for selection in target_gdp_selection_rows
        for entity, resource in [(selection["entity"], selection["resource"])]
    }
    counterevidence_cases = load_counterevidence_cases()
    family_normalising_constants = load_family_normalising_constants()
    live_values = parse_live_state_resources()
    vanilla_priors = load_vanilla_priors()
    arable_baskets = load_arable_baskets()
    _arable_land_weights = load_arable_land_class_weights()
    _wood_land_weights = load_wood_land_class_weights()
    vanilla_totals = load_adjusted_vanilla_totals(vanilla_priors)
    target_normalized, target_wheat_eq = build_target_observed_maps(selected_target_rows)
    comparator_maxima, comparator_totals = build_comparator_maxima(selected_comparator_rows)
    rankings = parse_agricultural_rankings()
    arable_target_capacity_rows = load_arable_target_capacity_rows()
    arable_comparator_capacity_rows = load_arable_comparator_capacity_rows(rankings)
    wood_target_capacity_rows = load_wood_target_capacity_rows()
    wood_comparator_capacity_rows = load_wood_comparator_capacity_rows()
    target_data_validation_rows = build_target_data_validation_rows(
        target_observations,
        arable_target_capacity_rows,
        wood_target_capacity_rows,
    )
    state_regional_advantage_rows = seed_regional_advantages_rows()
    state_review_status_rows = build_state_review_status_rows()
    mining_comparator_rows, mining_gdp_selection_rows = compute_mining_comparator_rows(load_benchmark_cases(), growth_index_lookup, gdp_reference, gdp_series, gdp_map)
    wood_state_payloads = compute_wood_state_payloads(wood_target_capacity_rows)
    wood_denominator_row, wood_comparator_summary_rows = compute_wood_denominator(wood_comparator_capacity_rows)
    resource_denominators, resource_denominator_rows = compute_resource_denominators(mining_comparator_rows, wood_denominator_row)
    arable_state_payloads, arable_state_mean_rows, arable_comparator_diagnostics = compute_arable_state_means(
        rankings,
        arable_target_capacity_rows,
        arable_comparator_capacity_rows,
        arable_baskets,
        live_values,
    )
    arable_shared_denominator = compute_shared_arable_denominator(arable_state_payloads)
    arable_resource_expectation_rows = build_arable_resource_expectations(live_values, arable_baskets)
    arable_coverage_payloads = arable_state_payloads
    base_rows = build_base_state_resource_rows(
        live_values,
        vanilla_priors,
        arable_baskets,
        selected_target_rows,
        target_selection_map,
        resource_denominators,
        arable_shared_denominator,
        arable_coverage_payloads,
        wood_state_payloads,
        gdp_reference,
        gdp_series,
        gdp_map,
    )
    seed_adjustment_inputs_if_needed(base_rows, arable_state_payloads)
    adjustment_inputs = load_adjustment_inputs()
    resource_adjustment_rows, final_caps, audit_rows = apply_adjustments(base_rows, adjustment_inputs, counterevidence_cases)
    regional_total_rows = build_regional_totals(vanilla_totals, final_caps)
    priority_rows = compute_priority_rows(resource_adjustment_rows)
    row_audit_rows = build_row_audit_rows(resource_adjustment_rows, build_selected_target_map(selected_target_rows), counterevidence_cases, family_normalising_constants)
    capacity_model_resources = {resource for category, resource in BINARY_RESOURCES if category == "Arable Resource"} | {"Arable Land", "Wood"}
    target_gdp_selection_rows = [row for row in target_gdp_selection_rows if row["resource"] not in capacity_model_resources]
    comparator_gdp_selection_rows = [row for row in comparator_gdp_selection_rows if row["resource"] not in capacity_model_resources]
    gdp_selection_rows = (
        target_gdp_selection_rows
        + comparator_gdp_selection_rows
        + mining_gdp_selection_rows
    )
    gdp_selection_rows = [row for row in gdp_selection_rows if row["resource"] != "Wood"]
    public_target_observations = [row for row in target_observations if row["resource"] != "Wood"]
    public_comparator_observations = [row for row in comparator_observations if row["resource"] != "Wood"]
    workbook = build_workbook(
        resource_adjustment_rows,
        regional_total_rows,
        arable_resource_expectation_rows,
        vanilla_priors,
        state_review_status_rows,
    )
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    write_public_outputs(
        growth_rows,
        public_target_observations,
        public_comparator_observations,
        target_data_validation_rows,
        gdp_selection_rows,
        family_normalising_constants,
        resource_denominator_rows,
        arable_target_capacity_rows,
        arable_comparator_capacity_rows,
        arable_state_mean_rows,
        arable_shared_denominator,
        arable_resource_expectation_rows,
        arable_comparator_diagnostics,
        wood_target_capacity_rows,
        wood_comparator_capacity_rows,
        resource_adjustment_rows,
        regional_total_rows,
        audit_rows,
        row_audit_rows,
        priority_rows,
        state_regional_advantage_rows,
        state_review_status_rows,
    )
    return {
        "growth_rows": growth_rows,
        "target_observations": public_target_observations,
        "comparator_observations": public_comparator_observations,
        "target_data_validation_rows": target_data_validation_rows,
        "family_normalising_constants": family_normalising_constants,
        "resource_denominator_rows": resource_denominator_rows,
        "arable_target_capacity_rows": arable_target_capacity_rows,
        "arable_comparator_capacity_rows": arable_comparator_capacity_rows,
        "arable_state_mean_rows": arable_state_mean_rows,
        "arable_shared_denominator": arable_shared_denominator,
        "arable_resource_expectation_rows": arable_resource_expectation_rows,
        "arable_comparator_diagnostics": arable_comparator_diagnostics,
        "wood_target_capacity_rows": wood_target_capacity_rows,
        "wood_comparator_capacity_rows": wood_comparator_capacity_rows,
        "wood_comparator_summary_rows": wood_comparator_summary_rows,
        "gdp_selection_rows": gdp_selection_rows,
        "resource_adjustment_rows": resource_adjustment_rows,
        "regional_total_rows": regional_total_rows,
        "audit_rows": audit_rows,
        "row_audit_rows": row_audit_rows,
        "priority_rows": priority_rows,
        "state_regional_advantage_rows": state_regional_advantage_rows,
        "state_review_status_rows": state_review_status_rows,
        "final_caps": final_caps,
        "sync_live": SYNC_LIVE_ON_BUILD,
    }


if __name__ == "__main__":
    result = build_public_workbook()
    print(f"Wrote {OUTPUT}")
    print(f"Wrote {DERIVED_DIR}")
    print(f"Wrote {AUDIT_DIR}")
    print(f"Tracked {len(result['resource_adjustment_rows'])} state-resource rows")
